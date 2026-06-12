"""
tests/test_parser.py — parser.py 單元測試

測試策略：不需要真實 Excel 檔，直接用 openpyxl 在記憶體中建立 Workbook。
"""
import pytest
import openpyxl
from openpyxl.utils import get_column_letter

from core.parser import _norm, _cell_str, _scan_header, _find_item_header_row, _parse_items, parse


# ── _norm ─────────────────────────────────────────────────────────────────────

def test_norm_removes_spaces():
    assert _norm("報 價 單 號") == "報價單號"

def test_norm_removes_fullwidth_space():
    assert _norm("報　價　單") == "報價單"

def test_norm_empty():
    assert _norm("") == ""

def test_norm_none():
    assert _norm(None) == ""


# ── _cell_str ─────────────────────────────────────────────────────────────────

def test_cell_str_text():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "  ABC  "
    assert _cell_str(ws["A1"]) == "ABC"

def test_cell_str_none():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = None
    assert _cell_str(ws["A1"]) == ""

def test_cell_str_number():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 123
    assert _cell_str(ws["A1"]) == "123"

def test_cell_str_datetime():
    from datetime import datetime
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = datetime(2026, 1, 15)
    assert _cell_str(ws["A1"]) == "2026/01/15"

def test_cell_str_none_cell():
    assert _cell_str(None) == ""


# ── _scan_header ──────────────────────────────────────────────────────────────

def _make_header_ws():
    """建立含標準表頭的測試工作表。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    # 左側標籤：標籤在 col A，值在 col B
    ws["A1"] = "報價單號"
    ws["B1"] = "QF-2026-001"
    ws["A2"] = "客戶全名"
    ws["B2"] = "立善科技股份有限公司"
    ws["A3"] = "電話"
    ws["B3"] = "02-1234-5678"
    ws["A4"] = "聯絡人"
    ws["B4"] = "張三"
    # 右側標籤：標籤在某欄，值在其右方
    ws["D1"] = "報價日期"
    ws["E1"] = "2026/06/01"
    ws["D2"] = "統一編號"
    ws["E2"] = "12345678"
    return ws


def test_scan_header_basic():
    ws = _make_header_ws()
    result = _scan_header(ws)
    assert result["quote_no"] == "QF-2026-001"
    assert result["customer"] == "立善科技股份有限公司"
    assert result["phone"] == "02-1234-5678"
    assert result["contact"] == "張三"

def test_scan_header_right_labels():
    ws = _make_header_ws()
    result = _scan_header(ws)
    assert result["quote_date"] == "2026/06/01"
    assert result["tax_id"] == "12345678"

def test_scan_header_missing_fields_empty():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "其他內容"
    result = _scan_header(ws)
    assert result["quote_no"] == ""
    assert result["customer"] == ""


# ── _find_item_header_row ─────────────────────────────────────────────────────

def test_find_item_header_row_found():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A5"] = "序"
    ws["B5"] = "品名"
    assert _find_item_header_row(ws) == 5

def test_find_item_header_row_not_found():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "無序號欄位"
    assert _find_item_header_row(ws) == -1


# ── _parse_items ──────────────────────────────────────────────────────────────

def _make_items_ws(header_row: int = 5):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(header_row, 1, "序")
    # 品項 1
    r = header_row + 1
    ws.cell(r, 1, "1")
    ws.cell(r, 2, "ITEM-A")
    ws.cell(r, 3, "產品甲")
    ws.cell(r, 6, 2)
    ws.cell(r, 7, "台")
    ws.cell(r, 8, 1000)
    ws.cell(r, 9, 2000)
    # 品項 2
    r2 = header_row + 2
    ws.cell(r2, 1, "2")
    ws.cell(r2, 2, "ITEM-B")
    ws.cell(r2, 3, "產品乙")
    ws.cell(r2, 6, 3)
    ws.cell(r2, 7, "個")
    ws.cell(r2, 8, 500)
    ws.cell(r2, 9, 1500)
    return ws, header_row


def test_parse_items_count():
    ws, hrow = _make_items_ws()
    items = _parse_items(ws, hrow)
    assert len(items) == 2

def test_parse_items_values():
    ws, hrow = _make_items_ws()
    items = _parse_items(ws, hrow)
    assert items[0]["seq"] == 1
    assert items[0]["qty"] == 2
    assert items[0]["unit"] == "台"
    assert items[0]["unit_price"] == 1000
    assert items[1]["seq"] == 2
    assert items[1]["unit_price"] == 500

def test_parse_items_empty_sheet():
    wb = openpyxl.Workbook()
    ws = wb.active
    items = _parse_items(ws, 1)
    assert items == []
