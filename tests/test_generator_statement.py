"""
tests/test_generator_statement.py — generator_statement.py 工具函式測試
"""
from core.generator_statement import (
    generate_statement, parse_amount, parse_paid_amount, split_product_quantity,
)


# ── parse_amount ─────────────────────────────────────────────────────────────

def test_parse_amount_plain_number():
    assert parse_amount("12000") == 12000.0

def test_parse_amount_with_currency_and_commas():
    assert parse_amount("NT$12,000") == 12000.0

def test_parse_amount_empty():
    assert parse_amount("") == 0.0

def test_parse_amount_non_numeric():
    assert parse_amount("未確認") == 0.0


# ── parse_paid_amount ─────────────────────────────────────────────────────────

def test_parse_paid_amount_with_deposit():
    assert parse_paid_amount("匯款，已付訂金5000") == 5000.0

def test_parse_paid_amount_with_prefix_label():
    assert parse_paid_amount("現金，已收3000元") == 3000.0

def test_parse_paid_amount_no_hint():
    assert parse_paid_amount("匯款") == 0.0

def test_parse_paid_amount_empty():
    assert parse_paid_amount("") == 0.0

def test_parse_paid_amount_percentage():
    assert parse_paid_amount("訂金50%(匯款)、尾款50%(月結30天)", 10000) == 5000.0

def test_parse_paid_amount_percentage_overrides_absolute():
    # 同時出現百分比與絕對數字時，優先採用百分比寫法
    assert parse_paid_amount("訂金30%已付", 20000) == 6000.0

def test_parse_paid_amount_percentage_without_total_is_zero():
    assert parse_paid_amount("訂金50%", 0) == 0.0


# ── split_product_quantity ──────────────────────────────────────────────────

def test_split_product_quantity_with_marker():
    assert split_product_quantity("電池*5", "1") == ("電池", "5")

def test_split_product_quantity_marker_overrides_fallback():
    name, qty = split_product_quantity("升降平台*2(訂製)", "1")
    assert qty == "2"
    assert name == "升降平台(訂製)"

def test_split_product_quantity_without_marker():
    assert split_product_quantity("電池", "3") == ("電池", "3")

def test_split_product_quantity_empty():
    assert split_product_quantity("", "1") == ("", "1")


# ── generate_statement：gemini_paid_amount 覆寫 ──────────────────────────────

def test_generate_statement_gemini_override(tmp_path):
    import openpyxl
    card = {
        "company": "測試公司", "product": "電池", "quantity": "1",
        "amount": "10000", "payment_raw": "訂金30%",
    }
    out = generate_statement(card, output_dir=tmp_path, gemini_paid_amount=8000.0)
    ws = openpyxl.load_workbook(out).active
    assert ws.cell(row=12, column=6).value == 8000.0   # 已收金額 採用 Gemini 結果
    assert ws.cell(row=12, column=7).value == 2000.0   # 未收金額 = 總金額 - Gemini 已收

def test_generate_statement_without_gemini_uses_regex(tmp_path):
    import openpyxl
    card = {
        "company": "測試公司", "product": "電池", "quantity": "1",
        "amount": "10000", "payment_raw": "訂金30%",
    }
    out = generate_statement(card, output_dir=tmp_path)
    ws = openpyxl.load_workbook(out).active
    assert ws.cell(row=12, column=6).value == 3000.0   # 沒有 Gemini 結果時回退 regex（30%）
