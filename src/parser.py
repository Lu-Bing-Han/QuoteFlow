"""
parser.py — 關鍵字搜尋版本
"""

import re
import openpyxl
from openpyxl.cell import MergedCell
from datetime import datetime


def _norm(s):
    return re.sub(r'[\s\u3000]+', '', str(s or ''))


def _cell_str(cell) -> str:
    if cell is None or isinstance(cell, MergedCell):
        return ""
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y/%m/%d")
    return str(v).strip()


LEFT_LABELS = {
    "報價單號":  "quote_no",
    "客戶全名":  "customer",
    "電話":      "phone",
    "聯絡人":    "contact",
    "聯絡地址":  "address",
}

RIGHT_LABELS = {
    "報價日期":  "quote_date",
    "有效日期":  "valid_date",   # ← 新增
    "有效期限":  "valid_date",   # ← 新增
    "幣別":      "currency",
    "統一編號":  "tax_id",
    "EMAIL":     "email",
}


def _scan_header(ws) -> dict:
    result = {k: "" for k in list(LEFT_LABELS.values()) + list(RIGHT_LABELS.values())}

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell) or cell.value is None:
                continue
            norm = _norm(cell.value)

            for kw, field in LEFT_LABELS.items():
                if kw in norm and not result[field]:
                    val_cell = ws.cell(row=cell.row, column=2)
                    result[field] = _cell_str(val_cell)

            for kw, field in RIGHT_LABELS.items():
                if kw in norm and not result[field]:
                    for c in range(cell.column + 1, cell.column + 5):
                        v = _cell_str(ws.cell(row=cell.row, column=c))
                        if v:
                            result[field] = v
                            break

    return result


def _find_item_header_row(ws) -> int:
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell, MergedCell) and _norm(cell.value) in ("序", "序號"):
                return cell.row
    return -1


def _parse_items(ws, header_row: int) -> list:
    items = []
    current = None

    for row_idx in range(header_row + 1, header_row + 200):
        a = _cell_str(ws.cell(row=row_idx, column=1))

        if "應收總金額" in a or "合計" in a and row_idx > header_row + 2:
            break

        if a.isdigit():
            if current:
                items.append(current)

            name_parts = []
            for col in range(2, 6):
                v = _cell_str(ws.cell(row=row_idx, column=col))
                if v:
                    name_parts.append(v)
            name = " ".join(name_parts)

            current = {
                "seq":        int(a),
                "name":       name,
                "qty":        ws.cell(row=row_idx, column=6).value or 0,
                "unit":       _cell_str(ws.cell(row=row_idx, column=7)),
                "unit_price": ws.cell(row=row_idx, column=8).value or 0,
                "subtotal":   ws.cell(row=row_idx, column=9).value or 0,
            }

        elif current:
            for col in range(3, 6):
                v = _cell_str(ws.cell(row=row_idx, column=col))
                if v:
                    current["name"] += "\n" + v
                    break

    if current:
        items.append(current)

    return items


def parse(filepath: str) -> dict:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    header = _scan_header(ws)
    item_header_row = _find_item_header_row(ws)
    items = _parse_items(ws, item_header_row) if item_header_row > 0 else []

    return {"header": header, "items": items}


if __name__ == "__main__":
    import sys, json
    path = sys.argv[1] if len(sys.argv) > 1 else "報價單.xlsx"
    result = parse(path)
    print(json.dumps(result, ensure_ascii=False, indent=2))