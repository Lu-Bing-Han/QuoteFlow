"""
mixin_schedule.py — 出貨排程頁籤 mixin
"""
import os
import tkinter as tk
from tkinter import messagebox, ttk
import customtkinter as ctk
from pathlib import Path
from ui.app_core import _mk_lf


class _ScheduleTab:
    """Mixin providing _build_tab_schedule and its callbacks."""

    def _build_tab_schedule(self, parent, FONT, FONTB, BG):
        from tkcalendar import DateEntry
        from _paths import TEMPLATE_DIR
        GRAY   = "#5d6d7e"
        FONT_S = ("Microsoft JhengHei UI", 9)

        _rows = []

        _addr_options: list[str] = []
        try:
            import openpyxl as _oxl
            _tmpl_path = TEMPLATE_DIR / "template_schedule.xlsx"
            if _tmpl_path.exists():
                _wb = _oxl.load_workbook(str(_tmpl_path), read_only=True, data_only=True)
                if "地址" in _wb.sheetnames:
                    _ws = _wb["地址"]
                    _addr_options = [
                        str(_ws.cell(row=r, column=3).value).strip()
                        for r in range(2, 30)
                        if _ws.cell(row=r, column=3).value
                    ]
                _wb.close()
        except Exception:
            pass

        # ── Credential section ────────────────────────────
        cred_outer, cred_frame = _mk_lf(parent, "Timetree 登入憑證", BG, FONTB)
        cred_outer.pack(fill="x", padx=12, pady=(12, 4))
        cred_frame.columnconfigure(1, weight=1)

        tt_cfg = self._config.get("timetree", {})
        sid_var  = tk.StringVar(value=tt_cfg.get("session_id", ""))
        csrf_var = tk.StringVar(value=tt_cfg.get("csrf_token", ""))

        ctk.CTkLabel(cred_frame, text="Session ID：", fg_color="transparent",
                      font=FONT_S, text_color=GRAY,
                      anchor="w").grid(row=0, column=0, sticky="w", padx=8, pady=3)
        sid_entry = ctk.CTkEntry(cred_frame, textvariable=sid_var, font=FONT_S,
                                  show="*", corner_radius=4, border_width=1)
        sid_entry.grid(row=0, column=1, sticky="ew", padx=8, pady=3)

        ctk.CTkLabel(cred_frame, text="CSRF Token：", fg_color="transparent",
                      font=FONT_S, text_color=GRAY,
                      anchor="w").grid(row=1, column=0, sticky="w", padx=8, pady=3)
        csrf_entry = ctk.CTkEntry(cred_frame, textvariable=csrf_var, font=FONT_S,
                                   show="*", corner_radius=4, border_width=1)
        csrf_entry.grid(row=1, column=1, sticky="ew", padx=8, pady=3)

        def _show_hide(entry, btn):
            if entry.cget("show") == "*":
                entry.configure(show="")
                btn.configure(text="隱藏")
            else:
                entry.configure(show="*")
                btn.configure(text="顯示")

        btn_show_sid  = ctk.CTkButton(cred_frame, text="顯示", font=FONT_S,
                                       fg_color=GRAY, hover_color="#4d5d6e",
                                       text_color="white", width=50, height=26,
                                       corner_radius=4)
        btn_show_sid.configure(command=lambda: _show_hide(sid_entry, btn_show_sid))
        btn_show_sid.grid(row=0, column=2, padx=(0, 8), pady=3)

        btn_show_csrf = ctk.CTkButton(cred_frame, text="顯示", font=FONT_S,
                                       fg_color=GRAY, hover_color="#4d5d6e",
                                       text_color="white", width=50, height=26,
                                       corner_radius=4)
        btn_show_csrf.configure(command=lambda: _show_hide(csrf_entry, btn_show_csrf))
        btn_show_csrf.grid(row=1, column=2, padx=(0, 8), pady=3)

        def _save_creds():
            self._config.setdefault("timetree", {})
            self._config["timetree"]["session_id"] = sid_var.get().strip()
            self._config["timetree"]["csrf_token"]  = csrf_var.get().strip()
            self._save_config(self._config)
            messagebox.showinfo("已儲存", "Timetree 憑證已儲存", parent=parent)

        ctk.CTkButton(cred_frame, text="儲存憑證", command=_save_creds,
                       fg_color="#2e86c1", hover_color="#1a5276", text_color="white",
                       font=FONT, width=90, height=28, corner_radius=4
                       ).grid(row=2, column=1, sticky="w", padx=8, pady=(4, 6))

        # ── Google Maps 設定 ──────────────────────────────
        maps_outer, maps_frame = _mk_lf(parent, "Google Maps（行車時間）", BG, FONTB)
        maps_outer.pack(fill="x", padx=12, pady=(0, 4))
        maps_frame.columnconfigure(1, weight=1)

        gm_cfg     = self._config.get("google_maps", {})
        gm_key_var = tk.StringVar(value=gm_cfg.get("api_key", ""))
        gm_org_var = tk.StringVar(value=gm_cfg.get("origin",  "406臺中市北屯區水景里景南巷1-1號"))

        ctk.CTkLabel(maps_frame, text="API Key：", fg_color="transparent",
                      font=FONT_S, text_color=GRAY,
                      anchor="w").grid(row=0, column=0, sticky="w", padx=8, pady=3)
        gm_key_entry = ctk.CTkEntry(maps_frame, textvariable=gm_key_var, font=FONT_S,
                                     show="*", corner_radius=4, border_width=1)
        gm_key_entry.grid(row=0, column=1, sticky="ew", padx=8, pady=3)
        btn_gm_key = ctk.CTkButton(maps_frame, text="顯示", font=FONT_S,
                                    fg_color=GRAY, hover_color="#4d5d6e",
                                    text_color="white", width=50, height=26, corner_radius=4)
        btn_gm_key.configure(command=lambda: _show_hide(gm_key_entry, btn_gm_key))
        btn_gm_key.grid(row=0, column=2, padx=(0, 8), pady=3)

        ctk.CTkLabel(maps_frame, text="出發地：", fg_color="transparent",
                      font=FONT_S, text_color=GRAY,
                      anchor="w").grid(row=1, column=0, sticky="w", padx=8, pady=3)
        ctk.CTkEntry(maps_frame, textvariable=gm_org_var, font=FONT_S,
                      corner_radius=4, border_width=1
                      ).grid(row=1, column=1, columnspan=2, sticky="ew", padx=8, pady=3)

        def _save_maps_cfg():
            self._config.setdefault("google_maps", {})
            self._config["google_maps"]["api_key"] = gm_key_var.get().strip()
            self._config["google_maps"]["origin"]  = gm_org_var.get().strip()
            self._save_config(self._config)
            messagebox.showinfo("已儲存", "Google Maps 設定已儲存", parent=parent)

        ctk.CTkButton(maps_frame, text="儲存設定", command=_save_maps_cfg,
                       fg_color="#2e86c1", hover_color="#1a5276", text_color="white",
                       font=FONT, width=90, height=28, corner_radius=4
                       ).grid(row=2, column=1, sticky="w", padx=8, pady=(4, 6))

        # ── Preview section ───────────────────────────────
        prev_outer, prev_frame = _mk_lf(parent, "排程預覽", BG, FONTB)
        prev_outer.pack(fill="both", expand=True, padx=12, pady=4)

        date_row = ctk.CTkFrame(prev_frame, fg_color="transparent", corner_radius=0)
        date_row.pack(fill="x", padx=8, pady=(6, 4))
        ctk.CTkLabel(date_row, text="日期：", fg_color="transparent",
                      font=FONT, text_color="#2c3e50").pack(side="left")
        date_entry = DateEntry(date_row, font=FONT, date_pattern="yyyy/mm/dd",
                               background="#2e86c1", foreground="white", width=14)
        date_entry.pack(side="left", padx=(0, 6))

        fetch_status = ctk.CTkLabel(date_row, text="", fg_color="transparent",
                                     font=FONT_S, text_color=GRAY)
        fetch_status.pack(side="left", padx=8)

        tree_frame = ctk.CTkFrame(prev_frame, fg_color="transparent", corner_radius=0)
        tree_frame.pack(fill="both", expand=True, padx=8, pady=(0, 2))

        cols = ("seq", "location", "note", "travel_time")
        tree = ttk.Treeview(tree_frame, columns=cols, show="headings",
                             height=7, selectmode="browse")
        tree.heading("seq",         text="順序")
        tree.heading("location",    text="地點")
        tree.heading("note",        text="備註")
        tree.heading("travel_time", text="行車時間")
        tree.column("seq",         width=45,  anchor="center", stretch=False)
        tree.column("location",    width=160, anchor="w",      stretch=False)
        tree.column("note",        width=200, anchor="w",      stretch=True)
        tree.column("travel_time", width=75,  anchor="center", stretch=False)
        tree.pack(side="left", fill="both", expand=True)

        sb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        sb.pack(side="right", fill="y")
        tree.configure(yscrollcommand=sb.set)

        def _display_loc(location: str) -> str:
            idx = location.find("(")
            return location[:idx].strip() if idx != -1 else location

        def _refresh_tree():
            tree.delete(*tree.get_children())
            for i, row in enumerate(_rows, 1):
                tree.insert("", "end", values=(
                    i, _display_loc(row["location"]), row["note_suffix"], row.get("travel_time", "")))

        btn_row = ctk.CTkFrame(prev_frame, fg_color="transparent", corner_radius=0)
        btn_row.pack(fill="x", padx=8, pady=(2, 8))

        def _move(delta):
            sel = tree.selection()
            if not sel: return
            idx = tree.index(sel[0])
            new_idx = idx + delta
            if 0 <= new_idx < len(_rows):
                _rows[idx], _rows[new_idx] = _rows[new_idx], _rows[idx]
                _refresh_tree()
                tree.selection_set(tree.get_children()[new_idx])

        def _delete_row():
            sel = tree.selection()
            if not sel: return
            _rows.pop(tree.index(sel[0]))
            _refresh_tree()

        _addr_map = {}
        for _full in _addr_options:
            _idx = _full.find("(")
            _disp = _full[:_idx].strip() if _idx != -1 else _full
            _addr_map[_disp] = _full
        _addr_display = list(_addr_map.keys())

        def _open_row_dialog(title, location="", note_suffix="", on_confirm=None):
            dlg = ctk.CTkToplevel(parent)
            dlg.title(title)
            dlg.resizable(False, False)
            dlg.after(100, dlg.grab_set)

            ctk.CTkLabel(dlg, text="地點：", fg_color="transparent",
                          font=FONT, text_color="#2c3e50"
                          ).grid(row=0, column=0, padx=10, pady=6, sticky="w")
            loc_var = tk.StringVar(value=_display_loc(location))
            cb = ttk.Combobox(dlg, textvariable=loc_var, font=FONT,
                               width=26, values=_addr_display)
            cb.grid(row=0, column=1, padx=10, pady=6)

            ctk.CTkLabel(dlg, text="備註：", fg_color="transparent",
                          font=FONT, text_color="#2c3e50"
                          ).grid(row=1, column=0, padx=10, pady=6, sticky="w")
            note_var = tk.StringVar(value=note_suffix)
            ctk.CTkEntry(dlg, textvariable=note_var, font=FONT,
                          width=224, height=28, corner_radius=4
                          ).grid(row=1, column=1, padx=10, pady=6)

            def _confirm():
                typed = loc_var.get().strip()
                full_loc = _addr_map.get(typed, typed)
                if on_confirm:
                    on_confirm(full_loc, note_var.get().strip())
                dlg.destroy()

            ctk.CTkButton(dlg, text="確認", command=_confirm,
                           fg_color="#1a5276", hover_color="#154360", text_color="white",
                           font=FONT, width=80, height=30, corner_radius=4
                           ).grid(row=2, column=1, sticky="e", padx=10, pady=8)

        def _edit_row(_event=None):
            sel = tree.selection()
            if not sel: return
            idx = tree.index(sel[0])
            row = _rows[idx]

            def _apply(loc, note):
                _rows[idx]["location"]    = loc
                _rows[idx]["note_suffix"] = note
                _refresh_tree()

            _open_row_dialog("編輯事件", row["location"], row["note_suffix"], _apply)

        def _add_row():
            def _apply(loc, note):
                if not loc: return
                _rows.append({"ev": {}, "location": loc, "note_suffix": note, "travel_time": ""})
                _refresh_tree()
                tree.selection_set(tree.get_children()[-1])

            _open_row_dialog("新增事件", on_confirm=_apply)

        tree.bind("<Double-1>", _edit_row)

        def _calc_travel():
            from core.generator_schedule import calculate_travel_times
            if not _rows:
                messagebox.showwarning("無資料", "請先抓取事件清單", parent=parent); return
            api_key = gm_key_var.get().strip()
            origin  = gm_org_var.get().strip()
            if not api_key:
                messagebox.showwarning("未設定", "請先填入 Google Maps API Key", parent=parent); return

            def _worker():
                return calculate_travel_times(_rows, api_key, origin)

            def _on_done(result):
                _, failed = result
                _refresh_tree()
                if failed:
                    fetch_status.configure(text_color="#e67e22")
                    detail = "\n".join(
                        f"  第{seq}站「{loc}」— {status}" for seq, loc, status in failed)
                    messagebox.showwarning("部分地址無法計算",
                        f"以下站點無法取得行車時間：\n{detail}\n\n"
                        "請在 Timetree 的「地點」欄位填入完整地址，或手動在備註中修改。",
                        parent=parent)

            self._run_task(_worker,
                            buttons=[_btn_refs["📍 計算時間"]],
                            status_label=fetch_status,
                            loading_text="計算行車時間中…",
                            success_text=lambda result: (
                                "行車時間計算完成" if not result[1]
                                else f"完成，{len(result[1])} 筆失敗（地址找不到）"),
                            on_success=_on_done)

        def _sort_location(south_to_north: bool):
            from core.generator_schedule import sort_rows_by_location
            if not _rows:
                messagebox.showwarning("無資料", "請先抓取事件清單", parent=parent); return
            api_key = gm_key_var.get().strip()
            if not api_key:
                messagebox.showwarning("未設定", "請先填入 Google Maps API Key", parent=parent); return

            def _worker():
                return sort_rows_by_location(_rows, api_key, south_to_north)

            def _on_done(result):
                sorted_rows, failed = result
                _rows.clear()
                _rows.extend(sorted_rows)
                _refresh_tree()
                if failed:
                    fetch_status.configure(text_color="#e67e22")

            btn_key = "🧭 南→北" if south_to_north else "🧭 北→南"
            self._run_task(_worker,
                            buttons=[_btn_refs[btn_key]],
                            status_label=fetch_status,
                            loading_text="Geocoding 中…",
                            success_text=lambda result: (
                                "排序完成" if not result[1] else
                                f"排序完成，{len(result[1])} 筆無法定位：" +
                                ", ".join(r["location"].split("(")[0] for r in result[1])),
                            on_success=_on_done)

        _BTN_DEFS = [
            ("↑ 上移",      lambda: _move(-1),            "#5d6d7e", "#4d5d6e"),
            ("↓ 下移",      lambda: _move(1),             "#5d6d7e", "#4d5d6e"),
            ("➕ 新增",     _add_row,                      "#117a65", "#0e6655"),
            ("✏ 編輯",      _edit_row,                     "#1a5276", "#154360"),
            ("🗑 刪除",     _delete_row,                   "#922b21", "#7b241c"),
            ("🧭 南→北",   lambda: _sort_location(True),  "#1a5276", "#154360"),
            ("🧭 北→南",   lambda: _sort_location(False), "#1a5276", "#154360"),
            ("📍 計算時間", _calc_travel,                  "#6c3483", "#5b2c6f"),
        ]
        _btn_refs: dict = {}
        for text, cmd, color, hover in _BTN_DEFS:
            btn = ctk.CTkButton(btn_row, text=text, command=cmd,
                                 fg_color=color, hover_color=hover, text_color="white",
                                 font=FONT_S, width=72, height=28, corner_radius=4)
            btn.pack(side="left", padx=(0, 4))
            _btn_refs[text] = btn

        def _fetch_preview():
            from core.generator_schedule import fetch_events, events_to_rows
            sid  = sid_var.get().strip()
            csrf = csrf_var.get().strip()
            if not sid or not csrf:
                messagebox.showwarning("憑證未填", "請先填入 Session ID 與 CSRF Token", parent=parent)
                return
            target = date_entry.get_date()

            def _worker():
                evs = fetch_events(target, sid, csrf)
                return events_to_rows(evs), len(evs)

            def _on_done(result):
                rows, _n = result
                _rows.clear()
                _rows.extend(rows)
                _refresh_tree()

            self._run_task(_worker,
                            buttons=[fetch_btn],
                            status_label=fetch_status,
                            loading_text="抓取中…",
                            success_text=lambda result: f"找到 {result[1]} 筆事件",
                            on_success=_on_done)

        fetch_btn = ctk.CTkButton(date_row, text="🔍 抓取", command=_fetch_preview,
                                   fg_color="#117a65", hover_color="#0e6655", text_color="white",
                                   font=FONT, width=80, height=28, corner_radius=4)
        fetch_btn.pack(side="left")

        # ── Write button ──────────────────────────────────
        out_label = ctk.CTkLabel(parent, text="", fg_color="transparent",
                                  font=FONT_S, text_color=GRAY,
                                  anchor="w", wraplength=700)
        out_label.pack(fill="x", padx=16, pady=(2, 0))

        def _write_schedule():
            from core.generator_schedule import generate_schedule
            if not _rows:
                messagebox.showwarning("無資料", "請先抓取並確認事件清單", parent=parent); return
            sid    = sid_var.get().strip()
            csrf   = csrf_var.get().strip()
            target = date_entry.get_date()
            try:
                out = generate_schedule(target, sid, csrf, rows=list(_rows),
                                        schedule_file=self._get_path("schedule_file"))
                out_label.configure(text=f"✔  已寫入：{out}", text_color="#1e8449")
                if messagebox.askyesno("寫入成功",
                        f"排程已寫入：\n{out}\n\n是否立即開啟？", parent=parent):
                    os.startfile(str(out))
            except Exception as e:
                out_label.configure(text=f"✘  {e}", text_color="#c0392b")
                messagebox.showerror("寫入失敗", str(e), parent=parent)

        bb = ctk.CTkFrame(parent, fg_color=BG, corner_radius=0)
        bb.pack(fill="x", padx=12, pady=8)
        ctk.CTkButton(bb, text="✅  確認寫入出貨行程表.xlsx", command=_write_schedule,
                       fg_color="#1a5276", hover_color="#154360", text_color="white",
                       font=("Microsoft JhengHei UI", 12, "bold"),
                       height=44, corner_radius=8).pack(fill="x")
