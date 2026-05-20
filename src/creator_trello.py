"""
creator_trello.py — 從 Excel 讀取資料並在 Trello 建立卡片
"""
from pathlib import Path
import re

import openpyxl
import requests

_API_BASE   = "https://api.trello.com/1"
_BOARD_NAME = "物流事業部1"
_LIST_NAME  = "0.待評估"


def _auth(api_key: str, token: str) -> dict:
    return {"key": api_key, "token": token}


def _get_target_list_id(api_key: str, token: str) -> str:
    resp = requests.get(
        f"{_API_BASE}/members/me/boards",
        params={**_auth(api_key, token), "fields": "name,id"},
        timeout=15,
    )
    resp.raise_for_status()
    board_id = None
    for b in resp.json():
        if b["name"] == _BOARD_NAME:
            board_id = b["id"]
            break
    if not board_id:
        raise ValueError(f"找不到看板「{_BOARD_NAME}」")

    resp = requests.get(
        f"{_API_BASE}/boards/{board_id}/lists",
        params={**_auth(api_key, token), "fields": "name,id"},
        timeout=15,
    )
    resp.raise_for_status()
    for lst in resp.json():
        if _LIST_NAME in lst["name"]:
            return lst["id"]
    raise ValueError(f"找不到清單「{_LIST_NAME}」")


def _parse_phones(cell_value) -> tuple[str, str]:
    """從儲存格解析手機（10碼）與電話（9碼），回傳 (mobile, phone)。"""
    if not cell_value:
        return "", ""
    mobile, phone = "", ""
    for part in re.split(r'[\n\r/,、 ]+', str(cell_value).strip()):
        part = part.strip()
        if not part:
            continue
        digits = re.sub(r'[^0-9]', '', part)
        if len(digits) == 10 and not mobile:
            mobile = part
        elif len(digits) == 9 and not phone:
            phone = part
    return mobile, phone


def get_sheet_names(excel_path: Path) -> list[str]:
    """回傳 Excel 中所有工作表名稱。"""
    wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def read_excel_cards(excel_path: Path, sheet_name: str | None = None) -> list[dict]:
    """讀 Excel：B欄=標題，C=公司名，D=聯絡人，F=電話/手機，G=信箱，H=統編。

    回傳 list of {"row": int, "title": str, "desc": str}
    """
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    cards = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        b_val = row[1].value  # B欄 — 標題
        if not b_val:
            continue
        title = str(b_val).strip()
        if not title:
            continue

        c_val = str(row[2].value or "").strip()   # 公司名
        d_val = str(row[3].value or "").strip()   # 聯絡人
        f_val = row[5].value                       # 電話/手機
        g_val = str(row[6].value or "").strip()   # 電子信箱
        h_val = str(row[7].value or "").strip()   # 統一編號

        mobile, phone = _parse_phones(f_val)

        desc = (
            f"公司名：{c_val}\n"
            f"聯絡人：{d_val}\n"
            f"手機：{mobile}\n"
            f"電話：{phone}\n"
            f"傳真：\n"
            f"電子信箱：{g_val}\n"
            f"統一編號：{h_val}\n"
            f"地址："
        )

        cards.append({"row": row_idx, "title": title, "desc": desc})

    wb.close()
    return cards


def create_cards(cards: list[dict], api_key: str, token: str) -> int:
    """建立 Trello 卡片，回傳成功建立筆數。"""
    list_id = _get_target_list_id(api_key, token)
    created = 0
    for card in cards:
        resp = requests.post(
            f"{_API_BASE}/cards",
            params={
                **_auth(api_key, token),
                "idList": list_id,
                "name":   card["title"],
                "desc":   card["desc"],
                "pos":    "top",
            },
            timeout=15,
        )
        resp.raise_for_status()
        created += 1
    return created
