"""
app.py  —  報價單 → 出貨單 / 驗機單 / 維修單 轉換工具 (Tkinter GUI)
"""

import json, os, subprocess, sys, tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime, date as date_type
from pathlib import Path

from tkcalendar import DateEntry

sys.path.insert(0, str(Path(__file__).parent))
from parser import parse
from generator import generate
from generator_inspection import generate_inspection
from generator_fix import generate_fix
from generator_tag import generate_tag
from generator_label import generate_labels
from generator_schedule import generate_schedule, fetch_events, events_to_rows, calculate_travel_times, sort_rows_by_location
from syncer_trello import fetch_po_cards
from syncer_sheets import sync_cards
from syncer_production import sync_production
from creator_trello import read_excel_cards, create_cards as trello_create_cards, get_sheet_names
from downloader_trello import get_board_lists, get_list_cards, download_cards as trello_download_cards
from generator_quote import parse_card_desc, next_quote_no

from _paths import CONFIG_PATH, ICON_PATH, TEMPLATE_DIR, EXE_DIR

_GSHEETS_TOKEN_PATH      = EXE_DIR  / "gsheets_token.json"
_SYNCED_CARDS_PATH       = EXE_DIR  / "synced_cards.json"
_GSHEETS_CREDS_PATH      = TEMPLATE_DIR / "credentials.json"
_PRODUCTION_SYNCED_PATH  = EXE_DIR  / "production_synced_cards.json"

def _load_config():
    if CONFIG_PATH.exists():
        try:
            return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {"operators": ["小皋"]}

def _save_config(cfg):
    CONFIG_PATH.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("報價單轉單工具｜立善科技")
        self.update_idletasks()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        win_h = min(sh - 60, 920)
        self.geometry(f"960x{win_h}+{(sw-960)//2}+0")
        self.resizable(True, True)
        self.configure(bg="#f4f6f8")
        if ICON_PATH.exists():
            self._icon = tk.PhotoImage(file=str(ICON_PATH))
            self.iconphoto(True, self._icon)
            _ico = ICON_PATH.with_suffix(".ico")
            if not _ico.exists():
                try:
                    from PIL import Image
                    Image.open(ICON_PATH).save(_ico, format="ICO")
                except Exception:
                    pass
            if _ico.exists():
                try:
                    self.iconbitmap(str(_ico))
                except Exception:
                    pass
        self._parsed_data = None
        self._src_path = None
        self._config = _load_config()
        self._build_ui()

    # ════════════════════════════════════════════════════════
    #  UI 建構
    # ════════════════════════════════════════════════════════
    def _build_ui(self):
        PAD   = {"padx": 12, "pady": 4}
        FONT  = ("Microsoft JhengHei UI", 10)
        FONTB = ("Microsoft JhengHei UI", 10, "bold")
        BG    = "#f4f6f8"
        NAV_BG      = "#1b2631"
        NAV_ACTIVE  = "#1a5276"
        NAV_HOVER   = "#2e4057"
        NAV_FG      = "#d5d8dc"
        NAV_GRP_FG  = "#7f8c8d"

        # ── Top bar ──────────────────────────────────────────
        top = tk.Frame(self, bg="#1a5276", pady=8)
        top.pack(fill="x")
        tk.Label(top, text="立善科技｜QuoteFlow",
                 bg="#1a5276", fg="white",
                 font=("Microsoft JhengHei UI", 14, "bold")).pack(side="left", padx=16)
        tk.Button(top, text="選擇報價單 .xlsx ▶", command=self._open_file,
                  bg="#2e86c1", fg="white", relief="flat",
                  font=("Microsoft JhengHei UI", 10), padx=10, pady=3).pack(side="right", padx=(4, 16))
        tk.Button(top, text="⚙", command=self._open_paths_dialog,
                  bg="#1a5276", fg="white", relief="flat",
                  font=("Microsoft JhengHei UI", 13), padx=6, pady=1).pack(side="right")

        self._file_label = tk.Label(self, text="⚠  尚未選擇報價單",
                                    bg=BG, fg="#c0392b", font=FONT)
        self._file_label.pack(anchor="w", padx=16, pady=(3, 0))

        # ── 主體：側邊列 + 內容區 ────────────────────────────
        body = tk.Frame(self, bg=BG)
        body.pack(fill="both", expand=True)

        # ── 左側導覽列 ────────────────────────────────────────
        nav = tk.Frame(body, bg=NAV_BG, width=130)
        nav.pack(side="left", fill="y")
        nav.pack_propagate(False)

        content_area = tk.Frame(body, bg=BG)
        content_area.pack(side="left", fill="both", expand=True)

        # ── 頁面 Frame（每個功能一個，重疊在 content_area）──────
        self._pages: dict[str, tk.Frame] = {}
        self._nav_btns: dict[str, tk.Label] = {}
        self._active_page: str = ""

        def _show(key: str):
            if self._active_page and self._active_page in self._nav_btns:
                self._nav_btns[self._active_page].config(bg=NAV_BG)
            for f in self._pages.values():
                f.pack_forget()
            self._pages[key].pack(fill="both", expand=True)
            self._nav_btns[key].config(bg=NAV_ACTIVE)
            self._active_page = key

        self._show_page = _show

        def _make_page(key: str) -> tk.Frame:
            f = tk.Frame(content_area, bg=BG)
            self._pages[key] = f
            return f

        # ── 導覽列項目生成 ────────────────────────────────────
        def _nav_group(text: str):
            tk.Label(nav, text=text, bg=NAV_BG, fg=NAV_GRP_FG,
                     font=("Microsoft JhengHei UI", 8),
                     anchor="w", padx=10, pady=0).pack(fill="x", pady=(10, 1))
            tk.Frame(nav, bg=NAV_HOVER, height=1).pack(fill="x", padx=8)

        def _nav_item(text: str, key: str):
            lbl = tk.Label(nav, text=f"  {text}", bg=NAV_BG, fg=NAV_FG,
                           font=("Microsoft JhengHei UI", 9),
                           anchor="w", padx=6, pady=7, cursor="hand2")
            lbl.pack(fill="x")
            lbl.bind("<Button-1>", lambda e, k=key: _show(k))
            lbl.bind("<Enter>",    lambda e, b=lbl, k=key:
                         b.config(bg=NAV_HOVER) if self._active_page != k else None)
            lbl.bind("<Leave>",    lambda e, b=lbl, k=key:
                         b.config(bg=NAV_ACTIVE if self._active_page == k else NAV_BG))
            self._nav_btns[key] = lbl

        # ── 導覽結構 ──────────────────────────────────────────
        _nav_group("📄  單據生成")
        _nav_item("出貨單",   "shipping")
        _nav_item("報價單",   "quote")
        _nav_item("驗機單",   "inspection")
        _nav_item("維修單",   "fix")
        _nav_item("維修掛件", "tag")
        _nav_item("標籤生成", "label")

        _nav_group("📅  排程")
        _nav_item("出貨排程", "schedule")

        _nav_group("🃏  Trello")
        _nav_item("出貨一覽表",   "overview")
        _nav_item("生產群組紀錄", "production")
        _nav_item("建立卡片",     "create")
        _nav_item("下載卡片",     "download")

        # ── 建立各頁面內容 ────────────────────────────────────
        self._build_tab_shipping(   _make_page("shipping"),   PAD, FONT, FONTB, BG)
        self._build_tab_quote(      _make_page("quote"),      FONT, FONTB, BG)
        self._build_tab_inspection( _make_page("inspection"), PAD, FONT, FONTB, BG)
        self._build_tab_fix(        _make_page("fix"),        PAD, FONT, FONTB, BG)
        self._build_tab_tag(        _make_page("tag"),        PAD, FONT, FONTB, BG)
        self._build_tab_label(      _make_page("label"),      FONT, FONTB, BG)
        self._build_tab_schedule(   _make_page("schedule"),   FONT, FONTB, BG)
        self._build_tab_overview(   _make_page("overview"),   FONT, FONTB, BG)
        self._build_tab_production( _make_page("production"), FONT, FONTB, BG)
        self._build_tab_create_cards( _make_page("create"),   FONT, FONTB, BG)
        self._build_tab_download_cards(_make_page("download"),FONT, FONTB, BG)

        # 預設顯示出貨單
        _show("shipping")

    # ════════════════════════════════════════════════════════
    #  報價單生成（麥當勞三步驟流程）
    # ════════════════════════════════════════════════════════
    def _build_tab_quote(self, parent, FONT, FONTB, BG):
        import json as _json
        from generator_quote import load_product_catalog, generate_quote_from_cart

        GRAY   = "#5d6d7e"
        BLUE   = "#1a4a7a"
        GREEN  = "#1e8449"
        FONT_S = ("Microsoft JhengHei UI", 9)
        MONO   = ("Consolas", 9)
        _TARGET_LIST = "1.待報價(evaluate)"

        # ── 狀態 ────────────────────────────────────────────
        _all_cards:      list[dict]      = []
        _filtered_cards: list[dict]      = []
        _card:           list[dict|None] = [None]   # 選中的 Trello card
        _customer:       dict            = {}        # 解析後客戶資料
        _cart:           dict            = {}        # code → {product, qty, price}
        _cur_cat:        list[str]       = ["all"]
        _products: list[dict] = load_product_catalog(TEMPLATE_DIR)
        CATS = ["全部"] + sorted({p["category"] for p in _products})

        # ════════════════════════════════════════════════════
        # 右側購物車 Panel（固定 272px，所有步驟皆顯示）
        # ════════════════════════════════════════════════════
        body = tk.Frame(parent, bg=BG)
        body.pack(fill="both", expand=True)

        left_area = tk.Frame(body, bg=BG)
        left_area.pack(side="left", fill="both", expand=True)

        right_border = tk.Frame(body, bg="#c8cfd6", width=1)
        right_border.pack(side="right", fill="y")
        right_border.pack_propagate(False)

        cart_panel = tk.Frame(body, bg=BG, width=272)
        cart_panel.pack(side="right", fill="y")
        cart_panel.pack_propagate(False)

        # 購物車 Header
        ch = tk.Frame(cart_panel, bg="#e8ecf0")
        ch.pack(fill="x")
        tk.Label(ch, text="🛒  報價清單", bg="#e8ecf0", fg=BLUE,
                 font=(FONTB[0], 10, "bold"), pady=7).pack(side="left", padx=10)
        cart_cnt_lbl = tk.Label(ch, text="", bg="#e8ecf0", fg=GRAY, font=FONT_S)
        cart_cnt_lbl.pack(side="right", padx=10)

        # 購物車 可捲動 items
        cart_mid = tk.Frame(cart_panel, bg=BG)
        cart_mid.pack(fill="both", expand=True)
        cart_cvs = tk.Canvas(cart_mid, bg=BG, highlightthickness=0)
        cart_sb  = ttk.Scrollbar(cart_mid, orient="vertical", command=cart_cvs.yview)
        cart_cvs.configure(yscrollcommand=cart_sb.set)
        cart_sb.pack(side="right", fill="y")
        cart_cvs.pack(side="left", fill="both", expand=True)
        cart_inner = tk.Frame(cart_cvs, bg=BG)
        _cart_win  = cart_cvs.create_window((0, 0), window=cart_inner, anchor="nw")

        def _on_cart_resize(e=None):
            cart_cvs.configure(scrollregion=cart_cvs.bbox("all"))
            if cart_cvs.winfo_width() > 1:
                cart_cvs.itemconfig(_cart_win, width=cart_cvs.winfo_width())
        cart_inner.bind("<Configure>", _on_cart_resize)
        cart_cvs.bind("<Configure>",   _on_cart_resize)

        # 購物車 Footer
        cf = tk.Frame(cart_panel, bg="#f0f4f8", bd=1, relief="ridge")
        cf.pack(fill="x")
        cart_sub_lbl = tk.Label(cf, text="小計：—", bg="#f0f4f8", fg=GRAY,
                                 font=FONT_S, anchor="e")
        cart_sub_lbl.pack(fill="x", padx=10, pady=(7, 1))
        cart_tax_lbl = tk.Label(cf, text="營業稅 5%：—", bg="#f0f4f8", fg=GRAY,
                                 font=FONT_S, anchor="e")
        cart_tax_lbl.pack(fill="x", padx=10, pady=1)
        cart_tot_lbl = tk.Label(cf, text="合計：—", bg="#f0f4f8",
                                 font=(FONTB[0], 10, "bold"), fg=BLUE, anchor="e")
        cart_tot_lbl.pack(fill="x", padx=10, pady=(1, 4))
        checkout_btn = tk.Button(cf, text="確認報價內容 →", state="disabled",
                                  bg=BLUE, fg="white", relief="flat",
                                  font=FONT_S, pady=5)
        checkout_btn.pack(fill="x", padx=8, pady=(0, 8))

        def _fmt(n): return f"${int(round(n)):,}"

        def _render_cart():
            for w in cart_inner.winfo_children():
                w.destroy()
            items = [(c, v) for c, v in _cart.items() if v["qty"] > 0]
            if not items:
                tk.Label(cart_inner, text="\n📦\n點擊左側產品加入",
                         bg=BG, fg=GRAY, font=FONT_S, justify="center"
                         ).pack(pady=20)
                cart_cnt_lbl.config(text="")
                cart_sub_lbl.config(text="小計：—")
                cart_tax_lbl.config(text="營業稅 5%：—")
                cart_tot_lbl.config(text="合計：—")
                checkout_btn.config(state="disabled")
                return
            total_qty = sum(v["qty"] for _, v in items)
            cart_cnt_lbl.config(text=f"{total_qty} 項")
            for code, item in items:
                if_f = tk.Frame(cart_inner, bg=BG)
                if_f.pack(fill="x", padx=6, pady=3)
                tk.Frame(cart_inner, bg="#dee2e6", height=1).pack(fill="x", padx=4)
                tk.Label(if_f, text=code, bg=BG, fg=GRAY, font=MONO, anchor="w"
                         ).pack(fill="x")
                name_t = item["product"]["name"]
                if len(name_t) > 22: name_t = name_t[:20] + "…"
                tk.Label(if_f, text=name_t, bg=BG, fg="#2c3e50",
                         font=(FONT_S[0], 9, "bold"), anchor="w").pack(fill="x")
                ctrl = tk.Frame(if_f, bg=BG)
                ctrl.pack(fill="x", pady=(2, 0))

                def _dec(c=code):
                    _cart[c]["qty"] -= 1
                    if _cart[c]["qty"] <= 0:
                        del _cart[c]
                    _render_cart()
                    _render_products()

                def _inc(c=code):
                    _cart[c]["qty"] += 1
                    _render_cart()
                    _render_products()

                tk.Button(ctrl, text="−", command=_dec, width=2, font=FONT_S,
                          relief="flat", bg="#dee2e6", fg="#333").pack(side="left")
                tk.Label(ctrl, text=str(item["qty"]), bg=BG, fg="#2c3e50",
                         font=FONT_S, width=3).pack(side="left")
                tk.Button(ctrl, text="＋", command=_inc, width=2, font=FONT_S,
                          relief="flat", bg="#dee2e6", fg="#333").pack(side="left")
                tk.Label(ctrl, text="  ￥", bg=BG, fg=GRAY, font=FONT_S
                         ).pack(side="left")
                pv = tk.StringVar(value=str(int(item["price"])))

                def _price_changed(pvar=pv, c=code, *_):
                    try:
                        _cart[c]["price"] = float(pvar.get().replace(",", ""))
                    except ValueError:
                        pass
                    _update_cart_totals()

                pv.trace_add("write", _price_changed)
                tk.Entry(ctrl, textvariable=pv, width=7, font=FONT_S
                         ).pack(side="left")
                sub = item["qty"] * item["price"]
                tk.Label(if_f, text=f"= {_fmt(sub)}", bg=BG, fg=GRAY,
                         font=FONT_S, anchor="e").pack(fill="x")
            _update_cart_totals()

        def _update_cart_totals():
            items = [v for v in _cart.values() if v["qty"] > 0]
            sub   = sum(v["qty"] * v["price"] for v in items)
            tax   = round(sub * 0.05)
            tot   = sub + tax
            cart_sub_lbl.config(text=f"小計：{_fmt(sub)}")
            cart_tax_lbl.config(text=f"營業稅 5%：{_fmt(tax)}")
            cart_tot_lbl.config(text=f"合計：{_fmt(tot)}")
            checkout_btn.config(state="normal" if items else "disabled")

        # ════════════════════════════════════════════════════
        # Step Bar（步驟進度列）
        # ════════════════════════════════════════════════════
        step_bar = tk.Frame(left_area, bg="#e8ecf0")
        step_bar.pack(fill="x")
        _step_btns: list[tk.Label] = []
        for i, txt in enumerate(["① 顧客資料", "② 選取品項", "③ 報價確認"], 1):
            lbl = tk.Label(step_bar, text=txt, bg="#e8ecf0", fg=GRAY,
                           font=FONT_S, pady=7, cursor="hand2")
            lbl.pack(side="left", expand=True, fill="x")
            lbl.bind("<Button-1>", lambda e, n=i: _go_step(n))
            _step_btns.append(lbl)
            if i < 3:
                tk.Frame(step_bar, bg="#c8cfd6", width=1).pack(side="left", fill="y")

        step_content = tk.Frame(left_area, bg=BG)
        step_content.pack(fill="both", expand=True)

        # ════════════════════════════════════════════════════
        # Step 1 — 顧客資料
        # ════════════════════════════════════════════════════
        p1 = tk.Frame(step_content, bg=BG)

        # Trello 載入列
        p1_top = tk.Frame(p1, bg=BG)
        p1_top.pack(fill="x", padx=10, pady=(8, 2))
        tk.Label(p1_top, text=f"來源：{_TARGET_LIST}",
                 bg=BG, font=FONT_S, fg=GRAY).pack(side="left")
        p1_status = tk.Label(p1_top, text="", bg=BG, font=FONT_S, fg=GRAY)
        p1_status.pack(side="left", padx=8)

        def _get_tr_creds():
            tr_cfg  = self._config.get("trello", {})
            api_key = tr_cfg.get("api_key", "").strip()
            token   = tr_cfg.get("token",   "").strip()
            if not api_key or not token:
                messagebox.showwarning("憑證未設定",
                    "請先至「出貨一覽表」頁籤填入並儲存 Trello 憑證", parent=parent)
                return None, None
            return api_key, token

        def _refresh_card_tree(cards):
            _filtered_cards.clear()
            _filtered_cards.extend(cards)
            card_tree.delete(*card_tree.get_children())
            for card in cards:
                desc    = parse_card_desc(card.get("desc", ""))
                company = desc.get("公司名") or desc.get("公司名稱", "")
                card_tree.insert("", "end", values=(
                    card["name"], company, desc.get("聯絡人", "")))
            if card_tree.get_children():
                card_tree.selection_set(card_tree.get_children()[0])
                _on_card_select()
            shown = len(cards)
            total = len(_all_cards)
            card_lf.config(
                text=f"待報價卡片（共 {total} 張{'，顯示 '+str(shown)+' 張' if shown<total else ''}）")

        def _apply_card_search(*_):
            kw = p1_search_var.get().strip().lower()
            matched = [c for c in _all_cards if kw in c["name"].lower()] if kw else list(_all_cards)
            _refresh_card_tree(matched)

        def _load_cards():
            api_key, token = _get_tr_creds()
            if not api_key: return
            p1_status.config(text="載入中…", fg=GRAY)
            parent.update_idletasks()
            try:
                lists  = get_board_lists(api_key, token)
                target = next((l for l in lists
                               if "待報價" in l["name"] or "evaluate" in l["name"].lower()), None)
                if not target:
                    p1_status.config(text=f'✘ 找不到清單', fg="#c0392b"); return
                cards = get_list_cards(target["id"], api_key, token)
                _all_cards.clear(); _all_cards.extend(cards)
                p1_search_var.set("")
                _refresh_card_tree(_all_cards)
                p1_status.config(text=f"✔ 找到 {len(cards)} 張", fg=GREEN)
            except Exception as e:
                p1_status.config(text=f"✘ {e}", fg="#c0392b")

        tk.Button(p1_top, text="🔄 載入", command=_load_cards,
                  bg="#2e86c1", fg="white", relief="flat",
                  font=FONT_S, padx=8).pack(side="right")

        # 搜索列
        p1_search_row = tk.Frame(p1, bg=BG)
        p1_search_row.pack(fill="x", padx=10, pady=(0, 2))
        tk.Label(p1_search_row, text="🔍", bg=BG, font=FONT_S, fg=GRAY).pack(side="left")
        p1_search_var = tk.StringVar()
        tk.Entry(p1_search_row, textvariable=p1_search_var,
                 font=FONT_S, width=28).pack(side="left", padx=(2, 4))
        tk.Button(p1_search_row, text="清除", command=lambda: p1_search_var.set(""),
                  bg="#5d6d7e", fg="white", relief="flat", font=FONT_S, padx=4
                  ).pack(side="left")
        p1_search_var.trace_add("write", _apply_card_search)

        # 卡片 Treeview
        card_lf = tk.LabelFrame(p1, text="待報價卡片（共 0 張）", bg=BG, font=FONT_S)
        card_lf.pack(fill="both", expand=True, padx=10, pady=2)
        _ccols = ("name", "company", "contact")
        card_tree = ttk.Treeview(card_lf, columns=_ccols, show="headings",
                                  selectmode="browse", height=5)
        card_tree.heading("name",    text="卡片標題")
        card_tree.heading("company", text="公司名")
        card_tree.heading("contact", text="聯絡人")
        card_tree.column("name",    width=230, anchor="w", stretch=True)
        card_tree.column("company", width=130, anchor="w", stretch=False)
        card_tree.column("contact", width=80,  anchor="w", stretch=False)
        ct_vsb = ttk.Scrollbar(card_lf, orient="vertical", command=card_tree.yview)
        card_tree.configure(yscrollcommand=ct_vsb.set)
        card_tree.pack(side="left", fill="both", expand=True)
        ct_vsb.pack(side="right", fill="y")

        # 報價設定列（放在客戶資料上方，日曆才不會被遮住）
        cfg_f = tk.LabelFrame(p1, text="報價設定", bg=BG, font=FONT_S)
        cfg_f.pack(fill="x", padx=10, pady=(4, 2))

        cfg_row0 = tk.Frame(cfg_f, bg=BG)
        cfg_row0.pack(fill="x", padx=6, pady=(4, 2))
        tk.Label(cfg_row0, text="製表人員：", bg=BG, font=FONT_S, fg=GRAY).pack(side="left")
        operators = self._config.get("operators", ["小皋"])
        op_var  = tk.StringVar(value=operators[0] if operators else "")
        op_cb   = ttk.Combobox(cfg_row0, textvariable=op_var, values=operators,
                                font=FONT_S, state="readonly", width=10)
        op_cb.pack(side="left", padx=(0, 16))

        tk.Label(cfg_row0, text="報價日期：", bg=BG, font=FONT_S, fg=GRAY).pack(side="left")
        date_entry = DateEntry(cfg_row0, font=FONT_S, date_pattern="yyyy/mm/dd",
                               width=12, background="#2e86c1", foreground="white",
                               borderwidth=1)
        date_entry.pack(side="left", padx=(0, 6))

        cfg_row1 = tk.Frame(cfg_f, bg=BG)
        cfg_row1.pack(fill="x", padx=6, pady=(2, 6))
        valid_lbl = tk.Label(cfg_row1, text="", bg=BG, font=FONT_S, fg=GRAY)
        valid_lbl.pack(side="left", padx=(0, 16))
        tk.Label(cfg_row1, text="報價單號：", bg=BG, font=FONT_S, fg=GRAY).pack(side="left")
        quote_no_var = tk.StringVar()
        tk.Entry(cfg_row1, textvariable=quote_no_var, font=FONT_S, width=16
                 ).pack(side="left", padx=(0, 6))

        def _update_valid_label(*_):
            try:
                from datetime import timedelta as _td
                d  = date_entry.get_date()
                vd = d + _td(days=15)
                valid_lbl.config(text=f"有效日期：{vd.strftime('%Y/%m/%d')}")
                _refresh_quote_no()
            except Exception:
                pass

        def _refresh_quote_no(*_):
            try:
                d     = date_entry.get_date()
                op    = op_var.get().strip()
                codes = self._config.get("operator_codes", {})
                code  = codes.get(op, op[:1].upper() if op else "X")
                no    = next_quote_no(self._get_path("output_quote"), code, d)
                quote_no_var.set(no)
            except Exception:
                pass

        date_entry.bind("<<DateEntrySelected>>", _update_valid_label)
        op_var.trace_add("write", _refresh_quote_no)
        _update_valid_label()

        # 客戶資訊預覽區
        cust_lf = tk.LabelFrame(p1, text="客戶資料", bg=BG, font=FONT_S)
        cust_lf.pack(fill="x", padx=10, pady=(2, 4))
        cust_lf.columnconfigure(1, weight=1)
        cust_lf.columnconfigure(3, weight=1)
        _cust_vars: dict[str, tk.StringVar] = {}
        for r, (label, key) in enumerate([
            ("公司名稱", "company"), ("聯絡人", "contact"),
            ("電話",     "phone"),   ("傳真",   "fax"),
            ("地址",     "address"), ("統一編號","tax_id"),
            ("E-MAIL",   "email"),
        ]):
            col = (r % 2) * 2
            row_i = r // 2
            tk.Label(cust_lf, text=label + "：", bg=BG, fg=GRAY,
                     font=FONT_S, anchor="e").grid(row=row_i, column=col, sticky="e",
                                                    padx=(6, 2), pady=2)
            sv = tk.StringVar()
            _cust_vars[key] = sv
            tk.Entry(cust_lf, textvariable=sv, font=FONT_S, state="readonly",
                     width=18).grid(row=row_i, column=col+1, sticky="ew", padx=(0, 8), pady=2)

        def _strip_md_link(text: str) -> str:
            """去除 Trello Markdown 連結格式，只保留顯示文字。
            [text](url) → text
            """
            import re as _re
            return _re.sub(r'\[([^\]]+)\]\([^)]*\)', r'\1', text).strip()

        def _on_card_select(*_):
            sel = card_tree.selection()
            if not sel: return
            idx  = card_tree.index(sel[0])
            card = _filtered_cards[idx]
            _card[0] = card
            desc = parse_card_desc(card.get("desc", ""))

            def _g(*keys):
                for k in keys:
                    v = desc.get(k)
                    if v: return _strip_md_link(v)
                return ""

            _customer.clear()
            _customer.update({
                "company": _g("公司名", "公司名稱"),
                "contact": _g("聯絡人"),
                "phone":   _g("手機", "電話"),
                "fax":     _g("傳真"),
                "address": _g("地址", "聯絡地址"),
                "tax_id":  _g("統一編號", "统一編號", "统編", "統編"),
                "email":   _g("電子信箱", "E-MAIL", "EMAIL"),
            })
            for key, sv in _cust_vars.items():
                sv.set(_customer.get(key, ""))

        card_tree.bind("<<TreeviewSelect>>", _on_card_select)

        p1_next_btn = tk.Button(p1, text="下一步：選取品項 →",
                                 command=lambda: _go_step(2),
                                 bg=BLUE, fg="white", relief="flat",
                                 font=(FONT_S[0], 10, "bold"), pady=7)
        p1_next_btn.pack(fill="x", padx=10, pady=(2, 8))

        # ════════════════════════════════════════════════════
        # Step 2 — 選取品項
        # ════════════════════════════════════════════════════
        p2 = tk.Frame(step_content, bg=BG)

        # 分類 Tabs
        cat_bar = tk.Frame(p2, bg=BG)
        cat_bar.pack(fill="x", padx=10, pady=(8, 4))
        _cat_btns: dict[str, tk.Button] = {}

        def _set_cat(cat: str):
            _cur_cat[0] = cat
            for c, b in _cat_btns.items():
                b.config(bg=BLUE if c == cat else "#dee2e6",
                         fg="white" if c == cat else "#333")
            _render_products()

        for cat in CATS:
            b = tk.Button(cat_bar, text=cat, font=FONT_S, relief="flat",
                          bg="#dee2e6", fg="#333", padx=8, pady=3,
                          command=lambda c=cat: _set_cat(c))
            b.pack(side="left", padx=(0, 4))
            _cat_btns[cat] = b
        # _set_cat(CATS[0]) 移到 _render_products 定義之後執行

        # 產品格狀列表（Canvas + 2欄）
        prod_outer = tk.Frame(p2, bg=BG)
        prod_outer.pack(fill="both", expand=True, padx=10, pady=(0, 6))

        prod_cvs = tk.Canvas(prod_outer, bg=BG, highlightthickness=0)
        prod_sb  = ttk.Scrollbar(prod_outer, orient="vertical", command=prod_cvs.yview)
        prod_cvs.configure(yscrollcommand=prod_sb.set)
        prod_sb.pack(side="right", fill="y")
        prod_cvs.pack(side="left", fill="both", expand=True)

        prod_inner = tk.Frame(prod_cvs, bg=BG)
        _prod_win  = prod_cvs.create_window((0, 0), window=prod_inner, anchor="nw")

        def _on_prod_resize(e=None):
            prod_cvs.configure(scrollregion=prod_cvs.bbox("all"))
            if prod_cvs.winfo_width() > 1:
                prod_cvs.itemconfig(_prod_win, width=prod_cvs.winfo_width())
        prod_inner.bind("<Configure>", _on_prod_resize)
        prod_cvs.bind("<Configure>",   _on_prod_resize)

        def _render_products():
            for w in prod_inner.winfo_children():
                w.destroy()
            cat = _cur_cat[0]
            filtered = [p for p in _products
                        if cat == "全部" or p.get("category") == cat]
            for idx, prod in enumerate(filtered):
                col   = idx % 2
                row_i = idx // 2
                in_cart = prod["code"] in _cart and _cart[prod["code"]]["qty"] > 0
                card_bg = "#d8eaf8" if in_cart else "white"
                bdr     = BLUE if in_cart else "#d0d7de"

                cf2 = tk.Frame(prod_inner, bg=card_bg, bd=1, relief="solid",
                               highlightbackground=bdr, highlightthickness=1)
                cf2.grid(row=row_i, column=col, padx=4, pady=4, sticky="nsew")
                prod_inner.columnconfigure(col, weight=1)

                tk.Label(cf2, text=prod["code"], bg=card_bg, fg=GRAY,
                         font=MONO, anchor="w").pack(fill="x", padx=8, pady=(6, 0))
                # 品名（去掉 code 前綴）
                disp_name = prod["name"].replace(prod["code"] + " ", "").replace(prod["code"], "")
                tk.Label(cf2, text=disp_name.strip(), bg=card_bg, fg="#1a1a1a",
                         font=(FONT_S[0], 9, "bold"), anchor="w",
                         wraplength=160, justify="left").pack(fill="x", padx=8, pady=(0, 2))
                tk.Label(cf2, text=prod.get("spec", ""), bg=card_bg, fg=GRAY,
                         font=(MONO[0], 8), anchor="w",
                         wraplength=160, justify="left").pack(fill="x", padx=8, pady=(0, 4))

                bot_f = tk.Frame(cf2, bg=card_bg)
                bot_f.pack(fill="x", padx=8, pady=(0, 6))
                tk.Label(bot_f, text=f"${prod['price']:,}/{prod['unit']}",
                         bg=card_bg, fg=BLUE,
                         font=(FONTB[0], 10, "bold")).pack(side="left")

                def _add(p=prod):
                    code = p["code"]
                    if code not in _cart:
                        _cart[code] = {"product": p, "qty": 0, "price": p["price"]}
                    _cart[code]["qty"] += 1
                    _render_cart()
                    _render_products()

                add_lbl = "✔" if in_cart else "+"
                add_bg  = GREEN if in_cart else BLUE
                tk.Button(bot_f, text=add_lbl, command=_add,
                          bg=add_bg, fg="white", relief="flat",
                          font=(FONTB[0], 12), width=2, pady=0
                          ).pack(side="right")

        # 初始化：設定第一個分類並渲染
        if CATS:
            _cur_cat[0] = CATS[0]
            for c, b in _cat_btns.items():
                b.config(bg=BLUE if c == CATS[0] else "#dee2e6",
                         fg="white" if c == CATS[0] else "#333")
        _render_products()

        p2_back_btn = tk.Button(p2, text="← 返回顧客資料",
                                 command=lambda: _go_step(1),
                                 bg="#5d6d7e", fg="white", relief="flat",
                                 font=FONT_S, padx=8, pady=4)
        p2_back_btn.pack(anchor="w", padx=10, pady=(0, 6))

        # ════════════════════════════════════════════════════
        # Step 3 — 報價確認
        # ════════════════════════════════════════════════════
        p3 = tk.Frame(step_content, bg=BG)

        # 客戶摘要
        p3_cust_f = tk.LabelFrame(p3, text="客戶資料", bg=BG, font=FONT_S)
        p3_cust_f.pack(fill="x", padx=10, pady=(8, 4))
        p3_cust_f.columnconfigure(1, weight=1)
        p3_cust_f.columnconfigure(3, weight=1)
        _p3_cust_lbls: dict[str, tk.Label] = {}
        for r, (label, key) in enumerate([
            ("公司名稱", "company"), ("聯絡人",  "contact"),
            ("報價單號", "_quote_no"), ("報價日期", "_quote_date"),
        ]):
            col = (r % 2) * 2
            row_i = r // 2
            tk.Label(p3_cust_f, text=label + "：", bg=BG, fg=GRAY,
                     font=FONT_S).grid(row=row_i, column=col, sticky="e", padx=(6, 2), pady=2)
            lbl = tk.Label(p3_cust_f, text="—", bg=BG, fg="#1a1a1a",
                           font=(FONT_S[0], 9, "bold"), anchor="w")
            lbl.grid(row=row_i, column=col+1, sticky="ew", padx=(0, 8), pady=2)
            _p3_cust_lbls[key] = lbl

        # 品項確認表
        conf_lf = tk.LabelFrame(p3, text="報價品項", bg=BG, font=FONT_S)
        conf_lf.pack(fill="both", expand=True, padx=10, pady=4)
        _ccols3 = ("code", "name", "qty", "price", "subtotal")
        conf_tree = ttk.Treeview(conf_lf, columns=_ccols3, show="headings",
                                  selectmode="none", height=6)
        conf_tree.heading("code",     text="品號")
        conf_tree.heading("name",     text="品名")
        conf_tree.heading("qty",      text="數量")
        conf_tree.heading("price",    text="單價")
        conf_tree.heading("subtotal", text="小計")
        conf_tree.column("code",     width=100, anchor="w")
        conf_tree.column("name",     width=200, anchor="w", stretch=True)
        conf_tree.column("qty",      width=55,  anchor="center", stretch=False)
        conf_tree.column("price",    width=80,  anchor="e", stretch=False)
        conf_tree.column("subtotal", width=90,  anchor="e", stretch=False)
        conf_tree.pack(fill="both", expand=True)

        # 摘要
        sum_f = tk.Frame(p3, bg="#f0f4f8", bd=1, relief="ridge")
        sum_f.pack(fill="x", padx=10, pady=(0, 4))
        p3_sub_lbl = tk.Label(sum_f, text="小計：—", bg="#f0f4f8", fg=GRAY,
                               font=FONT_S, anchor="e")
        p3_sub_lbl.pack(fill="x", padx=12, pady=(6, 1))
        p3_tax_lbl = tk.Label(sum_f, text="營業稅 5%：—", bg="#f0f4f8", fg=GRAY,
                               font=FONT_S, anchor="e")
        p3_tax_lbl.pack(fill="x", padx=12, pady=1)
        p3_tot_lbl = tk.Label(sum_f, text="應收總金額：—", bg="#f0f4f8",
                               font=(FONTB[0], 11, "bold"), fg=BLUE, anchor="e")
        p3_tot_lbl.pack(fill="x", padx=12, pady=(1, 6))

        p3_out_lbl = tk.Label(p3, text="", bg=BG, font=FONT_S, fg=GRAY,
                               anchor="w", wraplength=620)
        p3_out_lbl.pack(fill="x", padx=10)

        def _build_confirm():
            conf_tree.delete(*conf_tree.get_children())
            items = [v for v in _cart.values() if v["qty"] > 0]
            sub   = 0
            for v in items:
                s = v["qty"] * v["price"]
                sub += s
                conf_tree.insert("", "end", values=(
                    v["product"]["code"],
                    v["product"]["name"][:30],
                    f"{v['qty']} {v['product']['unit']}",
                    f"${int(v['price']):,}",
                    f"${int(s):,}",
                ))
            tax = round(sub * 0.05)
            p3_sub_lbl.config(text=f"小計：{_fmt(sub)}")
            p3_tax_lbl.config(text=f"營業稅 5%：{_fmt(tax)}")
            p3_tot_lbl.config(text=f"應收總金額：{_fmt(sub+tax)}")
            _p3_cust_lbls["company"].config(    text=_customer.get("company", "—"))
            _p3_cust_lbls["contact"].config(    text=_customer.get("contact", "—"))
            _p3_cust_lbls["_quote_no"].config(  text=quote_no_var.get())
            _p3_cust_lbls["_quote_date"].config( text=date_entry.get_date().strftime("%Y/%m/%d"))

        def _generate():
            if not _cart:
                messagebox.showwarning("購物車空白", "請先選取品項", parent=parent); return
            if not _customer.get("company"):
                messagebox.showwarning("無客戶資料", "請先在步驟①選取 Trello 卡片", parent=parent); return
            tpl_path = TEMPLATE_DIR / "template_quote.xlsx"
            if not tpl_path.exists():
                messagebox.showerror("找不到範本", f"找不到 {tpl_path}", parent=parent); return

            cart_items = [
                {**{"code":  v["product"]["code"],
                    "name":  v["product"]["name"],
                    "spec":  v["product"].get("spec", ""),
                    "unit":  v["product"]["unit"],
                    "qty":   v["qty"],
                    "price": v["price"]}}
                for v in _cart.values() if v["qty"] > 0
            ]
            q_date   = date_entry.get_date()
            quote_no = quote_no_var.get().strip()
            out_dir  = self._get_path("output_quote")

            p3_out_lbl.config(text="生成中…", fg=GRAY)
            parent.update_idletasks()
            try:
                out_path = generate_quote_from_cart(
                    _customer, cart_items, tpl_path, out_dir, quote_no, q_date,
                    operator=op_var.get().strip())
                p3_out_lbl.config(text=f"✔  已生成：{out_path}", fg=GREEN)
                if messagebox.askyesno("完成",
                        f"報價單已生成\n{out_path}\n\n是否立即開啟？", parent=parent):
                    os.startfile(str(out_path))
            except Exception as e:
                p3_out_lbl.config(text=f"✘  {e}", fg="#c0392b")
                messagebox.showerror("生成失敗", str(e), parent=parent)

        gen_btn = tk.Button(p3, text="📄  生成報價單 .xlsx", command=_generate,
                             bg=GREEN, fg="white", relief="flat",
                             font=(FONTB[0], 12, "bold"), pady=8)
        gen_btn.pack(fill="x", padx=10, pady=(0, 6))

        p3_back_btn = tk.Button(p3, text="← 返回選取品項",
                                 command=lambda: _go_step(2),
                                 bg="#5d6d7e", fg="white", relief="flat",
                                 font=FONT_S, padx=8, pady=3)
        p3_back_btn.pack(anchor="w", padx=10, pady=(0, 4))

        # ════════════════════════════════════════════════════
        # Step 切換
        # ════════════════════════════════════════════════════
        _panels = {1: p1, 2: p2, 3: p3}

        def _go_step(n: int):
            for i, lbl in enumerate(_step_btns, 1):
                if i == n:
                    lbl.config(bg=BLUE, fg="white")
                elif i < n:
                    lbl.config(bg="#d5e8f5", fg=BLUE)
                else:
                    lbl.config(bg="#e8ecf0", fg=GRAY)
            for i, panel in _panels.items():
                if i == n:
                    panel.pack(fill="both", expand=True)
                else:
                    panel.pack_forget()
            if n == 3:
                _build_confirm()

        checkout_btn.config(command=lambda: _go_step(3))
        _render_cart()
        _go_step(1)

    # ── Tab 1：出貨單 ─────────────────────────────────────────
    def _build_tab_shipping(self, parent, PAD, FONT, FONTB, BG):
        # 生成按鈕（先 pack bottom 確保不被擠掉）
        bb = tk.Frame(parent, bg=BG, pady=8)
        bb.pack(side="bottom", fill="x", padx=12)
        tk.Button(bb, text="＋ 新增", command=self._add_row,
                  bg="#27ae60", fg="white", relief="flat",
                  font=FONT, padx=10, pady=3).pack(side="left", padx=(0, 6))
        tk.Button(bb, text="－ 刪除", command=self._del_row,
                  bg="#c0392b", fg="white", relief="flat",
                  font=FONT, padx=10, pady=3).pack(side="left")
        tk.Button(bb, text="⬇  生成出貨單", command=self._generate,
                  bg="#1a5276", fg="white", relief="flat",
                  font=("Microsoft JhengHei UI", 11, "bold"),
                  padx=16, pady=6).pack(side="right")

        # 欄位區
        mid = tk.Frame(parent, bg=BG)
        mid.pack(fill="x", padx=12, pady=6)
        mid.columnconfigure(0, weight=1)
        mid.columnconfigure(1, weight=1)

        lf = tk.LabelFrame(mid, text="從報價單讀入", bg=BG, font=FONTB)
        lf.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
        lf.columnconfigure(1, weight=1)

        self._read_vars = {}
        for i, (label, key) in enumerate([
            ("客戶名稱", "customer"), ("聯絡電話", "phone"),
            ("聯絡人",   "contact"),  ("地址",     "address"),
            ("報價單號", "quote_no"), ("報價日期", "quote_date"),
        ]):
            tk.Label(lf, text=label + "：", bg=BG, anchor="w", font=FONT
                     ).grid(row=i, column=0, sticky="w", **PAD)
            var = tk.StringVar(value="—")
            tk.Entry(lf, textvariable=var, font=FONT
                     ).grid(row=i, column=1, sticky="ew", **PAD)
            self._read_vars[key] = var

        rf = tk.LabelFrame(mid, text="補填欄位", bg=BG, font=FONTB)
        rf.grid(row=0, column=1, sticky="nsew", padx=(6, 0))
        rf.columnconfigure(1, weight=1)

        self._fill_vars = {}
        for i, (label, key, default) in enumerate([
            ("出貨日期", "ship_date", datetime.today().strftime("%Y/%m/%d")),
            ("銷貨單號", "sale_no",   ""),
            ("附註",     "note",      ""),
        ]):
            tk.Label(rf, text=label + "：", bg=BG, anchor="w", font=FONT
                     ).grid(row=i, column=0, sticky="w", **PAD)
            var = tk.StringVar(value=default)
            tk.Entry(rf, textvariable=var, font=FONT
                     ).grid(row=i, column=1, sticky="ew", **PAD)
            self._fill_vars[key] = var

        tk.Label(rf, text="製表人員：", bg=BG, anchor="w", font=FONT
                 ).grid(row=3, column=0, sticky="w", **PAD)
        op_f = tk.Frame(rf, bg=BG)
        op_f.grid(row=3, column=1, sticky="ew", **PAD)
        self._operator_var = tk.StringVar()
        self._operator_cb  = ttk.Combobox(op_f, textvariable=self._operator_var,
                                           values=self._config["operators"],
                                           width=12, font=FONT, state="readonly")
        if self._config["operators"]:
            self._operator_var.set(self._config["operators"][0])
        self._operator_cb.pack(side="left")
        tk.Button(op_f, text="＋", command=self._add_operator,
                  bg="#27ae60", fg="white", relief="flat",
                  font=FONT, width=3).pack(side="left", padx=(4, 0))
        tk.Button(op_f, text="－", command=self._del_operator,
                  bg="#c0392b", fg="white", relief="flat",
                  font=FONT, width=3).pack(side="left", padx=(2, 0))

        tk.Label(rf, text="發票方式：", bg=BG, anchor="w", font=FONT
                 ).grid(row=4, column=0, sticky="w", **PAD)
        inv_f = tk.Frame(rf, bg=BG)
        inv_f.grid(row=4, column=1, sticky="w", **PAD)
        self._invoice_var = tk.StringVar(value="尚未確認")
        for lbl, val in [("尚未確認", "尚未確認"), ("隨貨", "隨貨"), ("直寄", "直寄")]:
            tk.Radiobutton(inv_f, text=lbl, variable=self._invoice_var,
                           value=val, bg=BG, font=FONT,
                           activebackground=BG).pack(side="left", padx=(0, 8))

        # 品項列表
        tf = tk.LabelFrame(parent, text="品項列表（雙擊儲存格可編輯）",
                           bg=BG, font=FONTB)
        tf.pack(fill="both", expand=True, padx=12, pady=4)

        cols     = ("seq", "name", "qty", "unit", "unit_price", "subtotal")
        col_lbls = ("序號", "品名 / 規格", "數量", "單位", "單價", "小計")
        col_ws   = (45, 330, 65, 65, 85, 85)

        self._tree = ttk.Treeview(tf, columns=cols, show="headings",
                                   selectmode="browse", height=8)
        for col, lbl, w in zip(cols, col_lbls, col_ws):
            self._tree.heading(col, text=lbl)
            self._tree.column(col, width=w, minwidth=w, anchor="center")
        self._tree.column("name", anchor="w")

        vsb = ttk.Scrollbar(tf, orient="vertical", command=self._tree.yview)
        self._tree.configure(yscrollcommand=vsb.set)
        self._tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        self._tree.bind("<Double-1>", self._on_cell_dclick)

    # ── Tab 2：驗機單 ─────────────────────────────────────────
    def _build_tab_inspection(self, parent, PAD, FONT, FONTB, BG):
        GRAY   = "#5d6d7e"
        FONT_S = ("Microsoft JhengHei UI", 9)

        info = tk.LabelFrame(parent, text="說明", bg=BG, font=FONTB)
        info.pack(fill="x", padx=12, pady=(12, 4))
        tk.Label(info, text="載入報價單後，點擊下方按鈕自動生成驗機單 Excel 及 Word。",
                 bg=BG, font=FONT, fg=GRAY).pack(padx=12, pady=8, anchor="w")

        pf = tk.Frame(parent, bg="#e8ecf0")
        pf.pack(fill="x", padx=12, pady=4)
        tk.Label(pf, text="輸出路徑：", bg="#e8ecf0", font=FONT_S, fg=GRAY,
                 anchor="w", width=12).pack(side="left", padx=8, pady=6)
        tk.Label(pf, text="（依⚙路徑設定）",
                 bg="#e8ecf0", font=FONT_S, fg=GRAY).pack(side="left", pady=6)

        bb = tk.Frame(parent, bg=BG, pady=8)
        bb.pack(side="bottom", fill="x", padx=12)
        tk.Button(bb, text="🔍  生成驗機單", command=self._generate_inspection,
                  bg="#6c3483", fg="white", relief="flat",
                  font=("Microsoft JhengHei UI", 12, "bold"), pady=10).pack(fill="x")

    # ── Tab 3：維修單 ─────────────────────────────────────────
    def _build_tab_fix(self, parent, PAD, FONT, FONTB, BG):
        GRAY   = "#5d6d7e"
        FONT_S = ("Microsoft JhengHei UI", 9)

        info = tk.LabelFrame(parent, text="說明", bg=BG, font=FONTB)
        info.pack(fill="x", padx=12, pady=(12, 4))
        tk.Label(info, text="載入報價單後，點擊下方按鈕生成維修單。",
                 bg=BG, font=FONT, fg=GRAY).pack(padx=12, pady=8, anchor="w")

        pf = tk.Frame(parent, bg="#e8ecf0")
        pf.pack(fill="x", padx=12, pady=4)
        row = tk.Frame(pf, bg="#e8ecf0")
        row.pack(fill="x")
        tk.Label(row, text="輸出路徑：", bg="#e8ecf0", font=FONT_S,
                 fg=GRAY, anchor="w", width=10).pack(side="left", padx=8)
        tk.Label(row, text="（依⚙路徑設定）",
                 bg="#e8ecf0", font=FONT_S, fg=GRAY).pack(side="left")

        bb = tk.Frame(parent, bg=BG, pady=8)
        bb.pack(side="bottom", fill="x", padx=12)
        tk.Button(bb, text="🔧  生成維修單", command=self._generate_fix,
                  bg="#d68910", fg="white", relief="flat",
                  font=("Microsoft JhengHei UI", 12, "bold"), pady=10).pack(fill="x")

    # ── Tab 4：維修掛件 ───────────────────────────────────────
    def _build_tab_tag(self, parent, PAD, FONT, FONTB, BG):
        from tkcalendar import DateEntry
        GRAY = "#5d6d7e"

        tgf = tk.LabelFrame(parent, text="維修掛件資料", bg=BG, font=FONTB)
        tgf.pack(fill="x", padx=12, pady=(12, 4))
        tgf.columnconfigure(1, weight=1)
        tgf.columnconfigure(3, weight=1)

        self._tag_vars = {}

        # 客戶名稱
        cust_var = tk.StringVar()
        self._tag_vars["customer"] = cust_var
        tk.Label(tgf, text="客戶名稱：", bg=BG, anchor="w", font=FONT
                 ).grid(row=0, column=0, sticky="w", padx=8, pady=2)
        tk.Entry(tgf, textvariable=cust_var, font=FONT
                 ).grid(row=0, column=1, sticky="ew", padx=8, pady=2)

        def _load_customer():
            if self._parsed_data:
                cust_var.set(self._parsed_data["header"].get("customer", ""))
                part_nos = [item.get("part_no", "") or item.get("name", "")
                            for item in self._parsed_data.get("items", [])]
                self._tag_partno_cb["values"] = part_nos
                if part_nos:
                    self._tag_partno_var.set(part_nos[0])
            else:
                messagebox.showwarning("尚未載入", "請先選擇並載入報價單", parent=parent)

        tk.Button(tgf, text="從報價單帶入", command=_load_customer,
                  bg="#2e86c1", fg="white", relief="flat",
                  font=FONT, padx=6).grid(row=0, column=2, padx=8, pady=2)

        # No.
        no_var = tk.StringVar(value="1")
        self._tag_vars["no"] = no_var
        tk.Label(tgf, text="No.：", bg=BG, anchor="w", font=FONT
                 ).grid(row=1, column=0, sticky="w", padx=8, pady=2)
        ttk.Combobox(tgf, textvariable=no_var,
                     values=[str(i) for i in range(1, 21)],
                     width=8, font=FONT).grid(row=1, column=1, sticky="w", padx=8, pady=2)

        # 品號
        self._tag_partno_var = tk.StringVar()
        self._tag_vars["part_no"] = self._tag_partno_var
        tk.Label(tgf, text="品號：", bg=BG, anchor="w", font=FONT
                 ).grid(row=1, column=2, sticky="w", padx=8, pady=2)
        self._tag_partno_cb = ttk.Combobox(tgf, textvariable=self._tag_partno_var,
                                            font=FONT, width=20)
        self._tag_partno_cb.grid(row=1, column=3, sticky="ew", padx=8, pady=2)

        # 序號 / 拉回日期
        seq_var = tk.StringVar()
        self._tag_vars["seq_no"] = seq_var
        tk.Label(tgf, text="序號：", bg=BG, anchor="w", font=FONT
                 ).grid(row=2, column=0, sticky="w", padx=8, pady=2)
        tk.Entry(tgf, textvariable=seq_var, font=FONT
                 ).grid(row=2, column=1, sticky="ew", padx=8, pady=2)

        tk.Label(tgf, text="拉回：", bg=BG, anchor="w", font=FONT
                 ).grid(row=2, column=2, sticky="w", padx=8, pady=2)
        self._tag_date_entry = DateEntry(
            tgf, font=FONT, date_pattern="yyyy/mm/dd",
            background="#2e86c1", foreground="white", width=14)
        self._tag_date_entry.grid(row=2, column=3, sticky="w", padx=8, pady=2)

        # 問題 / 維修狀況
        prob_var = tk.StringVar()
        self._tag_vars["problem"] = prob_var
        tk.Label(tgf, text="問題：", bg=BG, anchor="w", font=FONT
                 ).grid(row=3, column=0, sticky="w", padx=8, pady=2)
        tk.Entry(tgf, textvariable=prob_var, font=FONT
                 ).grid(row=3, column=1, sticky="ew", padx=8, pady=2)

        status_var = tk.StringVar()
        self._tag_vars["repair_status"] = status_var
        tk.Label(tgf, text="維修狀況：", bg=BG, anchor="w", font=FONT
                 ).grid(row=3, column=2, sticky="w", padx=8, pady=2)
        tk.Entry(tgf, textvariable=status_var, font=FONT
                 ).grid(row=3, column=3, sticky="ew", padx=8, pady=2)

        # 輸出路徑
        pf = tk.Frame(parent, bg="#e8ecf0")
        pf.pack(fill="x", padx=12, pady=4)
        tk.Label(pf, text="輸出路徑：", bg="#e8ecf0",
                 font=("Microsoft JhengHei UI", 9), fg=GRAY,
                 anchor="w", width=10).pack(side="left", padx=8, pady=6)
        tk.Label(pf, text=r"Z:\待維修機台資料",
                 bg="#e8ecf0", font=("Microsoft JhengHei UI", 9), fg=GRAY
                 ).pack(side="left", pady=6)

        bb = tk.Frame(parent, bg=BG, pady=8)
        bb.pack(side="bottom", fill="x", padx=12)
        tk.Button(bb, text="📋  生成維修掛件", command=self._generate_tag_doc,
                  bg="#1a5276", fg="white", relief="flat",
                  font=("Microsoft JhengHei UI", 12, "bold"), pady=10).pack(fill="x")

    # ── Tab 5：標籤生成 ───────────────────────────────────────
    def _build_tab_label(self, parent, FONT, FONTB, BG):
        from tksheet import Sheet
        import re as _re

        # 模板選擇
        tpl_frame = tk.Frame(parent, bg=BG)
        tpl_frame.pack(fill="x", padx=12, pady=(10, 0))
        tk.Label(tpl_frame, text="標籤樣式：", bg=BG, font=FONTB).pack(side="left")
        _tpl_var = tk.StringVar(value="銀標")
        _tpl_cb  = ttk.Combobox(tpl_frame, textvariable=_tpl_var,
                                 values=["銀標", "APT標", "無公司標", "上銀標"],
                                 state="readonly", font=FONT, width=10)
        _tpl_cb.pack(side="left", padx=(8, 0))

        # 欄位索引：0=型號 1=荷重 2=序號 3=機台尺寸 4=機台重量 5=出廠年份
        #           6=供應商代碼 7=機台序號 8=訂單編號 9=收貨人
        _HEADERS = ["型號", "荷重", "序號", "機台尺寸", "機台重量", "出廠年份",
                    "供應商代碼", "機台序號", "訂單編號", "收貨人"]
        _EMPTY   = [""] * len(_HEADERS)
        _COLS    = {
            "銀標":    [0, 1, 2],
            "APT標":   [0, 1, 2, 3, 4, 5],
            "無公司標": [0, 1, 2],
            "上銀標":   [6, 7, 8, 9],
        }
        _ALL_COLS = list(range(len(_HEADERS)))

        tf = tk.LabelFrame(parent, text="標籤資料", bg=BG, font=FONTB)
        tf.pack(fill="both", expand=True, padx=12, pady=(10, 4))

        sheet = Sheet(tf,
                      headers=_HEADERS,
                      data=[_EMPTY[:] for _ in range(50)],
                      column_width=130,
                      row_height=28)
        sheet.enable_bindings()
        sheet.pack(fill="both", expand=True)

        def _on_tpl_change(*_):
            visible = _COLS.get(_tpl_var.get(), [0, 1, 2])
            hidden  = [c for c in _ALL_COLS if c not in visible]
            sheet.show_columns(_ALL_COLS)
            if hidden:
                sheet.hide_columns(hidden)

        _tpl_var.trace_add("write", _on_tpl_change)
        _on_tpl_change()

        def _re_find(text, *patterns):
            for pat in patterns:
                m = _re.search(pat, text)
                if m:
                    return m.group(1).strip()
            return ""

        def _load_from_quote():
            today  = datetime.today()
            serial = f"{today.year % 100 + 12:02d}{today.month + 12:02d}"
            year   = str(today.year)
            rows   = []
            if self._parsed_data:
                for item in self._parsed_data.get("items", []):
                    name = item.get("name", "")
                    raw_load  = _re_find(name, r'載重[：:]\s*(\S+)')
                    load_num  = _re.sub(r'[kK][gG][sS]?$', '', raw_load).strip()
                    load      = (load_num + "kgs") if load_num else ""
                    length    = _re_find(name, r'牙叉長度\s*[：: ]+(\d+(?:\.\d+)?)')
                    width     = _re_find(name, r'牙叉外寬\s*[：: ]+(\d+(?:\.\d+)?)')
                    size      = (f"{length}mm*{width}mm" if length and width
                                 else (f"{length}mm" if length else f"{width}mm" if width else ""))
                    weight_raw = _re_find(name, r'自重\s*[：:]\s*(\d+(?:\.\d+)?)')
                    weight     = (weight_raw + "kgs") if weight_raw else ""
                    rows.append([item.get("part_no", ""), load, serial,
                                 size, weight, year])
            while len(rows) < 50:
                rows.append(_EMPTY[:])
            sheet.data = rows

        _load_from_quote()

        def _autofill_serial():
            data = sheet.data
            start_row, start_val = None, ""
            for i, row in enumerate(data):
                v = str(row[2]).strip() if len(row) > 2 else ""
                if v:
                    start_row, start_val = i, v
                    break
            if start_row is None:
                messagebox.showwarning("無起始序號", "請先在製造序號欄填入第一個序號", parent=parent)
                return
            m = _re.match(r'^(.*?)(\d+)$', start_val)
            if not m:
                messagebox.showwarning("格式不符", "序號結尾需為數字，例如 SN001 或 2026001", parent=parent)
                return
            prefix, num_str = m.group(1), m.group(2)
            w = len(num_str)
            counter = int(num_str)
            for i in range(start_row, len(data)):
                has_content = str(data[i][0]).strip() or str(data[i][1]).strip()
                if i == start_row or has_content:
                    data[i][2] = prefix + str(counter).zfill(w)
                    counter += 1
            sheet.data = data

        # 操作按鈕
        bb = tk.Frame(parent, bg=BG)
        bb.pack(fill="x", padx=12, pady=(0, 4))
        tk.Button(bb, text="從報價單讀入", command=_load_from_quote,
                  bg="#2e86c1", fg="white", relief="flat",
                  font=FONT, padx=8).pack(side="left", padx=(0, 6))
        tk.Button(bb, text="＋ 新增列",
                  command=lambda: sheet.insert_rows(number=1),
                  bg="#27ae60", fg="white", relief="flat",
                  font=FONT, padx=8).pack(side="left", padx=(0, 6))
        tk.Button(bb, text="－ 刪除列",
                  command=lambda: [sheet.delete_rows(row=r)
                                   for r in sorted(sheet.get_selected_rows(), reverse=True)],
                  bg="#c0392b", fg="white", relief="flat",
                  font=FONT, padx=8).pack(side="left", padx=(0, 6))
        tk.Button(bb, text="流水號↓", command=_autofill_serial,
                  bg="#7d3c98", fg="white", relief="flat",
                  font=FONT, padx=8).pack(side="left")

        # 生成按鈕
        def _generate():
            rows = sheet.data
            def _kgs(v):
                v = str(v).strip()
                return (v + "kgs") if v and not v.lower().endswith("kgs") else v
            def _s(v): return str(v).strip()
            def _g(r, i): return _s(r[i]) if len(r) > i else ""
            tpl = _tpl_var.get()
            is_silver_top = (tpl == "上銀標")
            data_list = [
                {"型號": _s(r[0]), "荷重": _kgs(r[1]),
                 "序號": _g(r, 7) if is_silver_top else _g(r, 2),
                 "製造序號": _g(r, 2),
                 "機台尺寸": _g(r, 3), "機台重量": _kgs(_g(r, 4)),
                 "出廠年份": _g(r, 5),
                 "供應商代碼": _g(r, 6), "機台序號": _g(r, 7),
                 "訂單編號": _g(r, 8), "收貨人": _g(r, 9)}
                for r in rows if any(str(r[i]).strip() for i in _COLS.get(tpl, [0, 1, 2]) if i < len(r))
            ]
            if not data_list:
                messagebox.showwarning("無資料", "請先填入標籤資料", parent=parent)
                return
            date_tag = datetime.today().strftime("%Y%m%d%H%M%S")
            out_path = self._get_path("output_label") / f"標籤-{date_tag}.pdf"
            try:
                result = generate_labels(data_list, out_path, template_key=tpl)
                if messagebox.askyesno("生成成功",
                        f"已生成 {len(data_list)} 張標籤：\n{result}\n\n是否立即開啟？",
                        parent=parent):
                    os.startfile(str(result))
            except Exception as e:
                messagebox.showerror("生成失敗", str(e), parent=parent)

        gf = tk.Frame(parent, bg=BG, pady=8)
        gf.pack(fill="x", padx=12)
        tk.Button(gf, text="🖨  生成標籤 PDF", command=_generate,
                  bg="#1e8449", fg="white", relief="flat",
                  font=("Microsoft JhengHei UI", 12, "bold"), pady=8).pack(fill="x")

    # ── Tab 5：出貨排程 ───────────────────────────────────────
    def _build_tab_schedule(self, parent, FONT, FONTB, BG):
        from tkcalendar import DateEntry
        GRAY   = "#5d6d7e"
        FONT_S = ("Microsoft JhengHei UI", 9)

        _rows = []  # list of {"ev": dict, "location": str, "note_suffix": str, "travel_time": str}

        # 從 template_schedule.xlsx「地址工作區」C2:C29 載入地址選項
        _addr_options: list[str] = []
        try:
            import openpyxl as _oxl
            _tmpl_path = TEMPLATE_DIR / "template_schedule.xlsx"
            if _tmpl_path.exists():
                _wb = _oxl.load_workbook(str(_tmpl_path), read_only=True, data_only=True)
                if "地址" in _wb.sheetnames:
                    _ws = _wb["地址"]
                    _addr_options = [
                        str(_ws.cell(row=r, column=3).value).strip()
                        for r in range(2, 30)
                        if _ws.cell(row=r, column=3).value
                    ]
                _wb.close()
        except Exception:
            pass

        # ── Credential section ────────────────────────────
        cred_frame = tk.LabelFrame(parent, text="Timetree 登入憑證", bg=BG, font=FONTB)
        cred_frame.pack(fill="x", padx=12, pady=(12, 4))
        cred_frame.columnconfigure(1, weight=1)

        tt_cfg = self._config.get("timetree", {})
        sid_var  = tk.StringVar(value=tt_cfg.get("session_id", ""))
        csrf_var = tk.StringVar(value=tt_cfg.get("csrf_token", ""))

        tk.Label(cred_frame, text="Session ID：", bg=BG, font=FONT_S, fg=GRAY,
                 anchor="w").grid(row=0, column=0, sticky="w", padx=8, pady=3)
        sid_entry = tk.Entry(cred_frame, textvariable=sid_var, font=FONT_S, show="*")
        sid_entry.grid(row=0, column=1, sticky="ew", padx=8, pady=3)

        tk.Label(cred_frame, text="CSRF Token：", bg=BG, font=FONT_S, fg=GRAY,
                 anchor="w").grid(row=1, column=0, sticky="w", padx=8, pady=3)
        csrf_entry = tk.Entry(cred_frame, textvariable=csrf_var, font=FONT_S, show="*")
        csrf_entry.grid(row=1, column=1, sticky="ew", padx=8, pady=3)

        def _show_hide(entry, btn):
            if entry.cget("show") == "*":
                entry.config(show="")
                btn.config(text="隱藏")
            else:
                entry.config(show="*")
                btn.config(text="顯示")

        btn_show_sid  = tk.Button(cred_frame, text="顯示", font=FONT_S,
                                  command=lambda: _show_hide(sid_entry,  btn_show_sid))
        btn_show_sid.grid(row=0, column=2, padx=(0, 8), pady=3)
        btn_show_csrf = tk.Button(cred_frame, text="顯示", font=FONT_S,
                                  command=lambda: _show_hide(csrf_entry, btn_show_csrf))
        btn_show_csrf.grid(row=1, column=2, padx=(0, 8), pady=3)

        def _save_creds():
            self._config.setdefault("timetree", {})
            self._config["timetree"]["session_id"] = sid_var.get().strip()
            self._config["timetree"]["csrf_token"]  = csrf_var.get().strip()
            _save_config(self._config)
            messagebox.showinfo("已儲存", "Timetree 憑證已儲存", parent=parent)

        tk.Button(cred_frame, text="儲存憑證", command=_save_creds,
                  bg="#2e86c1", fg="white", relief="flat",
                  font=FONT, padx=8).grid(row=2, column=1, sticky="w", padx=8, pady=(4, 6))

        # ── Google Maps 設定 ──────────────────────────────
        maps_frame = tk.LabelFrame(parent, text="Google Maps（行車時間）", bg=BG, font=FONTB)
        maps_frame.pack(fill="x", padx=12, pady=(0, 4))
        maps_frame.columnconfigure(1, weight=1)

        gm_cfg     = self._config.get("google_maps", {})
        gm_key_var = tk.StringVar(value=gm_cfg.get("api_key", ""))
        gm_org_var = tk.StringVar(value=gm_cfg.get("origin",  "406臺中市北屯區水景里景南巷1-1號"))

        tk.Label(maps_frame, text="API Key：", bg=BG, font=FONT_S, fg=GRAY,
                 anchor="w").grid(row=0, column=0, sticky="w", padx=8, pady=3)
        gm_key_entry = tk.Entry(maps_frame, textvariable=gm_key_var, font=FONT_S, show="*")
        gm_key_entry.grid(row=0, column=1, sticky="ew", padx=8, pady=3)
        btn_gm_key = tk.Button(maps_frame, text="顯示", font=FONT_S,
                               command=lambda: _show_hide(gm_key_entry, btn_gm_key))
        btn_gm_key.grid(row=0, column=2, padx=(0, 8), pady=3)

        tk.Label(maps_frame, text="出發地：", bg=BG, font=FONT_S, fg=GRAY,
                 anchor="w").grid(row=1, column=0, sticky="w", padx=8, pady=3)
        tk.Entry(maps_frame, textvariable=gm_org_var, font=FONT_S
                 ).grid(row=1, column=1, columnspan=2, sticky="ew", padx=8, pady=3)

        def _save_maps_cfg():
            self._config.setdefault("google_maps", {})
            self._config["google_maps"]["api_key"] = gm_key_var.get().strip()
            self._config["google_maps"]["origin"]  = gm_org_var.get().strip()
            _save_config(self._config)
            messagebox.showinfo("已儲存", "Google Maps 設定已儲存", parent=parent)

        tk.Button(maps_frame, text="儲存設定", command=_save_maps_cfg,
                  bg="#2e86c1", fg="white", relief="flat",
                  font=FONT, padx=8).grid(row=2, column=1, sticky="w", padx=8, pady=(4, 6))

        # ── Preview section ───────────────────────────────
        prev_frame = tk.LabelFrame(parent, text="排程預覽", bg=BG, font=FONTB)
        prev_frame.pack(fill="both", expand=True, padx=12, pady=4)

        # Date + fetch row
        date_row = tk.Frame(prev_frame, bg=BG)
        date_row.pack(fill="x", padx=8, pady=(6, 4))
        tk.Label(date_row, text="日期：", bg=BG, font=FONT).pack(side="left")
        date_entry = DateEntry(date_row, font=FONT, date_pattern="yyyy/mm/dd",
                               background="#2e86c1", foreground="white", width=14)
        date_entry.pack(side="left", padx=(0, 6))

        fetch_status = tk.Label(date_row, text="", bg=BG, font=FONT_S, fg=GRAY)
        fetch_status.pack(side="left", padx=8)

        # Treeview
        tree_frame = tk.Frame(prev_frame, bg=BG)
        tree_frame.pack(fill="both", expand=True, padx=8, pady=(0, 2))

        cols = ("seq", "location", "note", "travel_time")
        tree = ttk.Treeview(tree_frame, columns=cols, show="headings",
                            height=7, selectmode="browse")
        tree.heading("seq",         text="順序")
        tree.heading("location",    text="地點")
        tree.heading("note",        text="備註")
        tree.heading("travel_time", text="行車時間")
        tree.column("seq",         width=45,  anchor="center", stretch=False)
        tree.column("location",    width=160, anchor="w",      stretch=False)
        tree.column("note",        width=200, anchor="w",      stretch=True)
        tree.column("travel_time", width=75,  anchor="center", stretch=False)
        tree.pack(side="left", fill="both", expand=True)

        sb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        sb.pack(side="right", fill="y")
        tree.configure(yscrollcommand=sb.set)

        def _display_loc(location: str) -> str:
            idx = location.find("(")
            return location[:idx].strip() if idx != -1 else location

        def _refresh_tree():
            tree.delete(*tree.get_children())
            for i, row in enumerate(_rows, 1):
                tree.insert("", "end", values=(
                    i, _display_loc(row["location"]), row["note_suffix"], row.get("travel_time", "")))

        # Row action buttons
        btn_row = tk.Frame(prev_frame, bg=BG)
        btn_row.pack(fill="x", padx=8, pady=(2, 8))

        def _move(delta):
            sel = tree.selection()
            if not sel:
                return
            idx = tree.index(sel[0])
            new_idx = idx + delta
            if 0 <= new_idx < len(_rows):
                _rows[idx], _rows[new_idx] = _rows[new_idx], _rows[idx]
                _refresh_tree()
                tree.selection_set(tree.get_children()[new_idx])

        def _delete_row():
            sel = tree.selection()
            if not sel:
                return
            _rows.pop(tree.index(sel[0]))
            _refresh_tree()

        # display name → full address 對應表
        _addr_map = {}
        for _full in _addr_options:
            _idx = _full.find("(")
            _disp = _full[:_idx].strip() if _idx != -1 else _full
            _addr_map[_disp] = _full
        _addr_display = list(_addr_map.keys())

        def _open_row_dialog(title, location="", note_suffix="", on_confirm=None):
            dlg = tk.Toplevel(parent)
            dlg.title(title)
            dlg.resizable(False, False)
            dlg.grab_set()

            tk.Label(dlg, text="地點：", font=FONT).grid(row=0, column=0, padx=10, pady=6, sticky="w")
            # 顯示用：只顯示公司名稱
            loc_var = tk.StringVar(value=_display_loc(location))
            cb = ttk.Combobox(dlg, textvariable=loc_var, font=FONT, width=26, values=_addr_display)
            cb.grid(row=0, column=1, padx=10, pady=6)

            tk.Label(dlg, text="備註：", font=FONT).grid(row=1, column=0, padx=10, pady=6, sticky="w")
            note_var = tk.StringVar(value=note_suffix)
            tk.Entry(dlg, textvariable=note_var, font=FONT, width=28
                     ).grid(row=1, column=1, padx=10, pady=6)

            def _confirm():
                typed = loc_var.get().strip()
                # 從下拉選的 → 換回完整地址；手動輸入的 → 直接使用
                full_loc = _addr_map.get(typed, typed)
                if on_confirm:
                    on_confirm(full_loc, note_var.get().strip())
                dlg.destroy()

            tk.Button(dlg, text="確認", command=_confirm,
                      bg="#1a5276", fg="white", relief="flat",
                      font=FONT, padx=10).grid(row=2, column=1, sticky="e", padx=10, pady=8)

        def _edit_row(_event=None):
            sel = tree.selection()
            if not sel:
                return
            idx = tree.index(sel[0])
            row = _rows[idx]

            def _apply(loc, note):
                _rows[idx]["location"]    = loc
                _rows[idx]["note_suffix"] = note
                _refresh_tree()

            _open_row_dialog("編輯事件", row["location"], row["note_suffix"], _apply)

        def _add_row():
            def _apply(loc, note):
                if not loc:
                    return
                _rows.append({"ev": {}, "location": loc, "note_suffix": note, "travel_time": ""})
                _refresh_tree()
                tree.selection_set(tree.get_children()[-1])

            _open_row_dialog("新增事件", on_confirm=_apply)

        tree.bind("<Double-1>", _edit_row)

        def _calc_travel():
            if not _rows:
                messagebox.showwarning("無資料", "請先抓取事件清單", parent=parent)
                return
            api_key = gm_key_var.get().strip()
            origin  = gm_org_var.get().strip()
            if not api_key:
                messagebox.showwarning("未設定", "請先填入 Google Maps API Key", parent=parent)
                return
            fetch_status.config(text="計算行車時間中…", fg=GRAY)
            parent.update_idletasks()
            try:
                _, failed = calculate_travel_times(_rows, api_key, origin)
                _refresh_tree()
                if not failed:
                    fetch_status.config(text="行車時間計算完成", fg="#1e8449")
                else:
                    detail = "\n".join(
                        f"  第{seq}站「{loc}」— {status}" for seq, loc, status in failed
                    )
                    fetch_status.config(
                        text=f"完成，{len(failed)} 筆失敗（地址找不到）", fg="#e67e22")
                    messagebox.showwarning(
                        "部分地址無法計算",
                        f"以下站點無法取得行車時間：\n{detail}\n\n"
                        "請在 Timetree 的「地點」欄位填入完整地址，或手動在備註中修改。",
                        parent=parent)
            except Exception as e:
                fetch_status.config(text=f"✘ {e}", fg="#c0392b")

        def _sort_location(south_to_north: bool):
            if not _rows:
                messagebox.showwarning("無資料", "請先抓取事件清單", parent=parent)
                return
            api_key = gm_key_var.get().strip()
            if not api_key:
                messagebox.showwarning("未設定", "請先填入 Google Maps API Key", parent=parent)
                return
            fetch_status.config(text="Geocoding 中…", fg=GRAY)
            parent.update_idletasks()
            try:
                sorted_rows, failed = sort_rows_by_location(_rows, api_key, south_to_north)
                _rows.clear()
                _rows.extend(sorted_rows)
                _refresh_tree()
                if failed:
                    names = ", ".join(r["location"].split("(")[0] for r in failed)
                    fetch_status.config(text=f"排序完成，{len(failed)} 筆無法定位：{names}", fg="#e67e22")
                else:
                    fetch_status.config(text="排序完成", fg="#1e8449")
            except Exception as e:
                fetch_status.config(text=f"✘ {e}", fg="#c0392b")

        for text, cmd, color in [
            ("↑ 上移",       lambda: _move(-1),               "#5d6d7e"),
            ("↓ 下移",       lambda: _move(1),                "#5d6d7e"),
            ("➕ 新增",      _add_row,                         "#117a65"),
            ("✏ 編輯",       _edit_row,                        "#1a5276"),
            ("🗑 刪除",      _delete_row,                      "#922b21"),
            ("🧭 南→北",    lambda: _sort_location(True),     "#1a5276"),
            ("🧭 北→南",    lambda: _sort_location(False),    "#1a5276"),
            ("📍 計算時間",  _calc_travel,                     "#6c3483"),
        ]:
            tk.Button(btn_row, text=text, command=cmd,
                      bg=color, fg="white", relief="flat",
                      font=FONT_S, padx=8, pady=3).pack(side="left", padx=(0, 6))

        def _fetch_preview():
            sid  = sid_var.get().strip()
            csrf = csrf_var.get().strip()
            if not sid or not csrf:
                messagebox.showwarning("憑證未填", "請先填入 Session ID 與 CSRF Token", parent=parent)
                return
            target = date_entry.get_date()
            fetch_status.config(text="抓取中…", fg=GRAY)
            parent.update_idletasks()
            try:
                evs = fetch_events(target, sid, csrf)
                _rows.clear()
                _rows.extend(events_to_rows(evs))
                _refresh_tree()
                fetch_status.config(text=f"找到 {len(evs)} 筆事件", fg="#1e8449")
            except Exception as e:
                fetch_status.config(text=f"✘ {e}", fg="#c0392b")

        tk.Button(date_row, text="🔍 抓取", command=_fetch_preview,
                  bg="#117a65", fg="white", relief="flat",
                  font=FONT, padx=8).pack(side="left")

        # ── Write button ──────────────────────────────────
        out_label = tk.Label(parent, text="", bg=BG, font=FONT_S, fg=GRAY,
                             anchor="w", wraplength=700)
        out_label.pack(fill="x", padx=16, pady=(2, 0))

        def _write_schedule():
            if not _rows:
                messagebox.showwarning("無資料", "請先抓取並確認事件清單", parent=parent)
                return
            sid  = sid_var.get().strip()
            csrf = csrf_var.get().strip()
            target = date_entry.get_date()
            try:
                out = generate_schedule(target, sid, csrf, rows=list(_rows),
                                        schedule_file=self._get_path("schedule_file"))
                out_label.config(text=f"✔  已寫入：{out}", fg="#1e8449")
                if messagebox.askyesno("寫入成功",
                        f"排程已寫入：\n{out}\n\n是否立即開啟？", parent=parent):
                    os.startfile(str(out))
            except Exception as e:
                out_label.config(text=f"✘  {e}", fg="#c0392b")
                messagebox.showerror("寫入失敗", str(e), parent=parent)

        bb = tk.Frame(parent, bg=BG, pady=8)
        bb.pack(fill="x", padx=12)
        tk.Button(bb, text="✅  確認寫入出貨行程表.xlsx", command=_write_schedule,
                  bg="#1a5276", fg="white", relief="flat",
                  font=("Microsoft JhengHei UI", 12, "bold"), pady=8).pack(fill="x")

    # ════════════════════════════════════════════════════════
    #  Tab 7：出貨一覽表
    # ════════════════════════════════════════════════════════
    def _build_tab_overview(self, parent, FONT, FONTB, BG):
        GRAY   = "#5d6d7e"
        FONT_S = ("Microsoft JhengHei UI", 9)

        # ── Trello 憑證 ───────────────────────────────────
        cred_frame = tk.LabelFrame(parent, text="Trello 憑證", bg=BG, font=FONTB)
        cred_frame.pack(fill="x", padx=12, pady=(12, 4))
        cred_frame.columnconfigure(1, weight=1)

        tr_cfg   = self._config.get("trello", {})
        key_var  = tk.StringVar(value=tr_cfg.get("api_key", ""))
        tok_var  = tk.StringVar(value=tr_cfg.get("token",   ""))

        tk.Label(cred_frame, text="API Key：", bg=BG, font=FONT_S, fg=GRAY,
                 anchor="w").grid(row=0, column=0, sticky="w", padx=8, pady=3)
        key_entry = tk.Entry(cred_frame, textvariable=key_var, font=FONT_S, show="*")
        key_entry.grid(row=0, column=1, sticky="ew", padx=8, pady=3)

        tk.Label(cred_frame, text="Token：", bg=BG, font=FONT_S, fg=GRAY,
                 anchor="w").grid(row=1, column=0, sticky="w", padx=8, pady=3)
        tok_entry = tk.Entry(cred_frame, textvariable=tok_var, font=FONT_S, show="*")
        tok_entry.grid(row=1, column=1, sticky="ew", padx=8, pady=3)

        def _show_hide(entry, btn):
            if entry.cget("show") == "*":
                entry.config(show=""); btn.config(text="隱藏")
            else:
                entry.config(show="*"); btn.config(text="顯示")

        btn_key = tk.Button(cred_frame, text="顯示", font=FONT_S,
                            command=lambda: _show_hide(key_entry, btn_key))
        btn_key.grid(row=0, column=2, padx=(0, 8), pady=3)
        btn_tok = tk.Button(cred_frame, text="顯示", font=FONT_S,
                            command=lambda: _show_hide(tok_entry, btn_tok))
        btn_tok.grid(row=1, column=2, padx=(0, 8), pady=3)

        def _save_trello_creds():
            self._config.setdefault("trello", {})
            self._config["trello"]["api_key"] = key_var.get().strip()
            self._config["trello"]["token"]   = tok_var.get().strip()
            _save_config(self._config)
            messagebox.showinfo("已儲存", "Trello 憑證已儲存", parent=parent)

        tk.Button(cred_frame, text="儲存憑證", command=_save_trello_creds,
                  bg="#2e86c1", fg="white", relief="flat",
                  font=FONT, padx=8).grid(row=2, column=1, sticky="w", padx=8, pady=(4, 6))

        # ── Google Sheets 設定 ────────────────────────────
        gs_frame = tk.LabelFrame(parent, text="Google Sheets", bg=BG, font=FONTB)
        gs_frame.pack(fill="x", padx=12, pady=4)

        creds_status = tk.Label(
            gs_frame,
            text="✔  credentials.json 已就緒" if _GSHEETS_CREDS_PATH.exists()
                 else "✘  找不到 credentials.json（請放到 template 資料夾）",
            bg=BG, font=FONT_S,
            fg="#1e8449" if _GSHEETS_CREDS_PATH.exists() else "#c0392b",
            anchor="w",
        )
        creds_status.pack(fill="x", padx=8, pady=6)

        token_status = tk.Label(
            gs_frame,
            text="✔  已授權（gsheets_token.json 存在）" if _GSHEETS_TOKEN_PATH.exists()
                 else "尚未授權，點「同步」時會自動開啟瀏覽器",
            bg=BG, font=FONT_S,
            fg="#1e8449" if _GSHEETS_TOKEN_PATH.exists() else GRAY,
            anchor="w",
        )
        token_status.pack(fill="x", padx=8, pady=(0, 6))

        # ── 同步按鈕與狀態 ────────────────────────────────
        out_label = tk.Label(parent, text="", bg=BG, font=FONT_S, fg=GRAY,
                             anchor="w", wraplength=700)
        out_label.pack(fill="x", padx=16, pady=(4, 0))

        def _sync():
            api_key = key_var.get().strip()
            token   = tok_var.get().strip()
            if not api_key or not token:
                messagebox.showwarning("憑證未填", "請先填入 Trello API Key 與 Token", parent=parent)
                return
            if not _GSHEETS_CREDS_PATH.exists():
                messagebox.showerror("缺少憑證", f"找不到 {_GSHEETS_CREDS_PATH}", parent=parent)
                return

            out_label.config(text="抓取 Trello 卡片中…", fg=GRAY)
            parent.update_idletasks()
            try:
                cards = fetch_po_cards(api_key, token)
                out_label.config(text=f"找到 {len(cards)} 張卡片，同步至 Google Sheets…", fg=GRAY)
                parent.update_idletasks()

                added = sync_cards(cards, _GSHEETS_CREDS_PATH,
                                   _GSHEETS_TOKEN_PATH, _SYNCED_CARDS_PATH)

                token_status.config(text="✔  已授權（gsheets_token.json 存在）", fg="#1e8449")
                if added:
                    out_label.config(text=f"✔  同步完成，新增 {added} 筆資料", fg="#1e8449")
                else:
                    out_label.config(text="✔  同步完成，無新卡片", fg="#1e8449")
            except Exception as e:
                out_label.config(text=f"✘  {e}", fg="#c0392b")
                messagebox.showerror("同步失敗", str(e), parent=parent)

        bb = tk.Frame(parent, bg=BG, pady=8)
        bb.pack(fill="x", padx=12)
        tk.Button(bb, text="🔄  同步 Trello → Google Sheets", command=_sync,
                  bg="#1a5276", fg="white", relief="flat",
                  font=("Microsoft JhengHei UI", 12, "bold"), pady=8).pack(fill="x")

    # ════════════════════════════════════════════════════════
    #  Tab 8：生產群組紀錄
    # ════════════════════════════════════════════════════════
    def _build_tab_production(self, parent, FONT, FONTB, BG):
        GRAY   = "#5d6d7e"
        FONT_S = ("Microsoft JhengHei UI", 9)

        # ── 說明 ──────────────────────────────────────────
        info = tk.LabelFrame(parent, text="說明", bg=BG, font=FONTB)
        info.pack(fill="x", padx=12, pady=(12, 4))
        tk.Label(info,
                 text="從 Trello「本周下單」抓取 2026/5/15 之後的新卡片，附加到生產群組紀錄 Excel。\n"
                      "Trello 憑證與「出貨一覽表」頁籤共用，請先在該頁籤儲存憑證。",
                 bg=BG, font=FONT_S, fg=GRAY, justify="left",
                 ).pack(padx=12, pady=8, anchor="w")

        # ── 檔案路徑 ──────────────────────────────────────
        pf = tk.Frame(parent, bg="#e8ecf0")
        pf.pack(fill="x", padx=12, pady=4)
        tk.Label(pf, text="寫入檔案：", bg="#e8ecf0", font=FONT, fg=GRAY,
                 anchor="w", width=10).pack(side="left", padx=8, pady=6)
        tk.Label(pf, text="（依⚙路徑設定）",
                 bg="#e8ecf0", font=FONT_S, fg=GRAY).pack(side="left", pady=6)

        # ── 狀態與同步按鈕 ────────────────────────────────
        out_label = tk.Label(parent, text="", bg=BG, font=FONT_S, fg=GRAY,
                             anchor="w", wraplength=700)
        out_label.pack(fill="x", padx=16, pady=(8, 0))

        def _sync():
            tr_cfg  = self._config.get("trello", {})
            api_key = tr_cfg.get("api_key", "").strip()
            token   = tr_cfg.get("token",   "").strip()
            if not api_key or not token:
                messagebox.showwarning(
                    "憑證未設定",
                    "請先至「出貨一覽表」頁籤填入並儲存 Trello 憑證",
                    parent=parent)
                return

            out_label.config(text="抓取 Trello 卡片中…", fg=GRAY)
            parent.update_idletasks()
            try:
                cards = fetch_po_cards(api_key, token)
                out_label.config(text=f"找到 {len(cards)} 張卡片，寫入 Excel 中…", fg=GRAY)
                parent.update_idletasks()

                added = sync_production(cards, _PRODUCTION_SYNCED_PATH,
                                        production_file=self._get_path("production_file"))
                if added:
                    out_label.config(text=f"✔  同步完成，新增 {added} 筆資料", fg="#1e8449")
                else:
                    out_label.config(text="✔  同步完成，無新卡片（2026/5/15 之後）", fg="#1e8449")
                if messagebox.askyesno("同步完成",
                        f"{'新增 ' + str(added) + ' 筆資料' if added else '無新卡片'}\n\n是否立即開啟生產群組紀錄？",
                        parent=parent):
                    os.startfile(str(self._get_path("production_file")))
            except Exception as e:
                out_label.config(text=f"✘  {e}", fg="#c0392b")
                messagebox.showerror("同步失敗", str(e), parent=parent)

        bb = tk.Frame(parent, bg=BG, pady=8)
        bb.pack(fill="x", padx=12)
        tk.Button(bb, text="🔄  同步 Trello → 生產群組紀錄.xlsx", command=_sync,
                  bg="#1a5276", fg="white", relief="flat",
                  font=("Microsoft JhengHei UI", 12, "bold"), pady=8).pack(fill="x")

    # ════════════════════════════════════════════════════════
    #  Tab 9：建立卡片
    # ════════════════════════════════════════════════════════
    def _build_tab_create_cards(self, parent, FONT, FONTB, BG):
        from tksheet import Sheet
        GRAY   = "#5d6d7e"
        FONT_S = ("Microsoft JhengHei UI", 9)

        _HEADERS = ["序號", "標題", "描述"]
        _EMPTY   = ["", "", ""]

        # ── 頂部：來源 Excel（左）+ 使用說明（右）──────────
        top_row = tk.Frame(parent, bg=BG)
        top_row.pack(fill="x", padx=12, pady=(12, 4))

        src_frame = tk.LabelFrame(top_row, text="來源 Excel", bg=BG, font=FONTB)
        src_frame.pack(side="left", fill="y", padx=(0, 8))
        src_frame.columnconfigure(1, weight=1)

        hint_frame = tk.LabelFrame(top_row, text="使用說明", bg=BG, font=FONTB)
        hint_frame.pack(side="left", fill="both", expand=True)
        hint_text = (
            "導入 Excel 後即可編輯資料\n"
            "可以選取多筆資料搬移刪除\n"
            "不可跳號選取"
        )
        tk.Label(hint_frame, text=hint_text, bg=BG, font=FONT_S, fg=GRAY,
                 justify="left", anchor="nw").pack(padx=10, pady=8, anchor="nw")

        tk.Label(src_frame, text="檔案路徑：", bg=BG, font=FONT_S, fg=GRAY,
                 anchor="w").grid(row=0, column=0, sticky="w", padx=8, pady=6)
        path_var = tk.StringVar()
        tk.Entry(src_frame, textvariable=path_var, font=FONT_S
                 ).grid(row=0, column=1, sticky="ew", padx=4, pady=6)

        tk.Label(src_frame, text="工作表：", bg=BG, font=FONT_S, fg=GRAY,
                 anchor="w").grid(row=1, column=0, sticky="w", padx=8, pady=4)
        sheet_var = tk.StringVar()
        sheet_cb  = ttk.Combobox(src_frame, textvariable=sheet_var, font=FONT_S,
                                  state="readonly", width=20)
        sheet_cb.grid(row=1, column=1, sticky="w", padx=4, pady=4)

        def _pick_file():
            p = filedialog.askopenfilename(
                title="選擇 Excel 檔案",
                filetypes=[("Excel 檔案", "*.xlsx *.xls"), ("所有檔案", "*.*")])
            if not p:
                return
            path_var.set(p)
            try:
                names = get_sheet_names(Path(p))
                # 「全部」放在最前面，可一次讀入所有工作表
                all_options = ["全部"] + names
                sheet_cb["values"] = all_options
                sheet_var.set(all_options[0])
            except Exception:
                sheet_cb["values"] = []
                sheet_var.set("")

        tk.Button(src_frame, text="選擇", command=_pick_file,
                  bg="#2e86c1", fg="white", relief="flat",
                  font=FONT_S, padx=6).grid(row=0, column=2, padx=(0, 4), pady=6)

        status_lbl = tk.Label(src_frame, text="", bg=BG, font=FONT_S, fg=GRAY)
        status_lbl.grid(row=2, column=1, sticky="w", padx=4, pady=(0, 4))

        # ── 序號篩選列 ────────────────────────────────────
        filter_frame = tk.Frame(parent, bg=BG)
        filter_frame.pack(fill="x", padx=12, pady=(0, 2))
        tk.Label(filter_frame, text="序號篩選：", bg=BG, font=FONT_S, fg=GRAY
                 ).pack(side="left")
        filter_var = tk.StringVar()
        tk.Entry(filter_frame, textvariable=filter_var, font=FONT_S, width=30
                 ).pack(side="left", padx=(4, 6))
        tk.Label(filter_frame, text="（逗號分隔，例如 1,3,5；空白=全部顯示）",
                 bg=BG, font=FONT_S, fg=GRAY).pack(side="left")

        # ── 可編輯 Sheet 預覽 ─────────────────────────────
        prev_frame = tk.LabelFrame(parent, text="卡片預覽（雙擊儲存格可編輯，0 筆）",
                                   bg=BG, font=FONT)
        prev_frame.pack(fill="both", expand=True, padx=12, pady=4)

        sheet = Sheet(prev_frame,
                      headers=_HEADERS,
                      data=[_EMPTY[:] for _ in range(10)],
                      column_width=200,
                      row_height=28)
        sheet.set_column_widths([70, 300, 500])
        sheet.enable_bindings()
        sheet.pack(fill="both", expand=True)

        _all_data: list[dict] = []

        def _apply_filter():
            raw = filter_var.get().strip()
            if raw:
                keys = {s.strip() for s in raw.split(",") if s.strip()}
                filtered = [c for c in _all_data if str(c["seq"]) in keys]
            else:
                filtered = _all_data
            rows = [[c["seq"], c["title"], c["desc"]] for c in filtered]
            while len(rows) < len(filtered) + 5:
                rows.append(_EMPTY[:])
            sheet.data = rows
            prev_frame.config(text=f"卡片預覽（雙擊儲存格可編輯，{len(filtered)} 筆）")

        tk.Button(filter_frame, text="篩選", command=_apply_filter,
                  bg="#2e86c1", fg="white", relief="flat",
                  font=FONT_S, padx=6).pack(side="left")
        tk.Button(filter_frame, text="清除", command=lambda: (filter_var.set(""), _apply_filter()),
                  bg="#5d6d7e", fg="white", relief="flat",
                  font=FONT_S, padx=6).pack(side="left", padx=(4, 0))

        def _load_preview():
            p = path_var.get().strip()
            if not p:
                messagebox.showwarning("未選擇檔案", "請先選擇 Excel 檔案", parent=parent)
                return
            status_lbl.config(text="讀取中…", fg=GRAY)
            parent.update_idletasks()
            try:
                selected = sheet_var.get()
                _all_data.clear()
                if selected == "全部":
                    # 取所有工作表名稱（排除「全部」偽選項）
                    all_sheets = [v for v in sheet_cb["values"] if v != "全部"]
                    for sname in all_sheets:
                        _all_data.extend(read_excel_cards(Path(p), sheet_name=sname))
                else:
                    _all_data.extend(read_excel_cards(Path(p), sheet_name=selected or None))
                filter_var.set("")
                _apply_filter()
                status_lbl.config(text=f"✔  讀取完成，共 {len(_all_data)} 筆", fg="#1e8449")
            except Exception as e:
                status_lbl.config(text=f"✘  {e}", fg="#c0392b")

        tk.Button(src_frame, text="讀取預覽", command=_load_preview,
                  bg="#117a65", fg="white", relief="flat",
                  font=FONT_S, padx=6).grid(row=0, column=3, padx=(0, 8), pady=6)

        # ── 建立按鈕 ──────────────────────────────────────
        out_label = tk.Label(parent, text="", bg=BG, font=FONT_S, fg=GRAY,
                             anchor="w", wraplength=700)
        out_label.pack(fill="x", padx=16, pady=(4, 0))

        def _create():
            rows = sheet.data
            cards = []
            for row in rows:
                title = str(row[1]).strip() if len(row) > 1 else ""
                desc  = str(row[2]).strip() if len(row) > 2 else ""
                if not title:
                    continue
                cards.append({"title": title, "desc": desc})

            if not cards:
                messagebox.showwarning("無資料", "表格中沒有填入標題的列", parent=parent)
                return
            tr_cfg  = self._config.get("trello", {})
            api_key = tr_cfg.get("api_key", "").strip()
            token   = tr_cfg.get("token",   "").strip()
            if not api_key or not token:
                messagebox.showwarning(
                    "憑證未設定",
                    "請先至「出貨一覽表」頁籤填入並儲存 Trello 憑證",
                    parent=parent)
                return
            if not messagebox.askyesno(
                    "確認建立",
                    f"即將在「0.待評估」清單建立 {len(cards)} 張卡片，確定繼續？",
                    parent=parent):
                return

            out_label.config(text="建立中…", fg=GRAY)
            parent.update_idletasks()
            try:
                created = trello_create_cards(cards, api_key, token)
                out_label.config(text=f"✔  成功建立 {created} 張卡片", fg="#1e8449")
            except Exception as e:
                out_label.config(text=f"✘  {e}", fg="#c0392b")
                messagebox.showerror("建立失敗", str(e), parent=parent)

        bb = tk.Frame(parent, bg=BG, pady=8)
        bb.pack(fill="x", padx=12)
        tk.Button(bb, text="🃏  建立全部卡片 → Trello 0.待評估", command=_create,
                  bg="#1a5276", fg="white", relief="flat",
                  font=("Microsoft JhengHei UI", 12, "bold"), pady=8).pack(fill="x")

    # ════════════════════════════════════════════════════════
    #  開檔
    # ════════════════════════════════════════════════════════
    #  Tab 10：下載卡片
    # ════════════════════════════════════════════════════════
    def _build_tab_download_cards(self, parent, FONT, FONTB, BG):
        GRAY   = "#5d6d7e"
        FONT_S = ("Microsoft JhengHei UI", 9)

        _list_map: dict[str, str] = {}   # list name → list id
        _all_cards: list[dict]    = []   # 目前預覽的卡片資料

        # ── 清單選擇 ──────────────────────────────────────
        top = tk.LabelFrame(parent, text="Trello 清單", bg=BG, font=FONTB)
        top.pack(fill="x", padx=12, pady=(12, 4))
        top.columnconfigure(1, weight=1)

        tk.Label(top, text="選擇清單：", bg=BG, font=FONT_S, fg=GRAY,
                 anchor="w").grid(row=0, column=0, sticky="w", padx=8, pady=6)

        list_var = tk.StringVar()
        list_cb  = ttk.Combobox(top, textvariable=list_var, font=FONT_S,
                                 state="readonly", width=28)
        list_cb.grid(row=0, column=1, sticky="w", padx=4, pady=6)

        status_lbl = tk.Label(top, text="", bg=BG, font=FONT_S, fg=GRAY)
        status_lbl.grid(row=0, column=3, sticky="w", padx=8, pady=6)

        def _get_creds():
            tr_cfg  = self._config.get("trello", {})
            api_key = tr_cfg.get("api_key", "").strip()
            token   = tr_cfg.get("token",   "").strip()
            if not api_key or not token:
                messagebox.showwarning("憑證未設定",
                    "請先至「出貨一覽表」頁籤填入並儲存 Trello 憑證", parent=parent)
                return None, None
            return api_key, token

        def _fetch_lists():
            api_key, token = _get_creds()
            if not api_key:
                return
            status_lbl.config(text="抓取中…", fg=GRAY)
            parent.update_idletasks()
            try:
                lists = get_board_lists(api_key, token)
                _list_map.clear()
                _list_map.update({lst["name"]: lst["id"] for lst in lists})
                names = list(_list_map.keys())
                list_cb["values"] = names
                if names:
                    list_var.set(names[0])
                status_lbl.config(text=f"找到 {len(lists)} 個清單", fg="#1e8449")
            except Exception as e:
                status_lbl.config(text=f"✘ {e}", fg="#c0392b")

        def _preview_cards():
            selected = list_var.get().strip()
            if not selected or selected not in _list_map:
                messagebox.showwarning("未選擇清單", "請先抓取清單並選擇", parent=parent)
                return
            api_key, token = _get_creds()
            if not api_key:
                return
            status_lbl.config(text="載入卡片中…", fg=GRAY)
            parent.update_idletasks()
            try:
                cards = get_list_cards(_list_map[selected], api_key, token)
                _all_cards.clear()
                _all_cards.extend(cards)
                tree.delete(*tree.get_children())
                for card in cards:
                    labels = "、".join(
                        lbl.get("name") or lbl.get("color", "")
                        for lbl in card.get("labels", []))
                    att_count = len(card.get("attachments") or [])
                    tree.insert("", "end", values=(
                        card["name"], labels, att_count))
                tree.selection_set(tree.get_children())   # 預設全選
                status_lbl.config(text=f"找到 {len(cards)} 張卡片", fg="#1e8449")
                prev_frame.config(text=f"卡片預覽（共 {len(cards)} 張，可多選）")
            except Exception as e:
                status_lbl.config(text=f"✘ {e}", fg="#c0392b")

        tk.Button(top, text="抓取清單", command=_fetch_lists,
                  bg="#2e86c1", fg="white", relief="flat",
                  font=FONT_S, padx=6).grid(row=0, column=2, padx=(0, 4), pady=6)
        tk.Button(top, text="預覽卡片", command=_preview_cards,
                  bg="#117a65", fg="white", relief="flat",
                  font=FONT_S, padx=6).grid(row=0, column=3, padx=(0, 8), pady=6)
        # 覆蓋 status_lbl 位置（改放 column=4）
        status_lbl.grid(row=0, column=4, sticky="w", padx=4, pady=6)

        # ── 卡片預覽 Treeview ─────────────────────────────
        prev_frame = tk.LabelFrame(parent, text="卡片預覽（共 0 張，可多選）",
                                   bg=BG, font=FONT)
        prev_frame.pack(fill="both", expand=True, padx=12, pady=4)

        cols = ("title", "labels", "att")
        tree = ttk.Treeview(prev_frame, columns=cols, show="headings",
                            selectmode="extended", height=8)
        tree.heading("title",  text="標題")
        tree.heading("labels", text="標籤")
        tree.heading("att",    text="附件")
        tree.column("title",  width=340, anchor="w",      stretch=True)
        tree.column("labels", width=130, anchor="w",      stretch=False)
        tree.column("att",    width=50,  anchor="center", stretch=False)

        vsb = ttk.Scrollbar(prev_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        # 全選 / 取消全選
        sel_row = tk.Frame(parent, bg=BG)
        sel_row.pack(fill="x", padx=12, pady=(2, 0))
        tk.Button(sel_row, text="全選", command=lambda: tree.selection_set(tree.get_children()),
                  bg="#5d6d7e", fg="white", relief="flat", font=FONT_S, padx=8
                  ).pack(side="left", padx=(0, 4))
        tk.Button(sel_row, text="取消全選", command=lambda: tree.selection_remove(tree.get_children()),
                  bg="#5d6d7e", fg="white", relief="flat", font=FONT_S, padx=8
                  ).pack(side="left")

        # ── 輸出路徑 ──────────────────────────────────────
        pf = tk.Frame(parent, bg="#e8ecf0")
        pf.pack(fill="x", padx=12, pady=4)
        tk.Label(pf, text="輸出資料夾：", bg="#e8ecf0", font=FONT_S, fg=GRAY,
                 anchor="w", width=12).pack(side="left", padx=8, pady=5)
        tk.Label(pf, text="（依⚙路徑設定）",
                 bg="#e8ecf0", font=FONT_S, fg=GRAY).pack(side="left", pady=5)

        # ── 狀態與下載按鈕 ────────────────────────────────
        out_label = tk.Label(parent, text="", bg=BG, font=FONT_S, fg=GRAY,
                             anchor="w", wraplength=700)
        out_label.pack(fill="x", padx=16, pady=(4, 0))

        def _download():
            sel_ids = tree.selection()
            if not sel_ids:
                messagebox.showwarning("未選擇", "請先預覽並選取要下載的卡片", parent=parent)
                return
            api_key, token = _get_creds()
            if not api_key:
                return

            indices  = [tree.index(i) for i in sel_ids]
            selected_cards = [_all_cards[i] for i in indices]
            list_name  = list_var.get().strip()
            output_dir = self._get_path("download_cards_dir") / list_name

            out_label.config(text="下載中…", fg=GRAY)
            parent.update_idletasks()

            def _progress(cur, total, name):
                out_label.config(text=f"下載中… ({cur}/{total}) {name}", fg=GRAY)
                parent.update_idletasks()

            try:
                count, failed = trello_download_cards(
                    selected_cards, output_dir, api_key, token, progress_cb=_progress)
                if failed:
                    out_label.config(
                        text=f"✔  完成 {count} 張，{len(failed)} 個附件失敗", fg="#e67e22")
                    messagebox.showwarning("部分附件失敗",
                        "以下附件下載失敗：\n" + "\n".join(failed), parent=parent)
                else:
                    out_label.config(
                        text=f"✔  成功下載 {count} 張卡片至：{output_dir}", fg="#1e8449")
                if messagebox.askyesno("下載完成",
                        f"共 {count} 張卡片\n\n是否立即開啟資料夾？", parent=parent):
                    os.startfile(str(output_dir))
            except Exception as e:
                out_label.config(text=f"✘  {e}", fg="#c0392b")
                messagebox.showerror("下載失敗", str(e), parent=parent)

        bb = tk.Frame(parent, bg=BG, pady=6)
        bb.pack(fill="x", padx=12)
        tk.Button(bb, text="⬇  下載選取的卡片", command=_download,
                  bg="#1a5276", fg="white", relief="flat",
                  font=("Microsoft JhengHei UI", 12, "bold"), pady=8).pack(fill="x")

    # ════════════════════════════════════════════════════════
    #  路徑設定 Dialog
    # ════════════════════════════════════════════════════════
    def _open_paths_dialog(self):
        BG     = "#f4f6f8"
        GRAY   = "#5d6d7e"
        FONT   = ("Microsoft JhengHei UI", 10)
        FONTB  = ("Microsoft JhengHei UI", 10, "bold")
        FONT_S = ("Microsoft JhengHei UI", 9)
        PAD    = {"padx": 8, "pady": 5}

        dlg = tk.Toplevel(self)
        dlg.title("⚙  路徑與人員設定")
        dlg.configure(bg=BG)
        dlg.resizable(True, False)
        dlg.grab_set()
        dlg.update_idletasks()
        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        dlg.geometry(f"720x560+{(sw-720)//2}+{(sh-560)//2}")

        items = [
            ("output_shipping",   "出貨單 / 維修單 輸出資料夾", False),
            ("output_inspection", "驗機單 輸出資料夾",          False),
            ("output_tag",        "維修掛件 輸出資料夾",        False),
            ("output_label",      "標籤 輸出資料夾",            False),
            ("output_quote",      "報價單 輸出資料夾",          False),
            ("download_cards_dir","下載卡片 輸出資料夾",        False),
            ("schedule_file",     "出貨行程表 .xlsx",           True),
            ("production_file",   "生產群組紀錄 .xlsx",         True),
        ]

        lf = tk.LabelFrame(dlg, text="輸出路徑設定", bg=BG, font=FONTB)
        lf.pack(fill="x", padx=16, pady=(16, 6))
        lf.columnconfigure(1, weight=1)

        path_vars: dict[str, tk.StringVar] = {}
        paths_cfg = self._config.get("paths", {})

        for i, (key, label, is_file) in enumerate(items):
            tk.Label(lf, text=label + "：", bg=BG, font=FONT_S, fg=GRAY,
                     anchor="w").grid(row=i, column=0, sticky="w", **PAD)
            var = tk.StringVar(value=paths_cfg.get(key) or self._PATH_DEFAULTS[key])
            path_vars[key] = var
            tk.Entry(lf, textvariable=var, font=FONT_S
                     ).grid(row=i, column=1, sticky="ew", padx=(0, 4), pady=5)

            def _pick(v=var, f=is_file):
                if f:
                    p = filedialog.askopenfilename(
                        filetypes=[("Excel 檔案", "*.xlsx *.xls"), ("所有檔案", "*.*")])
                else:
                    p = filedialog.askdirectory()
                if p:
                    v.set(p)

            tk.Button(lf, text="選擇", command=_pick,
                      bg="#2e86c1", fg="white", relief="flat",
                      font=FONT_S, padx=6).grid(row=i, column=2, padx=(0, 8), pady=5)

        # ── 人員設定 ─────────────────────────────────────────
        op_lf = tk.LabelFrame(dlg, text="人員設定（報價單製表人員與單號代號）",
                              bg=BG, font=FONTB)
        op_lf.pack(fill="x", padx=16, pady=(0, 6))

        # Treeview 顯示現有人員
        op_cols = ("name", "code")
        op_tree = ttk.Treeview(op_lf, columns=op_cols, show="headings",
                               selectmode="browse", height=4)
        op_tree.heading("name", text="名稱（放入 Excel 製表人欄）")
        op_tree.heading("code", text="代號（報價單號前綴）")
        op_tree.column("name", width=200, anchor="w")
        op_tree.column("code", width=120, anchor="center")
        op_tree.pack(fill="x", padx=8, pady=(6, 2))

        # 填入現有資料
        operators = self._config.get("operators", ["小皋"])
        op_codes  = self._config.get("operator_codes", {})
        for name in operators:
            op_tree.insert("", "end", values=(name, op_codes.get(name, "")))

        # 新增 / 刪除 列
        edit_row = tk.Frame(op_lf, bg=BG)
        edit_row.pack(fill="x", padx=8, pady=(0, 6))

        tk.Label(edit_row, text="名稱：", bg=BG, font=FONT_S, fg=GRAY).pack(side="left")
        name_var = tk.StringVar()
        tk.Entry(edit_row, textvariable=name_var, font=FONT_S, width=12,
                 relief="solid", borderwidth=1).pack(side="left", padx=(0, 8))

        tk.Label(edit_row, text="代號：", bg=BG, font=FONT_S, fg=GRAY).pack(side="left")
        code_var = tk.StringVar()
        tk.Entry(edit_row, textvariable=code_var, font=FONT_S, width=8,
                 relief="solid", borderwidth=1).pack(side="left", padx=(0, 8))

        tk.Label(edit_row, text="（英文字串，如 K）", bg=BG, font=FONT_S, fg=GRAY
                 ).pack(side="left", padx=(0, 12))

        def _op_add():
            n = name_var.get().strip()
            c = code_var.get().strip().upper()
            if not n or not c:
                messagebox.showwarning("欄位不完整", "請填入名稱和代號", parent=dlg)
                return
            # 更新已存在的同名項
            for item in op_tree.get_children():
                if op_tree.item(item)["values"][0] == n:
                    op_tree.item(item, values=(n, c))
                    name_var.set(""); code_var.set("")
                    return
            op_tree.insert("", "end", values=(n, c))
            name_var.set(""); code_var.set("")

        def _op_del():
            sel = op_tree.selection()
            if sel:
                op_tree.delete(sel[0])

        def _op_select(*_):
            sel = op_tree.selection()
            if sel:
                vals = op_tree.item(sel[0])["values"]
                name_var.set(vals[0])
                code_var.set(vals[1])

        op_tree.bind("<<TreeviewSelect>>", _op_select)

        tk.Button(edit_row, text="新增 / 更新", command=_op_add,
                  bg="#117a65", fg="white", relief="flat",
                  font=FONT_S, padx=8).pack(side="left", padx=(0, 4))
        tk.Button(edit_row, text="刪除選取", command=_op_del,
                  bg="#c0392b", fg="white", relief="flat",
                  font=FONT_S, padx=8).pack(side="left")

        def _reset_defaults():
            for key, var in path_vars.items():
                var.set(self._PATH_DEFAULTS[key])

        def _save():
            self._config.setdefault("paths", {})
            for key, var in path_vars.items():
                self._config["paths"][key] = var.get().strip()

            # 收集人員資料
            new_operators: list[str] = []
            new_codes:     dict[str, str] = {}
            for item in op_tree.get_children():
                vals = op_tree.item(item)["values"]
                name = str(vals[0]).strip()
                code = str(vals[1]).strip().upper()
                if name:
                    new_operators.append(name)
                    new_codes[name] = code or name[:1].upper()
            if new_operators:
                self._config["operators"]      = new_operators
                self._config["operator_codes"] = new_codes

            _save_config(self._config)
            messagebox.showinfo("已儲存", "設定已儲存", parent=dlg)
            dlg.destroy()

        bb = tk.Frame(dlg, bg=BG)
        bb.pack(fill="x", padx=16, pady=8)
        tk.Button(bb, text="還原路徑預設值", command=_reset_defaults,
                  bg="#5d6d7e", fg="white", relief="flat",
                  font=FONT, padx=10).pack(side="left", padx=(0, 8))
        tk.Button(bb, text="儲存並關閉", command=_save,
                  bg="#1a5276", fg="white", relief="flat",
                  font=FONTB, padx=16, pady=6).pack(side="left")

    # ════════════════════════════════════════════════════════
    def _open_file(self):
        path = filedialog.askopenfilename(
            title="選擇報價單",
            filetypes=[("Excel 檔案", "*.xlsx *.xls"), ("所有檔案", "*.*")])
        if not path:
            return
        self._src_path = path
        try:
            data = parse(path)
            self._parsed_data = data
            self._file_label.config(text=f"✔  已載入：{path}", fg="#1e8449")
            h = data["header"]
            for key, var in self._read_vars.items():
                var.set(h.get(key, "") or "—")
            for row_id in self._tree.get_children():
                self._tree.delete(row_id)
            for item in data["items"]:
                self._tree.insert("", "end", values=(
                    item["seq"], item.get("part_no", ""),
                    item["qty"], item["unit"], "", ""))
            self._fill_vars["sale_no"].set(h.get("quote_no", ""))
        except Exception as e:
            messagebox.showerror("讀取失敗", f"無法解析報價單：\n{e}")

    # ════════════════════════════════════════════════════════
    #  品項表操作
    # ════════════════════════════════════════════════════════
    def _on_cell_dclick(self, event):
        item_id = self._tree.identify_row(event.y)
        col_id  = self._tree.identify_column(event.x)
        if not item_id or not col_id:
            return
        col_idx  = int(col_id.replace("#", "")) - 1
        col_keys = ("seq", "name", "qty", "unit", "unit_price", "subtotal")
        col_disp = ("序號", "品名", "數量", "單位", "單價", "小計")
        old_val  = self._tree.item(item_id, "values")[col_idx]

        bbox = self._tree.bbox(item_id, col_id)
        if not bbox:
            return
        x, y, _, h = bbox

        pop = tk.Toplevel(self)
        pop.title(f"編輯「{col_disp[col_idx]}」")
        pop.geometry(f"300x80+{self.winfo_rootx()+x}+{self.winfo_rooty()+y+h}")
        pop.grab_set()

        var   = tk.StringVar(value=old_val)
        entry = tk.Entry(pop, textvariable=var, font=("Microsoft JhengHei UI", 11))
        entry.pack(fill="x", padx=10, pady=8)
        entry.select_range(0, "end")
        entry.focus()

        def save(_=None):
            vals = list(self._tree.item(item_id, "values"))
            new  = var.get()
            col_name = col_keys[col_idx]
            if col_name in ("seq", "qty", "unit_price", "subtotal"):
                if new.strip():
                    try:
                        new = float(new) if "." in new else int(new)
                    except ValueError:
                        messagebox.showwarning("格式錯誤", "此欄位請輸入數字", parent=pop)
                        return
            vals[col_idx] = new
            if col_name in ("qty", "unit_price"):
                try:
                    vals[5] = round(float(vals[2]) * float(vals[4]), 2)
                except Exception:
                    pass
            self._tree.item(item_id, values=vals)
            pop.destroy()

        entry.bind("<Return>", save)
        tk.Button(pop, text="確認", command=save,
                  bg="#2e86c1", fg="white", relief="flat").pack(pady=2)

    def _add_row(self):
        n = len(self._tree.get_children()) + 1
        self._tree.insert("", "end", values=(n, "新品項", 1, "組", 0, 0))

    def _del_row(self):
        sel = self._tree.selection()
        if not sel:
            return
        self._tree.delete(sel[0])
        for i, rid in enumerate(self._tree.get_children()):
            v = list(self._tree.item(rid, "values"))
            v[0] = i + 1
            self._tree.item(rid, values=v)

    # ════════════════════════════════════════════════════════
    #  製表人員
    # ════════════════════════════════════════════════════════
    def _add_operator(self):
        pop = tk.Toplevel(self)
        pop.title("新增製表人員")
        pop.geometry("260x80")
        pop.grab_set()
        var   = tk.StringVar()
        entry = tk.Entry(pop, textvariable=var, font=("Microsoft JhengHei UI", 11))
        entry.pack(fill="x", padx=10, pady=8)
        entry.focus()

        def save(_=None):
            name = var.get().strip()
            if not name:
                return
            if name not in self._config["operators"]:
                self._config["operators"].append(name)
                _save_config(self._config)
                self._operator_cb["values"] = self._config["operators"]
            self._operator_var.set(name)
            pop.destroy()

        entry.bind("<Return>", save)
        tk.Button(pop, text="新增", command=save,
                  bg="#27ae60", fg="white", relief="flat").pack(pady=2)

    def _del_operator(self):
        cur = self._operator_var.get()
        if not cur:
            return
        if len(self._config["operators"]) <= 1:
            messagebox.showwarning("無法刪除", "至少要保留一位製表人員")
            return
        if messagebox.askyesno("確認刪除", f"刪除「{cur}」？"):
            self._config["operators"].remove(cur)
            _save_config(self._config)
            self._operator_cb["values"] = self._config["operators"]
            self._operator_var.set(self._config["operators"][0])

    # ════════════════════════════════════════════════════════
    #  生成
    # ════════════════════════════════════════════════════════
    _PATH_DEFAULTS = {
        "output_shipping":   r"Z:\出貨單\Quoteflow_output",
        "output_inspection": r"Z:\Mika\驗收單及改造記錄單\Quoteflow_output",
        "output_tag":        r"Z:\待維修機台資料",
        "output_label":      r"Z:\出貨單\Quoteflow_output",
        "output_quote":      r"Z:\出貨單\Quoteflow_output\報價單",
        "schedule_file":     r"Z:\會計\5.出貨相關\出貨行程表.xlsx",
        "production_file":   r"Z:\會計\●使用表格\公司帳務\1.帳務資料\▲生產群組紀錄(新版)\生產群組紀錄2026(115年).xlsx",
        "download_cards_dir": r"Z:\出貨單\Quoteflow_output\下載卡片",
    }

    def _get_path(self, key: str) -> Path:
        return Path(self._config.get("paths", {}).get(key) or self._PATH_DEFAULTS[key])

    @property
    def _OUT_SHIPPING(self):   return self._get_path("output_shipping")
    @property
    def _OUT_TAG(self):        return self._get_path("output_tag")

    def _sync_header(self):
        for key, var in self._read_vars.items():
            val = var.get()
            self._parsed_data["header"][key] = "" if val == "—" else val

    @staticmethod
    def _to_num(s, default=0):
        try:
            return float(s) if "." in str(s) else int(s)
        except (ValueError, TypeError):
            return default

    def _generate(self):
        if not self._parsed_data:
            messagebox.showwarning("尚未載入", "請先選擇並載入報價單")
            return
        items = []
        for i, rid in enumerate(self._tree.get_children()):
            v = self._tree.item(rid, "values")
            items.append({"seq": i+1, "name": v[1],
                          "qty": self._to_num(v[2]), "unit": v[3],
                          "unit_price": self._to_num(v[4]),
                          "subtotal":   self._to_num(v[5])})
        self._parsed_data["items"] = items
        self._sync_header()
        extra = {
            "ship_date":      self._fill_vars["ship_date"].get(),
            "sale_no":        self._fill_vars["sale_no"].get(),
            "note":           self._fill_vars["note"].get(),
            "operator":       self._operator_var.get(),
            "invoice_choice": self._invoice_var.get(),
        }
        try:
            result = generate(self._parsed_data, extra, output_dir=self._OUT_SHIPPING)
            paths  = result if isinstance(result, list) else [result]
            msg    = "\n".join(str(p) for p in paths)
            if messagebox.askyesno("生成成功",
                    f"已生成 {len(paths)} 份出貨單：\n{msg}\n\n是否立即開啟？"):
                for p in paths:
                    os.startfile(p) if sys.platform == "win32" else subprocess.run(["open", str(p)])
        except Exception as e:
            messagebox.showerror("生成失敗", str(e))

    def _generate_inspection(self):
        if not self._parsed_data or not self._src_path:
            messagebox.showwarning("尚未載入", "請先選擇並載入報價單")
            return
        try:
            excel_path, word_paths = generate_inspection(
                self._src_path, self._parsed_data,
                output_dir=self._get_path("output_inspection"))
            msg = f"驗機單 Excel 已儲存至：\n{excel_path}"
            if word_paths:
                msg += f"\n\n驗機單 Word（共 {len(word_paths)} 份）："
                for wp in word_paths:
                    msg += f"\n  {wp.name}"
            if messagebox.askyesno("生成成功", msg + "\n\n是否立即開啟？"):
                os.startfile(excel_path)
                for wp in word_paths:
                    os.startfile(wp)
        except Exception as e:
            messagebox.showerror("生成失敗", str(e))

    def _generate_fix(self):
        if not self._parsed_data:
            messagebox.showwarning("尚未載入", "請先選擇並載入報價單")
            return
        items = []
        for i, rid in enumerate(self._tree.get_children()):
            v = self._tree.item(rid, "values")
            items.append({"seq": i+1, "name": v[1],
                          "qty": self._to_num(v[2]), "unit": v[3],
                          "unit_price": self._to_num(v[4]),
                          "subtotal":   self._to_num(v[5])})
        self._parsed_data["items"] = items
        self._sync_header()
        extra = {
            "ship_date":      self._fill_vars["ship_date"].get(),
            "sale_no":        self._fill_vars["sale_no"].get(),
            "note":           self._fill_vars["note"].get(),
            "operator":       self._operator_var.get(),
            "invoice_choice": self._invoice_var.get(),
        }
        try:
            result = generate_fix(self._parsed_data, extra, output_dir=self._OUT_SHIPPING)
            paths  = result if isinstance(result, list) else [result]
            msg = "\n".join(str(p) for p in paths)
            if messagebox.askyesno("生成成功",
                    f"已生成 {len(paths)} 份檔案：\n{msg}\n\n是否立即開啟？"):
                for p in paths:
                    os.startfile(p) if sys.platform == "win32" else subprocess.run(["open", str(p)])
        except Exception as e:
            messagebox.showerror("生成失敗", str(e))

    def _generate_tag_doc(self):
        customer = self._tag_vars["customer"].get().strip()
        if not customer:
            messagebox.showwarning("客戶名稱未填", "請填入客戶名稱或從報價單帶入")
            return
        tag_data = {
            "no":            self._tag_vars["no"].get(),
            "part_no":       self._tag_vars["part_no"].get(),
            "seq_no":        self._tag_vars["seq_no"].get(),
            "problem":       self._tag_vars["problem"].get(),
            "pullback_date": self._tag_date_entry.get_date().strftime("%Y/%m/%d"),
            "repair_status": self._tag_vars["repair_status"].get(),
        }
        data = {"header": {"customer": customer}}
        try:
            path = generate_tag(data, tag_data, output_dir=self._OUT_TAG)
            if messagebox.askyesno("生成成功",
                    f"維修掛件已生成：\n{path}\n\n是否立即開啟？"):
                os.startfile(str(path)) if sys.platform == "win32" else subprocess.run(["open", str(path)])
        except Exception as e:
            messagebox.showerror("生成失敗", str(e))


if __name__ == "__main__":
    App().mainloop()
