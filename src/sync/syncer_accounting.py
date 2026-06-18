"""
syncer_accounting.py — 會計對帳：讀取收款紀錄，比對 Trello 已出貨尚未付款，移至對帳完成
"""
import re
import requests
from pathlib import Path

_API_BASE   = "https://api.trello.com/1"
_BOARD_NAME = "物流事業部1"
_SRC_LISTS  = [
    "已出貨尚未付款",
    "3.1. 待匯尾款/待確認交期(Processing)",
    "3.2. 已匯款待出貨(Wait for shipping)",
]
_DST_LIST   = "會計對帳完成"

_SKIP_WORDS = {"轉帳存", "跨電匯", "媒體轉", "ATM轉", "FXML轉", "電匯", "匯款", "帳存"}


def _auth(api_key: str, token: str) -> dict:
    return {"key": api_key, "token": token}


def _get_board_id(api_key: str, token: str) -> str:
    resp = requests.get(
        f"{_API_BASE}/members/me/boards",
        params={**_auth(api_key, token), "fields": "name,id"},
        timeout=15,
    )
    resp.raise_for_status()
    for b in resp.json():
        if b["name"] == _BOARD_NAME:
            return b["id"]
    raise ValueError(f"找不到看板「{_BOARD_NAME}」")


def _find_list_id(api_key: str, token: str, board_id: str, name_contains: str) -> str:
    resp = requests.get(
        f"{_API_BASE}/boards/{board_id}/lists",
        params={**_auth(api_key, token), "fields": "name,id"},
        timeout=15,
    )
    resp.raise_for_status()
    for lst in resp.json():
        if name_contains in lst["name"]:
            return lst["id"]
    raise ValueError(f"找不到包含「{name_contains}」的 Trello 清單")


def _find_list_ids(api_key: str, token: str, board_id: str,
                   names_contains: list[str]) -> list[str]:
    """依序找出多個清單的 id；找不到的清單會被忽略（不中斷整體流程）。"""
    resp = requests.get(
        f"{_API_BASE}/boards/{board_id}/lists",
        params={**_auth(api_key, token), "fields": "name,id"},
        timeout=15,
    )
    resp.raise_for_status()
    lists = resp.json()
    ids = []
    for name_contains in names_contains:
        for lst in lists:
            if name_contains in lst["name"]:
                ids.append(lst["id"])
                break
    if not ids:
        raise ValueError(f"找不到任何指定的 Trello 清單：{', '.join(names_contains)}")
    return ids


def parse_payment_row(text: str) -> dict:
    """解析銀行收款明細文字，提取日期、金額與客戶名稱。"""
    text = str(text).strip()

    # 提取日期（民國格式 0115/05/07 或 115/05/07 → 轉西元）
    date_str      = ""
    date_sort_key = (0, 0, 0)
    m_date = re.match(r'^(\d{3,4})/(\d{2})/(\d{2})', text)
    if m_date:
        yr = int(m_date.group(1))
        mo = int(m_date.group(2))
        dy = int(m_date.group(3))
        if yr < 200:       # 民國轉西元
            yr += 1911
        date_str      = f"{yr}/{mo:02d}/{dy:02d}"
        date_sort_key = (yr, mo, dy)

    # 提取金額（如 165,045.00）
    m_amt = re.search(r'([\d,]+\.00)', text)
    amount_str = m_amt.group(1) if m_amt else ""
    try:
        amount = float(amount_str.replace(",", "")) if amount_str else 0.0
    except ValueError:
        amount = 0.0

    # 在金額之後提取中文客戶名稱
    company = ""
    after = text[m_amt.end():] if m_amt else text
    cleaned = re.sub(r'\d{6,}', ' ', after)
    cleaned = re.sub(r'TWD.*$', '', cleaned)
    cleaned = re.sub(r'\d{3,4}/\d{2}/\d{2}', '', cleaned)
    cleaned = re.sub(r'[\\]', ' ', cleaned)
    cjk = re.findall(r'[一-鿿]{2,}', cleaned)
    cjk = [g for g in cjk if g not in _SKIP_WORDS]
    if cjk:
        company = max(cjk, key=len)

    # Fallback：全文搜尋
    if not company:
        all_cjk = re.findall(r'[一-鿿]{2,}', text)
        useful = [g for g in all_cjk
                  if g not in _SKIP_WORDS and "轉" not in g and "匯" not in g]
        if useful:
            company = max(useful, key=len)

    return {
        "raw":            text,
        "date_str":       date_str,
        "date_sort_key":  date_sort_key,
        "amount":         amount,
        "amount_str":     amount_str,
        "company":        company,
    }


def parse_check_row(date_val, e_val: str) -> dict:
    """解析支票收款 E 欄（格式：客戶 支票 日期 到期日 票號 NT$金額），A 欄提供收到日期。"""
    text = str(e_val).strip()

    # 日期從 A 欄（收到日期）取得，可能是 datetime 物件或字串
    date_str      = ""
    date_sort_key = (0, 0, 0)
    if date_val:
        if hasattr(date_val, "year"):
            yr, mo, dy = date_val.year, date_val.month, date_val.day
        else:
            m = re.match(r'(\d{4})[/\-](\d{1,2})[/\-](\d{1,2})', str(date_val))
            yr = int(m.group(1)) if m else 0
            mo = int(m.group(2)) if m else 0
            dy = int(m.group(3)) if m else 0
        if yr > 0:
            date_str      = f"{yr}/{mo:02d}/{dy:02d}"
            date_sort_key = (yr, mo, dy)

    # 客戶名稱：「支票」之前的文字
    company = ""
    m_co = re.match(r'^(.+?)支票', text)
    if m_co:
        company = m_co.group(1).strip()

    # 金額：NT$X,XXX 或 NT$XXXX
    amount_str = ""
    amount     = 0.0
    m_amt = re.search(r'NT\$([\d,]+)', text)
    if m_amt:
        amount_str = m_amt.group(1)
        try:
            amount = float(amount_str.replace(",", ""))
        except ValueError:
            pass

    return {
        "raw":            text,
        "date_str":       date_str,
        "date_sort_key":  date_sort_key,
        "amount":         amount,
        "amount_str":     amount_str,
        "company":        company,
    }


def read_payment_records(excel_path: Path,
                          sheet_name: str = "收款紀錄") -> list[dict]:
    """
    讀取 Excel 收款紀錄或支票收款工作表，回傳解析後的列表。
    每筆記錄含 is_original_color：True 表示儲存格背景是預設原色（未標記）。
    """
    import openpyxl
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"找不到工作表「{sheet_name}」\n可用的工作表：{', '.join(wb.sheetnames)}")
    ws = wb[sheet_name]
    records = []

    if sheet_name == "支票收款":
        # A=收到日期（index 0）、E=核對狀況（index 4）
        for row in ws.iter_rows(min_row=2):
            if len(row) < 5:
                continue
            a_cell = row[0]
            e_cell = row[4]
            if not e_cell.value:
                continue
            fill = e_cell.fill
            is_original = (fill.fill_type is None or fill.fill_type == "none")
            parsed = parse_check_row(a_cell.value, str(e_cell.value))
            if parsed["amount"] > 0:
                parsed["is_original_color"] = is_original
                records.append(parsed)
    else:
        # 收款紀錄：D 欄（index 3）
        for row in ws.iter_rows(min_row=2):
            if len(row) < 4:
                continue
            d_cell = row[3]
            if not d_cell.value:
                continue
            fill = d_cell.fill
            is_original = (fill.fill_type is None or fill.fill_type == "none")
            parsed = parse_payment_row(str(d_cell.value))
            if parsed["amount"] > 0:
                parsed["is_original_color"] = is_original
                records.append(parsed)

    return records


def _fetch_trello_cards(list_id: str, api_key: str, token: str) -> list[dict]:
    """抓取指定清單的所有卡片，解析描述中的應收總金額。"""
    resp = requests.get(
        f"{_API_BASE}/lists/{list_id}/cards",
        params={**_auth(api_key, token), "fields": "id,name,desc"},
        timeout=15,
    )
    resp.raise_for_status()
    result = []
    for card in resp.json():
        desc = card.get("desc", "")
        m = re.search(r'應收總金額[：:]\s*([\d,，\.]+)', desc)
        card_amount = 0.0
        if m:
            try:
                card_amount = float(
                    m.group(1).replace(",", "").replace("，", "").split(".")[0])
            except ValueError:
                pass
        result.append({
            "id":     card["id"],
            "title":  card["name"],
            "amount": card_amount,
        })
    return result


def _match_payment(payment: dict, cards: list[dict]) -> dict | None:
    """根據客戶名稱 + 金額比對 Trello 卡片，回傳第一個匹配的卡片或 None。"""
    company = payment.get("company", "").strip()
    amount  = payment.get("amount", 0.0)
    if not company:
        return None

    # 嚴格比對：名稱 + 金額（容差 30 元）
    for card in cards:
        if company in card["title"] and abs(card["amount"] - amount) < 30.0:
            return card

    # 寬鬆比對：名稱符合，卡片金額未填
    for card in cards:
        if company in card["title"] and card["amount"] == 0.0:
            return card

    return None


def preview_matches(payments: list[dict],
                    api_key: str,
                    token: str) -> dict:
    """
    比對付款紀錄與 Trello 多個來源清單（_SRC_LISTS），不執行移動。
    回傳 {"matched": [(payment, card)], "unmatched": [payment], "dst_id": str}
    """
    board_id     = _get_board_id(api_key, token)
    src_ids      = _find_list_ids(api_key, token, board_id, _SRC_LISTS)
    dst_id       = _find_list_id(api_key, token, board_id, _DST_LIST)
    trello_cards = []
    for src_id in src_ids:
        trello_cards.extend(_fetch_trello_cards(src_id, api_key, token))

    matched:   list[tuple] = []
    unmatched: list[dict]  = []

    for payment in payments:
        card = _match_payment(payment, trello_cards)
        if card:
            matched.append((payment, card))
        else:
            unmatched.append(payment)

    return {
        "matched":   matched,
        "unmatched": unmatched,
        "dst_id":    dst_id,
    }


def execute_moves(matched: list[tuple], dst_id: str,
                  api_key: str, token: str) -> int:
    """將比對成功的卡片移至會計對帳完成清單，回傳移動數量。"""
    count = 0
    for _, card in matched:
        resp = requests.put(
            f"{_API_BASE}/cards/{card['id']}",
            params={**_auth(api_key, token), "idList": dst_id},
            timeout=15,
        )
        resp.raise_for_status()
        count += 1
    return count
