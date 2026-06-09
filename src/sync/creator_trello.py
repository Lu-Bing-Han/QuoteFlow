"""
creator_trello.py — 從 Excel 讀取資料並在 Trello 建立卡片
"""
from pathlib import Path
import re

import openpyxl
import requests

_API_BASE   = "https://api.trello.com/1"
_BOARD_NAME = "物流事業部1"
_LIST_NAME  = "0.待評估"

_REPAIR_BOARD = "維修保養部門"
_REPAIR_LIST  = "0.1. 待評估"


def _auth(api_key: str, token: str) -> dict:
    return {"key": api_key, "token": token}


def _get_target_list_id(api_key: str, token: str,
                         board_name: str = _BOARD_NAME,
                         list_name: str = _LIST_NAME) -> str:
    resp = requests.get(
        f"{_API_BASE}/members/me/boards",
        params={**_auth(api_key, token), "fields": "name,id"},
        timeout=15,
    )
    resp.raise_for_status()
    board_id = None
    for b in resp.json():
        if b["name"] == board_name:
            board_id = b["id"]
            break
    if not board_id:
        raise ValueError(f"找不到看板「{board_name}」")

    resp = requests.get(
        f"{_API_BASE}/boards/{board_id}/lists",
        params={**_auth(api_key, token), "fields": "name,id"},
        timeout=15,
    )
    resp.raise_for_status()
    for lst in resp.json():
        if list_name in lst["name"]:
            return lst["id"]
    raise ValueError(f"找不到清單「{list_name}」")


def _parse_phones(cell_value) -> tuple[str, str, str]:
    """從儲存格解析手機、市話、分機，回傳 (mobile, phone, ext)。
    - 手機：09 開頭、10碼
    - 市話：0[2-8] 開頭（含區碼）
    - 分機：# 後的數字，附屬於同一段市話（例如 04-9225-7851 #160）
    空格不作為分隔符，避免 "04-9225-7851 #160" 被切斷。
    """
    if not cell_value:
        return "", "", ""
    mobile, phone, ext = "", "", ""
    for part in re.split(r'[\n\r/,、]+', str(cell_value).strip()):
        part = part.strip()
        if not part:
            continue
        # 分離分機（# 之後）
        if '#' in part:
            base, ext_part = part.split('#', 1)
            base = base.strip()
            ext_part = ext_part.strip()
        else:
            base = part
            ext_part = ""
        digits = re.sub(r'[^0-9]', '', base)
        if re.match(r'^09', digits) and len(digits) == 10 and not mobile:
            mobile = base
        elif re.match(r'^0[2-8]', digits) and not phone:
            phone = base
            if ext_part and not ext:
                ext = ext_part
    return mobile, phone, ext


def get_sheet_names(excel_path: Path) -> list[str]:
    """回傳 Excel 中所有工作表名稱。"""
    wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def read_excel_cards(excel_path: Path, sheet_name: str | None = None) -> list[dict]:
    """讀 Excel：B欄=標題，C=公司名，D=聯絡人，F=電話/手機，G=傳真，H=地址，
    I=電子信箱，J=統編，K=手寫備註/產品需求。

    回傳 list of {"row": int, "seq": str, "title": str, "desc": str, "notes": str}
    """
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    cards = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        b_val = row[1].value  # B欄 — 標題
        if not b_val:
            continue
        title = str(b_val).strip()
        if not title:
            continue

        a_val = str(row[0].value or "").strip()                            # A — 序號
        c_val = str(row[2].value or "").strip()                            # C — 公司名
        d_val = str(row[3].value or "").strip()                            # D — 聯絡人
        f_val = row[5].value                                               # F — 電話/手機
        g_val = str(row[6].value or "").strip()                            # G — 傳真
        h_val = str(row[7].value or "").strip()                            # H — 地址
        i_val = str(row[8].value or "").strip()                            # I — 電子信箱
        j_val = str(row[9].value or "").strip()                            # J — 統一編號
        k_val = str(row[10].value or "").strip() if len(row) > 10 else ""  # K — 備註/需求

        mobile, phone, ext = _parse_phones(f_val)
        phone_display = f"{phone}#{ext}" if ext else phone

        desc = (
            f"公司名：{c_val}\n"
            f"聯絡人：{d_val}\n"
            f"手機：{mobile}\n"
            f"電話：{phone_display}\n"
            f"傳真：{g_val}\n"
            f"電子信箱：{i_val}\n"
            f"統一編號：{j_val}\n"
            f"地址：{h_val}"
        )
        cards.append({"row": row_idx, "seq": a_val, "title": title,
                      "desc": desc, "notes": k_val})

    wb.close()
    return cards


def create_cards(cards: list[dict], api_key: str, token: str,
                  board_name: str = _BOARD_NAME,
                  list_name: str = _LIST_NAME) -> int:
    """建立 Trello 卡片，回傳成功建立筆數。"""
    list_id = _get_target_list_id(api_key, token, board_name, list_name)
    created = 0
    for card in cards:
        resp = requests.post(
            f"{_API_BASE}/cards",
            params={
                **_auth(api_key, token),
                "idList": list_id,
                "name":   card["title"],
                "desc":   card["desc"],
                "pos":    "top",
            },
            timeout=15,
        )
        resp.raise_for_status()
        created += 1
    return created
