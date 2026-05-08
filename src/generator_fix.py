"""
generator_fix.py — 填入維修單模板，輸出 xlsx
與 generator.py 相同邏輯，標題與檔名改為「維修單」
"""

import shutil
from copy import copy
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.cell import MergedCell

TEMPLATE_PATH = Path(__file__).parent.parent / "template" / "template_fix.xlsx"
OUTPUT_DIR    = Path(__file__).parent.parent / "output"

ITEM_ROW     = 10
FOOTER_START = 11


def _format_date(s):
    try:
        dt = datetime.strptime(s, "%Y/%m/%d")
        return f"{dt.year} 年  {dt.month:02d} 月  {dt.day:02d}  日"
    except Exception:
        return s


def _safe_write(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        cell.value = value


def _copy_row_style(ws, src_row, dst_row):
    for col in range(1, 11):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        if isinstance(dst, MergedCell):
            continue
        if src.has_style:
            dst.font          = copy(src.font)
            dst.border        = copy(src.border)
            dst.fill          = copy(src.fill)
            dst.number_format = src.number_format
            dst.alignment     = copy(src.alignment)
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def _shift_footer_merges(ws, n_extra):
    to_shift = [
        (r.min_col, r.min_row, r.max_col, r.max_row)
        for r in ws.merged_cells.ranges
        if r.min_row >= FOOTER_START
    ]
    for (c1, r1, c2, r2) in to_shift:
        ws.unmerge_cells(start_row=r1, start_column=c1,
                         end_row=r2,   end_column=c2)
    for (c1, r1, c2, r2) in to_shift:
        ws.merge_cells(start_row=r1 + n_extra, start_column=c1,
                       end_row=r2   + n_extra, end_column=c2)


def _append_invoice_to_footer(ws, invoice_choice):
    if invoice_choice == '隨貨':
        suffix = '    ■發票隨貨 □發票直寄'
    elif invoice_choice == '直寄':
        suffix = '    □發票隨貨 ■發票直寄'
    else:
        suffix = '    □發票隨貨 □發票直寄'

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell) or cell.value is None:
                continue
            if '第一聯' in str(cell.value) and '白聯' in str(cell.value):
                cell.value = str(cell.value) + suffix
                return


def generate_fix(data, extra, out_filename=""):
    OUTPUT_DIR.mkdir(exist_ok=True)
    tmp = OUTPUT_DIR / "_working_fix.xlsx"
    shutil.copy(TEMPLATE_PATH, tmp)

    wb = openpyxl.load_workbook(tmp)
    ws = wb.active

    h       = data["header"]
    items   = data["items"]
    n       = len(items)
    n_extra = n - 1

    # ── 1. 表頭 ───────────────────────────────────────────────
    ws["A5"] = "維修日期：" + _format_date(extra.get("ship_date", ""))
    _safe_write(ws, 6, 3, h.get("customer", ""))
    _safe_write(ws, 7, 3, h.get("phone", ""))
    _safe_write(ws, 8, 3, h.get("address", ""))
    ws["F7"] = "聯絡人：" + h.get("contact", "")

    sale_no = extra.get("sale_no", "").strip()
    if sale_no:
        ws["I6"] = "銷貨單號：" + sale_no

    tax_id = h.get("tax_id", "").strip()
    if tax_id:
        ws["F6"] = "統一編號：" + tax_id

    # ── 2. 插入多餘列 ─────────────────────────────────────────
    if n_extra > 0:
        _shift_footer_merges(ws, n_extra)
        ws.insert_rows(FOOTER_START, amount=n_extra)

    # ── 3. 品項 ───────────────────────────────────────────────
    for i, item in enumerate(items):
        r = ITEM_ROW + i
        if i > 0:
            ws.merge_cells(f"A{r}:B{r}")
            ws.merge_cells(f"C{r}:E{r}")
            _copy_row_style(ws, ITEM_ROW, r)
        _safe_write(ws, r, 1, i + 1)
        _safe_write(ws, r, 3, item["name"].replace("\n", " "))
        _safe_write(ws, r, 6, item["qty"])
        _safe_write(ws, r, 7, item["unit"])
        _safe_write(ws, r, 8, item["unit_price"])
        _safe_write(ws, r, 9, item["subtotal"])

    # ── 4. Footer ─────────────────────────────────────────────
    note_row     = ITEM_ROW + n
    operator_row = note_row + 1

    note_text = extra.get("note", "").strip()
    if note_text:
        _safe_write(ws, note_row, 2, note_text)
    _safe_write(ws, operator_row, 3, extra.get("operator", ""))

    # ── 5. 發票方式（追加到※第一聯那行結尾）──────────────────────
    _append_invoice_to_footer(ws, extra.get("invoice_choice", "尚未確認"))

    # ── 6. 存檔 ───────────────────────────────────────────────
    if not out_filename:
        customer = h.get("customer", "客戶")
        date_tag = extra.get("ship_date", "").replace("/", "") or datetime.today().strftime("%Y%m%d")
        out_filename = f"維修單-{customer}-{date_tag}.xlsx"

    out_path = OUTPUT_DIR / out_filename
    wb.save(out_path)
    tmp.unlink(missing_ok=True)
    return out_path


if __name__ == "__main__":
    import sys
    sys.path.insert(0, str(Path(__file__).parent))
    from parser import parse

    if len(sys.argv) < 2:
        print("用法：python generator_fix.py 報價單.xlsx")
        sys.exit(1)

    data  = parse(sys.argv[1])
    extra = {
        "ship_date":      datetime.today().strftime("%Y/%m/%d"),
        "sale_no":        "",
        "operator":       "小皋",
        "note":           "",
        "invoice_choice": "隨貨",
    }
    out = generate_fix(data, extra)
    print(f"維修單已輸出：{out}")
