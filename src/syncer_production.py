"""
syncer_production.py — 將 Trello 卡片同步到生產群組紀錄 Excel
"""
import json
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font

PRODUCTION_FILE = Path(
    r"Z:\會計\●使用表格\公司帳務\1.帳務資料\▲生產群組紀錄(新版)\生產群組紀錄2026(115年).xlsx"
)
_CUT_DATE = date(2026, 5, 15)
_FONT     = Font(name="Microsoft JhengHei UI")
_CENTER   = Alignment(horizontal="center", vertical="center")
_RIGHT    = Alignment(horizontal="right",  vertical="center")


def _fmt_date(date_str: str) -> str:
    """'M/D' → 'M月D日'；非此格式原樣回傳。"""
    if "/" in date_str:
        parts = date_str.split("/", 1)
        if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
            return f"{int(parts[0])}月{int(parts[1])}日"
    return date_str


def _to_number(s: str):
    """字串轉數字；去除千分位逗號。無法轉換則原樣回傳。"""
    cleaned = str(s).replace(",", "").strip()
    try:
        v = float(cleaned)
        return int(v) if v == int(v) else v
    except (ValueError, TypeError):
        return s


def _card_to_row(card: dict) -> list:
    delivery = _fmt_date(card.get("delivery", "").replace("前", "").strip())
    return [
        "",                                       # A：業務（不填）
        _fmt_date(card["created_date"]),          # B：下單日期
        card["company"],                          # C：客戶名稱
        card["product"],                          # D：品號
        _to_number(card["quantity"]),             # E：數量（數字）
        delivery,                                 # F：出貨日期
        "",                                       # G：出貨單
        "",                                       # H：發票
        "",                                       # I：空欄
        card.get("payment_raw", ""),              # J：付款條件
        _to_number(card.get("amount", "")),       # K：綜額(含稅)（數字）
    ]


def _read_synced_ids(synced_path: Path) -> set:
    if synced_path.exists():
        return set(json.loads(synced_path.read_text(encoding="utf-8")))
    return set()


def _save_synced_ids(synced_path: Path, ids: set):
    synced_path.write_text(
        json.dumps(sorted(ids), ensure_ascii=False, indent=2), encoding="utf-8"
    )


def sync_production(cards: list[dict], synced_path: Path) -> int:
    """將 2026/5/15 之後且尚未同步的卡片附加到生產群組紀錄 Excel。"""
    synced_ids = _read_synced_ids(synced_path)

    new_cards = [
        c for c in cards
        if c["card_id"] not in synced_ids
        and (c.get("created_dt") or date.min) >= _CUT_DATE
    ]
    if not new_cards:
        return 0

    # 按下單日期由早到晚排序
    new_cards.sort(key=lambda c: c.get("created_dt") or date.max)

    wb = openpyxl.load_workbook(str(PRODUCTION_FILE))
    ws = wb.active
    for card in new_cards:
        ws.append(_card_to_row(card))
        r = ws.max_row
        for col in range(1, 12):
            ws.cell(r, col).font = _FONT
        ws.cell(r, 2).alignment  = _CENTER  # B 下單日期
        ws.cell(r, 5).alignment  = _CENTER  # E 數量
        ws.cell(r, 11).alignment = _RIGHT   # K 綜額
    wb.save(str(PRODUCTION_FILE))
    wb.close()

    synced_ids.update(c["card_id"] for c in new_cards)
    _save_synced_ids(synced_path, synced_ids)
    return len(new_cards)
