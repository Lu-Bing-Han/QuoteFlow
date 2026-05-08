"""
generator.py — 填入出貨單模板，輸出 xlsx
"""

import shutil
from copy import copy
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.cell import MergedCell

TEMPLATE_PATH = Path(__file__).parent.parent / "template" / "template.xlsx"
OUTPUT_DIR    = Path(__file__).parent.parent / "output"

ITEM_ROW     = 8
FOOTER_START = 9   # 附註列（insert 插入點，也是 footer 合併格的起始列）


def _format_date(s):
    try:
        dt = datetime.strptime(s, "%Y/%m/%d")
        return f"{dt.year} 年  {dt.month:02d} 月  {dt.day:02d}  日"
    except Exception:
        return s


def _safe_write(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        cell.value = value


def _copy_row_style(ws, src_row, dst_row):
    for col in range(1, 11):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        if isinstance(dst, MergedCell):
            continue
        if src.has_style:
            dst.font          = copy(src.font)
            dst.border        = copy(src.border)
            dst.fill          = copy(src.fill)
            dst.number_format = src.number_format
            dst.alignment     = copy(src.alignment)
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def _shift_footer_merges(ws, n_extra):
    """
    insert_rows 不會自動移動在插入點「之後」的合併格。
    手動處理：找出所有 min_row >= FOOTER_START 的合併格，
    unmerge → 重新以 (row + n_extra) merge 回去。
    """
    to_shift = [
        (r.min_col, r.min_row, r.max_col, r.max_row)
        for r in ws.merged_cells.ranges
        if r.min_row >= FOOTER_START
    ]

    for (c1, r1, c2, r2) in to_shift:
        ws.unmerge_cells(
            start_row=r1, start_column=c1,
            end_row=r2,   end_column=c2
        )

    for (c1, r1, c2, r2) in to_shift:
        ws.merge_cells(
            start_row=r1 + n_extra, start_column=c1,
            end_row=r2   + n_extra, end_column=c2
        )


def _append_invoice_to_footer(ws, invoice_choice):
    if invoice_choice == '隨貨':
        suffix = '    ■發票隨貨 □發票直寄'
    elif invoice_choice == '直寄':
        suffix = '    □發票隨貨 ■發票直寄'
    else:
        suffix = '    □發票隨貨 □發票直寄'

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell) or cell.value is None:
                continue
            if '第一聯' in str(cell.value) and '白聯' in str(cell.value):
                cell.value = str(cell.value) + suffix
                return


def generate(data, extra, out_filename=""):
    OUTPUT_DIR.mkdir(exist_ok=True)
    tmp = OUTPUT_DIR / "_working.xlsx"
    shutil.copy(TEMPLATE_PATH, tmp)

    wb = openpyxl.load_workbook(tmp)
    ws = wb.active

    h     = data["header"]
    items = data["items"]
    n     = len(items)
    n_extra = n - 1   # 需插入的額外列數

    # ── 1. 固定表頭（幣別 I5 不動）────────────────────────────
    ws["A3"] = "出貨日期：" + _format_date(extra.get("ship_date", ""))
    _safe_write(ws, 4, 3, h.get("customer", ""))
    _safe_write(ws, 5, 3, h.get("phone", ""))
    _safe_write(ws, 6, 3, h.get("address", ""))   # C6 地址
    ws["F5"] = "聯絡人：" + h.get("contact", "")

    sale_no = extra.get("sale_no", "").strip()
    if sale_no:
        ws["I4"] = "銷貨單號：" + sale_no

    tax_id = h.get("tax_id", "").strip()
    if tax_id:
        ws["F4"] = "統一編號：" + tax_id

    # ── 2. 多品項時：先移動 footer 合併格，再 insert ──────────
    if n_extra > 0:
        _shift_footer_merges(ws, n_extra)   # ← 先移，才不會被 insert 搞亂
        ws.insert_rows(FOOTER_START, amount=n_extra)

    # ── 3. 寫入各品項 ─────────────────────────────────────────
    for i, item in enumerate(items):
        r = ITEM_ROW + i
        if i > 0:
            ws.merge_cells(f"A{r}:B{r}")
            ws.merge_cells(f"C{r}:E{r}")
            _copy_row_style(ws, ITEM_ROW, r)

        _safe_write(ws, r, 1, i + 1)
        _safe_write(ws, r, 3, item["name"].replace("\n", " "))
        _safe_write(ws, r, 6, item["qty"])
        _safe_write(ws, r, 7, item["unit"])
        _safe_write(ws, r, 8, item["unit_price"])
        _safe_write(ws, r, 9, item["subtotal"])

    # ── 4. Footer 補填 ─────────────────────────────────────────
    note_row     = ITEM_ROW + n
    operator_row = note_row + 1

    note_text = extra.get("note", "").strip()
    if note_text:
        _safe_write(ws, note_row, 2, note_text)

    _safe_write(ws, operator_row, 3, extra.get("operator", ""))

    # ── 5. 發票方式（追加到※第一聯那行結尾）──────────────────────
    _append_invoice_to_footer(ws, extra.get("invoice_choice", "尚未確認"))

    # ── 6. 存檔 ────────────────────────────────────────────────
    if not out_filename:
        customer = h.get("customer", "客戶")
        date_tag = extra.get("ship_date", "").replace("/", "") or datetime.today().strftime("%Y%m%d")
        out_filename = f"出貨單-{customer}-{date_tag}.xlsx"

    out_path = OUTPUT_DIR / out_filename
    wb.save(out_path)
    tmp.unlink(missing_ok=True)
    return out_path


if __name__ == "__main__":
    import sys
    sys.path.insert(0, str(Path(__file__).parent))
    from parser import parse

    data  = parse(sys.argv[1])
    extra = {
        "ship_date":      datetime.today().strftime("%Y/%m/%d"),
        "sale_no":        "TEST-001",
        "operator":       "小皋",
        "note":           "",
        "invoice_choice": "隨貨",
    }
    out = generate(data, extra)
    print(f"輸出：{out}")
