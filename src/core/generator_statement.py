"""
generator_statement.py — 依 Trello「本周下單」卡片填入對帳單模板，輸出 xlsx
"""
import re
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.cell import MergedCell

from _paths import TEMPLATE_DIR, OUTPUT_DIR
TEMPLATE_PATH = TEMPLATE_DIR / "template_statement.xlsx"

TITLE_ROW  = 6
HEADER_ROW = 11
ITEM_ROW   = 12   # 模板保留 12~14 三列供品項使用（總金額為 SUM 公式）


def _safe_write(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        cell.value = value


def _append_to_label(ws, row, col, value):
    """保留模板原有的標籤文字（含其間距），於後方補上資料。"""
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    label = cell.value or ""
    cell.value = f"{label}{value}"


def parse_amount(raw) -> float:
    """從「應收總金額」之類的原始字串解析數字金額，無法解析則回傳 0。"""
    s = re.sub(r'[^\d.]', '', str(raw or ''))
    try:
        return float(s) if s else 0.0
    except ValueError:
        return 0.0


def parse_paid_amount(payment_raw: str, total_amount: float = 0.0) -> float:
    """從付款方式原始文字抓出已收金額。
    優先比對「訂金50%」之類的百分比寫法（= 總金額 × 百分比），
    沒有百分比才比對「已付/已收/訂金/預收 + 金額」的絕對數字寫法。
    抓不到則回傳 0。
    """
    text = payment_raw or ''
    m = re.search(r'(?:訂金|已付|已收|預收)[^\d%]*(\d+(?:\.\d+)?)\s*%', text)
    if m:
        return round(total_amount * parse_amount(m.group(1)) / 100, 2)
    m = re.search(r'(?:已付|已收|訂金|預收)[^\d]*([\d,]+(?:\.\d+)?)', text)
    return parse_amount(m.group(1)) if m else 0.0


def split_product_quantity(product: str, fallback_quantity=""):
    """若品名中含「*數字」（如「電池*5」），取出數量並從品名移除；否則使用 fallback_quantity。"""
    m = re.search(r'\*\s*(\d+)', product or '')
    if not m:
        return product, fallback_quantity
    name = (product[:m.start()] + product[m.end():]).strip()
    return name, m.group(1)


def generate_statement(card: dict, location_lookup: dict | None = None,
                        output_dir=None, gemini_paid_amount: float | None = None) -> Path:
    """依單張 Trello 卡片資料產出對帳單，回傳輸出檔案路徑。
    gemini_paid_amount：若有提供（由呼叫端先呼叫 Gemini 讀留言判斷），則優先採用其結果，
    否則回退到從付款方式文字解析的已收金額。
    """
    out_dir = Path(output_dir) if output_dir else OUTPUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active

    title_company = card.get("company", "")
    company       = card.get("company_desc", "") or title_company
    address = (
        card.get("address", "")
        or (location_lookup or {}).get(company, "")
        or (location_lookup or {}).get(title_company, "")
        or card.get("location", "")
    )

    total_amount = parse_amount(card.get("amount", ""))
    paid_amount  = (gemini_paid_amount if gemini_paid_amount is not None
                    else parse_paid_amount(card.get("payment_raw", ""), total_amount))
    due_amount   = total_amount - paid_amount

    product, quantity = split_product_quantity(card.get("product", ""), card.get("quantity", "") or "")
    unit_price = total_amount / float(quantity) if str(quantity).replace('.', '', 1).isdigit() and float(quantity) else ""

    today = datetime.today()
    title_cell = ws.cell(row=TITLE_ROW, column=1)
    if not isinstance(title_cell, MergedCell) and title_cell.value:
        title_cell.value = f"{title_cell.value.rstrip()}  {today.year}.{today.month:02d}"

    _append_to_label(ws, HEADER_ROW - 3, 1, company)
    _append_to_label(ws, HEADER_ROW - 2, 1, address)
    _append_to_label(ws, HEADER_ROW - 1, 1, card.get("payment_raw", ""))

    _append_to_label(ws, HEADER_ROW - 3, 6, card.get("contact", ""))
    _append_to_label(ws, HEADER_ROW - 2, 6, card.get("phone", ""))
    _append_to_label(ws, HEADER_ROW - 1, 6, card.get("fax", ""))

    _safe_write(ws, ITEM_ROW, 2, product)
    _safe_write(ws, ITEM_ROW, 3, quantity)
    _safe_write(ws, ITEM_ROW, 4, unit_price)
    _safe_write(ws, ITEM_ROW, 5, total_amount)
    _safe_write(ws, ITEM_ROW, 6, paid_amount)
    _safe_write(ws, ITEM_ROW, 7, due_amount)
    _safe_write(ws, ITEM_ROW, 8, card.get("card_url", ""))

    date_tag = datetime.today().strftime("%Y%m%d")
    safe_company = re.sub(r'[\\/:*?"<>|]', '_', company) or "客戶"
    out_path = out_dir / f"對帳單-{safe_company}-{date_tag}.xlsx"
    wb.save(out_path)
    return out_path
