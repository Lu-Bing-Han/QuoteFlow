"""
generator.py — 填入出貨單模板，輸出 xlsx
"""

import shutil
from copy import copy
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.cell import MergedCell

from _paths import TEMPLATE_DIR, OUTPUT_DIR
TEMPLATE_PATH = TEMPLATE_DIR / "template.xlsx"
SERIES_PATH   = TEMPLATE_DIR / "template_series.xlsx"

ITEM_ROW     = 10
FOOTER_START = 11   # 附註列（insert 插入點，也是 footer 合併格的起始列）


def _load_series_lookup() -> dict:
    """掃描 template_series.xlsx 所有分頁，建立 品號(B欄) → 品名/規格(C欄) 對照表。"""
    lookup = {}
    if not SERIES_PATH.exists():
        return lookup
    wb = openpyxl.load_workbook(SERIES_PATH, data_only=True, read_only=True)
    try:
        for ws in wb.worksheets:
            for code, name in ws.iter_rows(min_row=2, min_col=2, max_col=3, values_only=True):
                if code and name:
                    lookup[str(code).strip()] = str(name).strip()
    finally:
        wb.close()
    return lookup


def _resolve_item_names(items: list, lookup: dict) -> list:
    """若品項名稱完全符合 template_series 的品號，替換為對照表中的完整品名。"""
    resolved = []
    for item in items:
        key = str(item.get("name", "")).strip()
        if key in lookup:
            item = dict(item, name=lookup[key])
        resolved.append(item)
    return resolved


def _codes_for_filename(items: list) -> str:
    """取各品項的品號（簡寫），去重後串接，供檔名使用。"""
    codes = []
    for item in items:
        code = str(item.get("part_no", "")).strip()
        if code and code not in codes:
            codes.append(code)
    return "+".join(codes)


def _format_date(s):
    try:
        dt = datetime.strptime(s, "%Y/%m/%d")
        return f"{dt.year} 年  {dt.month:02d} 月  {dt.day:02d}  日"
    except Exception:
        return s


def _safe_write(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        cell.value = value



def _copy_row_style(ws, src_row, dst_row):
    for col in range(1, 11):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        if isinstance(dst, MergedCell):
            continue
        if src.has_style:
            dst.font          = copy(src.font)
            dst.border        = copy(src.border)
            dst.fill          = copy(src.fill)
            dst.number_format = src.number_format
            dst.alignment     = copy(src.alignment)
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def _shift_footer_merges(ws, n_extra):
    """
    insert_rows 不會自動移動在插入點「之後」的合併格。
    手動處理：找出所有 min_row >= FOOTER_START 的合併格，
    unmerge → 重新以 (row + n_extra) merge 回去。
    """
    to_shift = [
        (r.min_col, r.min_row, r.max_col, r.max_row)
        for r in ws.merged_cells.ranges
        if r.min_row >= FOOTER_START
    ]

    for (c1, r1, c2, r2) in to_shift:
        ws.unmerge_cells(
            start_row=r1, start_column=c1,
            end_row=r2,   end_column=c2
        )

    for (c1, r1, c2, r2) in to_shift:
        ws.merge_cells(
            start_row=r1 + n_extra, start_column=c1,
            end_row=r2   + n_extra, end_column=c2
        )


def _write_items_compact(ws, items):
    """1~3 個品項：全部塞進 ITEM_ROW 那一列，多品項以兩個換行分隔。"""
    from openpyxl.styles import Alignment
    sep  = "\n\n"
    wrap = Alignment(wrap_text=True, vertical="top", horizontal="center")
    wrap_left = Alignment(wrap_text=True, vertical="top", horizontal="left")

    def _join(values):
        return sep.join(str(v) if v not in (None, "") else "" for v in values)

    r = ITEM_ROW
    seqs      = [str(i + 1) for i in range(len(items))]
    names     = [item["name"].replace("\n", " ") for item in items]
    qtys      = [item["qty"]        for item in items]
    units     = [item["unit"]       for item in items]
    for col, vals, al in [
        (1, seqs,   wrap),
        (3, names,  wrap_left),
        (6, qtys,   wrap),
        (7, units,  wrap),
    ]:
        cell = ws.cell(row=r, column=col)
        if not isinstance(cell, MergedCell):
            cell.value     = _join(vals)
            cell.alignment = al

    # 清空模板預設的 0
    for col in (8, 9):
        cell = ws.cell(row=r, column=col)
        if not isinstance(cell, MergedCell):
            cell.value = None

    pass  # preserve template row height


def _append_invoice_to_footer(ws, invoice_choice):
    if invoice_choice == '隨貨':
        suffix = '    ■發票隨貨 □發票直寄'
    elif invoice_choice == '直寄':
        suffix = '    □發票隨貨 ■發票直寄'
    else:
        suffix = '    □發票隨貨 □發票直寄'

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell) or cell.value is None:
                continue
            if '第一聯' in str(cell.value) and '白聯' in str(cell.value):
                cell.value = str(cell.value) + suffix
                return


def generate(data, extra, out_filename="", output_dir=None):
    out_dir = Path(output_dir) if output_dir else OUTPUT_DIR
    items = _resolve_item_names(data["items"], _load_series_lookup())
    n     = len(items)

    # 超過 3 個品項：每 3 個一份，依序遞迴生成
    if n > 3:
        h        = data["header"]
        customer = h.get("customer", "客戶")
        date_tag = extra.get("ship_date", "").replace("/", "") or datetime.today().strftime("%Y%m%d")
        chunks   = [items[i:i+3] for i in range(0, n, 3)]
        paths    = []
        for idx, chunk in enumerate(chunks):
            if out_filename:
                stem = Path(out_filename).stem
                fn = out_filename if idx == 0 else f"{stem}-{idx + 1}.xlsx"
            else:
                suffix = f"-{idx + 1}" if idx > 0 else ""
                fn = f"出貨單-{customer}{_codes_for_filename(chunk)}-{date_tag}{suffix}.xlsx"
            paths.append(generate(dict(data, items=chunk), extra, fn, output_dir=output_dir))
        return paths

    out_dir.mkdir(exist_ok=True)
    tmp = out_dir / "_working.xlsx"
    shutil.copy(TEMPLATE_PATH, tmp)

    wb = openpyxl.load_workbook(tmp)
    ws = wb.active

    h       = data["header"]
    n_extra = n - 1   # 需插入的額外列數

    # ── 1. 固定表頭 ───────────────────────────────────────────
    ws["A5"] = "出貨日期：" + _format_date(extra.get("ship_date", ""))
    _safe_write(ws, 6, 3, h.get("customer", ""))
    _safe_write(ws, 7, 3, h.get("phone", ""))
    _safe_write(ws, 8, 3, h.get("address", ""))
    ws["F7"] = "聯絡人：" + h.get("contact", "")

    sale_no = extra.get("sale_no", "").strip()
    if sale_no:
        ws["I6"] = "銷貨單號：" + sale_no

    tax_id = h.get("tax_id", "").strip()
    if tax_id:
        ws["F6"] = "統一編號：" + tax_id

    # ── 2 & 3. 品項寫入 ───────────────────────────────────────
    if n <= 3:
        # 緊湊模式：全部塞進 ITEM_ROW，不插列
        _write_items_compact(ws, items)
        note_row = ITEM_ROW + 1
    else:
        # 一般模式：每個品項佔一列
        if n_extra > 0:
            _shift_footer_merges(ws, n_extra)
            ws.insert_rows(FOOTER_START, amount=n_extra)
        for i, item in enumerate(items):
            r = ITEM_ROW + i
            if i > 0:
                ws.merge_cells(f"A{r}:B{r}")
                ws.merge_cells(f"C{r}:E{r}")
                _copy_row_style(ws, ITEM_ROW, r)
            _safe_write(ws, r, 1, i + 1)
            _safe_write(ws, r, 3, item["name"].replace("\n", " "))
            _safe_write(ws, r, 6, item["qty"])
            _safe_write(ws, r, 7, item["unit"])
            _safe_write(ws, r, 8, item["unit_price"])
            _safe_write(ws, r, 9, item["subtotal"])
        note_row = ITEM_ROW + n

    # ── 4. Footer 補填 ─────────────────────────────────────────
    operator_row = note_row + 1

    note_text = extra.get("note", "").strip()
    if note_text:
        _safe_write(ws, note_row, 2, note_text)

    _safe_write(ws, operator_row, 3, extra.get("operator", ""))

    # ── 5. 發票方式（追加到※第一聯那行結尾）──────────────────────
    _append_invoice_to_footer(ws, extra.get("invoice_choice", "尚未確認"))

    # ── 6. 存檔 ────────────────────────────────────────────────
    if not out_filename:
        customer = h.get("customer", "客戶")
        date_tag = extra.get("ship_date", "").replace("/", "") or datetime.today().strftime("%Y%m%d")
        out_filename = f"出貨單-{customer}{_codes_for_filename(items)}-{date_tag}.xlsx"

    out_path = out_dir / out_filename
    wb.save(out_path)
    tmp.unlink(missing_ok=True)
    return out_path


if __name__ == "__main__":
    import sys
    sys.path.insert(0, str(Path(__file__).parent))
    from parser import parse

    data  = parse(sys.argv[1])
    extra = {
        "ship_date":      datetime.today().strftime("%Y/%m/%d"),
        "sale_no":        "TEST-001",
        "operator":       "小皋",
        "note":           "",
        "invoice_choice": "隨貨",
    }
    out = generate(data, extra)
    print(f"輸出：{out}")
