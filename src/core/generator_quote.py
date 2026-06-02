"""
generator_quote.py — 從 Trello 卡片資料填入報價單範本，生成 .xlsx
"""
import copy as _copy
import json
from openpyxl.cell.cell import MergedCell as _MergedCell
from openpyxl.cell.rich_text import CellRichText as _CellRichText, TextBlock as _TextBlock
from openpyxl.cell.text import InlineFont as _InlineFont
import re
from datetime import date, timedelta
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as _XlImage
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter as _col_letter


# ── 描述解析 ────────────────────────────────────────────────────


def _strip_md_link(text: str) -> str:
    """去除 Trello Markdown 連結格式，只保留顯示文字。
    [text](url) → text
    """
    return re.sub(r'\[([^\]]+)\]\([^)]*\)', r'\1', text).strip()


def parse_card_desc(desc: str) -> dict:
    """解析 TR 卡片描述（「欄位：值\n」格式）→ dict。
    同時支援全形 ：與半形 : 冒號。
    Trello 的 Markdown 連結格式（如 email 超連結）也會自動還原為純文字。
    """
    result: dict[str, str] = {}
    for line in desc.splitlines():
        # 全形優先；找不到再試半形
        colon = "：" if "：" in line else (":" if ":" in line else None)
        if colon is None:
            continue
        key, _, val = line.partition(colon)
        result[key.strip()] = _strip_md_link(val.strip())
    return result


# ── 範本工具 ────────────────────────────────────────────────────


def list_templates(template_dir: Path) -> list[Path]:
    """回傳 template_dir 內所有 報價單-*.xlsx，依檔名排序。"""
    return sorted(template_dir.glob("報價單-*.xlsx"))


def _norm(text: str) -> str:
    """正規化：移除空白、全形空格、冒號、連字號，轉大寫，方便模糊比對。
    ex: "E-MAIL" → "EMAIL"、"統一編號：" → "統一編號"
    """
    return re.sub(r"[\s：:　：\-]", "", str(text)).upper()


def _label_merge_end_col(ws, row: int, label_col: int) -> int:
    """回傳 label 所在合併儲存格的最右欄（無合併則回傳 label_col 本身）。"""
    for rng in ws.merged_cells.ranges:
        if (rng.min_row <= row <= rng.max_row
                and rng.min_col <= label_col <= rng.max_col):
            return rng.max_col
    return label_col


def _find_value_col(ws, row: int, label_col: int) -> int:
    """跳過 label 的合併範圍，找同列第一個可填入的值格（非 MergedCell）。"""
    start = _label_merge_end_col(ws, row, label_col) + 1
    for c in range(start, start + 8):
        cell = ws.cell(row=row, column=c)
        if isinstance(cell, _MergedCell):
            continue  # 合併區域非左上角格，無法寫入，跳過
        val = cell.value
        if val is None:
            return c
        if isinstance(val, str) and _norm(val) == "":
            return c
    return start


# label 關鍵字 → desc 欄位名（None 表示由呼叫端直接給值）
_LABEL_DESC_MAP: list[tuple[str, str | None]] = [
    ("客戶全名",  "公司名"),
    ("CUSTOMERNAME", "公司名"),   # 英文 label fallback
    ("電話",      None),          # 手機 or 電話，由呼叫端決定
    ("傳真",      "傳真"),
    ("聯絡人",    "聯絡人"),
    ("聯絡地址",  "地址"),
    ("統一編號",  "統一編號"),
    ("報價日期",  None),          # auto
    ("有效日期",  None),          # auto
    ("報價單號",  None),          # auto
]


def _fill_field(ws, label_kw_norm: str, value) -> bool:
    """掃描整張工作表，找到含 label_kw 的格後填入相鄰格。"""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            if label_kw_norm in _norm(str(cell.value)):
                target_col = _find_value_col(ws, cell.row, cell.column)
                out_cell = ws.cell(row=cell.row, column=target_col)
                if isinstance(out_cell, _MergedCell):
                    return False  # 找不到可寫格，跳過此欄位
                out_cell.value = value
                # 強制純文字，防止 Excel 自動將 email / URL 轉成超連結
                if isinstance(value, str) and ("@" in value or value.startswith("http")):
                    out_cell.number_format = "@"
                return True
    return False


# ── 需求解析 ────────────────────────────────────────────────────


def _parse_demand_items(demand_str: str) -> list[tuple[str, int]]:
    """從需求字串中找出所有「品號*數量」格式，回傳 [(品號, 數量), ...]。

    支援多行或空白/逗號分隔：
        "AP-50B*1"          → [("AP-50B", 1)]
        "AP-50B*1 AP-30B*2" → [("AP-50B", 1), ("AP-30B", 2)]
    """
    items: list[tuple[str, int]] = []
    for token in re.split(r"[\s,，]+", demand_str):
        m = re.match(r"([A-Za-z0-9][A-Za-z0-9\-]*)\*(\d+)$", token.strip())
        if m:
            items.append((m.group(1), int(m.group(2))))
    return items


def _extract_demand_from_desc(raw_desc: str) -> str:
    """從原始描述字串中取出「需求:」或「需求：」之後的所有內容。

    卡片描述常見格式：
        需求:
        AP-50B*2
        AP-30B*1

    parse_card_desc 只抓同行 val，會漏掉換行後的品項，
    此函式直接從原始文字擷取 需求 標題之後的多行內容。
    """
    for keyword in ("需求：", "需求:"):
        idx = raw_desc.find(keyword)
        if idx == -1:
            continue
        after = raw_desc[idx + len(keyword):]
        # 同行有值就包含進來，後續行繼續收集直到遇到下一個「key:」行為止
        lines = []
        for line in after.splitlines():
            stripped = line.strip()
            # 遇到下一個欄位（含冒號的行）就停止
            if stripped and ("：" in stripped or ":" in stripped) and not re.match(
                r"[A-Za-z0-9][A-Za-z0-9\-]*\*\d+", stripped
            ):
                break
            lines.append(stripped)
        return " ".join(l for l in lines if l)
    return ""


# ── 報價單號 ────────────────────────────────────────────────────


def next_quote_no(output_dir: Path, operator_code: str, quote_date: date) -> str:
    """依輸出資料夾內當日已有的報價單數，產生下一組報價單號。

    格式：{代號}{民國年}{月:02d}{日:02d}{序號:03d}
    Ex: K114052100 1 → 「K1140521001」
    """
    roc_year = quote_date.year - 1911
    prefix   = f"{operator_code}{roc_year}{quote_date.month:02d}{quote_date.day:02d}"
    count    = 0
    if output_dir.exists():
        for f in output_dir.iterdir():
            if f.stem.startswith(f"報價單_") and prefix in f.stem:
                count += 1
    return f"{prefix}{count + 1:03d}"


# ── 主要生成函式 ────────────────────────────────────────────────


def generate_quote(
    card: dict,
    template_path: Path,
    output_dir: Path,
    quote_no: str,
    quote_date: date,
) -> Path:
    """將 Trello 卡片資料填入報價單範本並儲存。

    回傳輸出 .xlsx 的路徑。
    """
    desc_fields = parse_card_desc(card.get("desc", ""))
    valid_date  = quote_date + timedelta(days=15)

    def _get(*keys: str) -> str:
        """依序嘗試多個 key，回傳第一個非空值，否則回傳空字串。"""
        for k in keys:
            v = desc_fields.get(k)
            if v:
                return v
        return ""

    # 客戶全名：可能叫「公司名」或「公司名稱」
    company = _get("公司名", "公司名稱").strip()

    # 電話：優先手機，fallback 電話
    phone = _get("手機", "電話")

    fill_map: dict[str, object] = {
        "客戶全名":  company,
        "電話":      phone,
        "傳真":      _get("傳真"),
        "聯絡人":    _get("聯絡人"),
        "聯絡地址":  _get("地址", "聯絡地址"),
        # 統一編號：繁體/簡體「統」，或縮寫「统編」
        "統一編號":  _get("統一編號", "统一編號", "统編", "統編"),
        # EMAIL：可能叫「電子信箱」或直接寫「E-MAIL」
        "EMAIL":     _get("電子信箱", "E-MAIL", "EMAIL", "E-Mail", "Mail"),
        "報價日期":  quote_date.strftime("%Y/%m/%d"),
        "有效日期":  valid_date.strftime("%Y/%m/%d"),
        "報價單號":  quote_no,
    }

    try:
        wb = openpyxl.load_workbook(str(template_path))
    except Exception as e:
        # 部分範本含有損壞的圖片/圖表引用（xl/drawings/NULL），
        # 改用 keep_links=False 再試一次
        if "xl/drawings" in str(e).lower() or "no item named" in str(e).lower():
            wb = openpyxl.load_workbook(str(template_path), keep_links=False)
        else:
            raise
    ws = wb.active

    for label_kw, value in fill_map.items():
        if value:
            _fill_field(ws, _norm(label_kw), value)

    # ── 需求明細：「品號*數量」→ 從第 15 列起填入 ──────────────
    # 需求項目可能換行寫在「需求:」下面，parse_card_desc 會漏掉，
    # 改用 _extract_demand_from_desc 直接從原始描述抓。
    demand_raw = _extract_demand_from_desc(card.get("desc", ""))
    if demand_raw:
        items = _parse_demand_items(demand_raw)
        for i, (_part_no, qty) in enumerate(items):
            row = 15 + i
            ws.cell(row=row, column=6, value=qty)                   # F = 數量
            ws.cell(row=row, column=9, value=f"=F{row}*H{row}")    # I = 小計

    output_dir.mkdir(parents=True, exist_ok=True)
    safe_company = re.sub(r'[\\/:*?"<>|]', "_", company).strip() or "客戶"
    out_path     = output_dir / f"報價單_{safe_company}.xlsx"
    wb.save(str(out_path))
    wb.close()
    return out_path


# ── 產品目錄 ────────────────────────────────────────────────────


def load_product_catalog(template_dir: Path) -> list[dict]:
    """從 template_dir/products.json 載入產品目錄，找不到回傳空清單。"""
    path = template_dir / "products.json"
    if not path.exists():
        return []
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return []


# ── 品號列範圍對照表 ─────────────────────────────────────────────
# 格式：品號 → (工作表名稱, 起始列, 結束列)
# 工作表名稱對應 template_series.xlsx 內的 Sheet 頁籤名稱

PRODUCT_ROWS: dict[str, tuple[str, int, int]] = {
    # ── APB 電動升降台車（Sheet: APB）───────────────────────────
    "AP-15WB(經銷)": ("APB", 2,  7),   # 6 行
    "AP-15WB":       ("APB", 9,  22),  # 14 行
    "AP-30B":        ("APB", 24, 37),  # 14 行
    "AP-30SB":       ("APB", 39, 51),  # 13 行
    "AP-50B":        ("APB", 53, 65),  # 13 行
    "AP-50WB":       ("APB", 67, 79),  # 13 行
    "AP-80B":        ("APB", 81, 93),  # 13 行
    # ── AM 拖板車（Sheet: AM）────────────────────────────────
    "AM15SS-L":  ("AM",  2,  10),  # 9 行
    "AM20C":     ("AM", 12,  20),  # 9 行
    "AM20L":     ("AM", 22,  30),  # 9 行
    "AM20LL":    ("AM", 32,  40),  # 9 行
    "AM20M":     ("AM", 42,  50),  # 9 行
    "AM20S":     ("AM", 52,  60),  # 9 行
    "AM20SS":    ("AM", 62,  70),  # 9 行
    "AM25":      ("AM", 72,  80),  # 9 行
    "AM25C":     ("AM", 82,  90),  # 9 行
    "AM25L":     ("AM", 92, 100),  # 9 行
    "AM25LL":    ("AM",102, 110),  # 9 行
    "AM25LLL":   ("AM",112, 120),  # 9 行
    "AM25M":     ("AM",122, 130),  # 9 行
    "AM25ML":    ("AM",132, 140),  # 9 行
    "AM25S":     ("AM",142, 150),  # 9 行
    "AM25SS":    ("AM",152, 160),  # 9 行
    "AM30LL":    ("AM",162, 170),  # 9 行
    "AM30M":     ("AM",172, 180),  # 9 行
    "AM40LL":    ("AM",182, 190),  # 9 行
    "AM50L4":    ("AM",192, 200),  # 9 行
    # ── APL 彈簧高度定位台車（Sheet: APL）───────────────────────
    "APL05":   ("APL",  2, 11),  # 10 行
    "APL08":   ("APL", 13, 22),  # 10 行
    "APL10":   ("APL", 24, 33),  # 10 行
    "APL10L":  ("APL", 35, 44),  # 10 行
    "APL21L":  ("APL", 46, 55),  # 10 行
    "APL21":   ("APL", 57, 66),  # 10 行
    "APL40":   ("APL", 68, 77),  # 10 行
    "APL60":   ("APL", 79, 88),  # 10 行
    # ── APS 不鏽鋼升降台車（Sheet: APS）────────────────────────
    "APS-10":  ("APS",  2, 12),  # 11 行
    "APS-20":  ("APS", 14, 24),  # 11 行
    "APS-25":  ("APS", 26, 36),  # 11 行
    "APS-50":  ("APS", 38, 48),  # 11 行
    # ── APT 系列（Sheet: APT）───────────────────────────────────
    "APT15F":  ("APT",  2, 24),  # 23 行
    "APT15L":  ("APT", 26, 48),  # 23 行
    "APT15LL": ("APT", 50, 65),  # 16 行
    # ── AP 油壓台車（Sheet: AP油壓）──────────────────────────────
    "AP-15":              ("AP油壓",   2,  13),  # 12 行
    "AP15W":              ("AP油壓",  15,  26),  # 12 行
    "AP15W+進口地(經銷)": ("AP油壓",  28,  34),  #  7 行
    "AP-30":              ("AP油壓",  36,  47),  # 12 行
    "AP-30SW":            ("AP油壓",  49,  60),  # 12 行
    "AP30W":              ("AP油壓",  62,  72),  # 11 行
    "AP-50":              ("AP油壓",  74,  84),  # 11 行
    "AP-50SW":            ("AP油壓",  86,  96),  # 11 行
    "AP-50W":             ("AP油壓",  98, 108),  # 11 行
    "AP-80":              ("AP油壓", 110, 120),  # 11 行
    "APXH-200":           ("AP油壓", 122, 129),  #  8 行
    "APXH-500":           ("AP油壓", 131, 143),  # 13 行
    # ── BMT 拖板車（Sheet: BMT）────────────────────────────────
    # 注意：BMT-20LL 在模板 B欄顯示為「BMT-20M」，但規格為 1220×685mm（LL型）
    "BMT-20LL":  ("BMT",  2, 10),  #  9 行
    "BMT-20LLL": ("BMT", 12, 20),  #  9 行
    "BMT-20M":   ("BMT", 22, 30),  #  9 行
    "BMT-20SS":  ("BMT", 32, 40),  #  9 行
    "BMT-25LL":  ("BMT", 42, 47),  #  6 行
    "BMT-30LL":  ("BMT", 49, 57),  #  9 行
    # ── BST 吊捍機（Sheet: BST）─────────────────────────────────
    "BST-80":    ("BST",  2, 15),  # 14 行
    "BST-200":   ("BST", 17, 30),  # 14 行
    # ── BXT 油壓台車（Sheet: BXT）───────────────────────────────
    "BXT-100":   ("BXT",  2,  8),  #  7 行
    "BXT-100B":  ("BXT", 10, 17),  #  8 行
    # ── EZ Lifter 自動升降疊棧機（Sheet: EZ lifter）──────────────
    "EZ Lifter":   ("EZ lifter",  2,  9),  #  8 行
    "EZO-25E":     ("EZ lifter", 11, 20),  # 10 行
    "EZO-25E-3S":  ("EZ lifter", 22, 31),  # 10 行
    # ── HT 重量型電動油壓升降平台（Sheet: HT）────────────────────
    "2HT101211": ("HT",   2,  15),  # 14 行 - 載重 1T, 1200×1150mm
    "2HT101214": ("HT",  17,  30),  # 14 行 - 載重 1T, 1200×1400mm
    "2HT101218": ("HT",  32,  45),  # 14 行 - 載重 1T, 1200×1850mm
    "2HT106010": ("HT",  47,  60),  # 14 行 - 載重 1T, 600×1000mm
    "2HT106011": ("HT",  62,  75),  # 14 行 - 載重 1T, 600×1150mm
    "2HT106014": ("HT",  77,  90),  # 14 行 - 載重 1T, 600×1400mm
    "2HT106018": ("HT",  92, 105),  # 14 行 - 載重 1T, 600×1850mm
    "2HT106078": ("HT", 107, 122),  # 16 行 - 載重 1T, 600×780mm
    "2HT108510": ("HT", 124, 137),  # 14 行 - 載重 1T, 850×1000mm
    "2HT108511": ("HT", 139, 152),  # 14 行 - 載重 1T, 850×1150mm
    "2HT108514": ("HT", 154, 167),  # 14 行 - 載重 1T, 850×1400mm
    "2HT108518": ("HT", 169, 182),  # 14 行 - 載重 1T, 850×1850mm
    "HT10410":   ("HT", 184, 197),  # 14 行 - 載重 1T, 400×1000mm
    "HT10413":   ("HT", 199, 212),  # 14 行 - 載重 1T, 400×1340mm
    "HT10415":   ("HT", 214, 227),  # 14 行 - 載重 1T, 400×1500mm
    "HT101013":  ("HT", 229, 242),  # 14 行 - 載重 1T, 1000×1340mm
    "HT101015":  ("HT", 244, 257),  # 14 行 - 載重 1T, 1000×1500mm
    "HT101020":  ("HT", 259, 272),  # 14 行 - 載重 1T, 1000×2000mm
    "HT101224":  ("HT", 274, 287),  # 14 行 - 載重 1T, 1200×2400mm
    "HT106410":  ("HT", 289, 302),  # 14 行 - 載重 1T, 640×1000mm
    "HT106413":  ("HT", 304, 317),  # 14 行 - 載重 1T, 640×1340mm
    "HT106415":  ("HT", 319, 332),  # 14 行 - 載重 1T, 640×1500mm
    "HT106418":  ("HT", 334, 347),  # 14 行 - 載重 1T, 640×1800mm
    "HT106485":  ("HT", 349, 362),  # 14 行 - 載重 1T, 640×850mm
    "HT107810":  ("HT", 364, 377),  # 14 行 - 載重 1T, 780×1000mm
    "HT107813":  ("HT", 379, 392),  # 14 行 - 載重 1T, 780×1340mm
    "HT107815":  ("HT", 394, 407),  # 14 行 - 載重 1T, 780×1500mm
    "HT107818":  ("HT", 409, 422),  # 14 行 - 載重 1T, 780×1800mm
    "HT107885":  ("HT", 424, 437),  # 14 行 - 載重 1T, 780×850mm
    "HT201013":  ("HT", 439, 452),  # 14 行 - 載重 2T, 1000×1300mm
    "HT201015":  ("HT", 454, 467),  # 14 行 - 載重 2T, 1000×1500mm
    "HT201020":  ("HT", 469, 482),  # 14 行 - 載重 2T, 1000×2000mm
    "HT201224":  ("HT", 484, 497),  # 14 行 - 載重 2T, 1200×2400mm
    "HT207013":  ("HT", 499, 512),  # 14 行 - 載重 2T, 700×1300mm
    "HT207015":  ("HT", 514, 528),  # 15 行 - 載重 2T, 700×1500mm
    "HT301013":  ("HT", 530, 543),  # 14 行 - 載重 3T, 1000×1300mm
    "HT301015":  ("HT", 545, 558),  # 14 行 - 載重 3T, 1000×1500mm
    "HT301020":  ("HT", 560, 573),  # 14 行 - 載重 3T, 1000×1200mm
    "HT301224":  ("HT", 575, 588),  # 14 行 - 載重 3T, 1200×2400mm
    "HT307013":  ("HT", 590, 603),  # 14 行 - 載重 3T, 700×1300mm
    "HT307015":  ("HT", 605, 618),  # 14 行 - 載重 3T, 700×1500mm
    # ── KGM（Sheet: KGM）
    "KGM20F": ("KGM",   2,  13),  # 12 行
    "KGM20HF": ("KGM",  15,  26),  # 12 行
    "KGM25": ("KGM",  28,  38),  # 11 行
    "KGM25H": ("KGM",  40,  50),  # 11 行
    "KGM40F": ("KGM",  52,  63),  # 12 行
    "KGM40HF": ("KGM",  65,  76),  # 12 行
    "KGM45": ("KGM",  78,  88),  # 11 行
    "KGM45H": ("KGM",  90, 100),  # 11 行
    # ── KGX（Sheet: KGX）
    "KGX40-12": ("KGX",   2,   9),  # 8 行
    "KGX-40-12DC": ("KGX",  11,  20),  # 10 行
    "KGX-40-15": ("KGX",  22,  29),  # 8 行
    "KGX-40-15DC": ("KGX",  31,  40),  # 10 行
    "KGX-40-17": ("KGX",  42,  49),  # 8 行
    "KGX-40-17DC": ("KGX",  51,  60),  # 10 行
    # ── LSCM（Sheet: LSCM）
    "LSCM340": ("LSCM",   2,  12),  # 11 行
    "LSCM520": ("LSCM",  14,  25),  # 12 行
    # ── LSES（Sheet: LSES）
    "LS400EQ": ("LSES",   2,  14),  # 13 行
    "LSES-100": ("LSES",  16,  28),  # 13 行
    # ── LSF（Sheet: LSF）
    "LSF-1025": ("LSF",   2,  27),  # 26 行
    "LSF-1525": ("LSF",  29,  54),  # 26 行
    "LSF-2025": ("LSF",  56,  81),  # 26 行
    # ── LSGR（Sheet: LSGR ）
    "LSGR6515": ("LSGR ",   2,  16),  # 15 行
    # ── LSJ（Sheet: LSJ）
    "LSJ-20": ("LSJ",   2,  32),  # 31 行
    "LSJ-25": ("LSJ",  34,  64),  # 31 行
    "LSJ-30": ("LSJ",  66,  96),  # 31 行
    # ── LSL（Sheet: LSL）
    "LSL-1028": ("LSL",   2,  27),  # 26 行
    "LSL-1228": ("LSL",  29,  54),  # 26 行
    "LSL-2028": ("LSL",  56,  81),  # 26 行
    # ── LSLM（Sheet: LSLM）
    "LSLM1210F": ("LSLM",   2,  11),  # 10 行
    "LSLM1210H": ("LSLM",  13,  22),  # 10 行
    # ── LSM（Sheet: LSM）
    "LSM1028": ("LSM",   2,  26),  # 25 行
    "LSM1528": ("LSM",  28,  52),  # 25 行
    "LSM2028": ("LSM",  54,  78),  # 25 行
    # ── LSP（Sheet: LSP）
    "LSP-1025": ("LSP",   2,  26),  # 25 行
    "LSP-1225": ("LSP",  28,  52),  # 25 行
    # ── LSPV（Sheet: LSPV）
    "LSPV-0327": ("LSPV",   2,  28),  # 27 行
    "LSPV-0330": ("LSPV",  30,  56),  # 27 行
    "LSPV-0340": ("LSPV",  58,  84),  # 27 行
    "LSPV-0350": ("LSPV",  86, 112),  # 27 行
    # ── LSR（Sheet: LSR）
    "LSR-1028": ("LSR",   2,  29),  # 28 行
    "LSR-1228": ("LSR",  31,  58),  # 28 行
    "LSR-1528": ("LSR",  60,  87),  # 28 行
    "LSR-1828": ("LSR",  89, 116),  # 28 行
    # ── LSS（Sheet: LSS）
    "LSS-1028": ("LSS",   2,  27),  # 26 行
    "LSS-1628": ("LSS",  29,  54),  # 26 行
    "LSS-2028": ("LSS",  56,  81),  # 26 行
    # ── LSV（Sheet: LSV）
    "LSV-1016": ("LSV",   2,  14),  # 13 行
    "LSV-1025": ("LSV",  16,  28),  # 13 行
    "LSV-1225": ("LSV",  30,  42),  # 13 行
    "LSV-1316": ("LSV",  44,  56),  # 13 行
    "LSV-1516": ("LSV",  58,  70),  # 13 行
    "LSV-1525": ("LSV",  72,  84),  # 13 行
    # ── LS傾斜（Sheet: LS傾斜）
    "LS-E50":  ("LS傾斜",   2,  12),  # 11 行
    "LS-E100": ("LS傾斜",  14,  24),  # 11 行
    "LS-H50":  ("LS傾斜",  26,  35),  # 10 行
    "LS-H100": ("LS傾斜",  37,  46),  # 10 行
    # ── LT（Sheet: LT）
    "LT25210":  ("LT",   2,  16),  # 15 行
    "LT25281":  ("LT",  18,  32),  # 15 行
    "LT55210":  ("LT",  34,  48),  # 15 行
    "LT55281":  ("LT",  50,  64),  # 15 行
    "LT105212": ("LT",  66,  80),  # 15 行
    "LT107812": ("LT",  82,  96),  # 15 行
    "LT201015": ("LT",  98, 112),  # 15 行
    "LT201020": ("LT", 114, 128),  # 15 行
    "LT207313": ("LT", 130, 144),  # 15 行
    "LT301015": ("LT", 146, 160),  # 15 行
    "LT301020": ("LT", 162, 176),  # 15 行
    "LT307313": ("LT", 178, 192),  # 15 行
    # ── NWL（Sheet: NWL）
    "NWL18H": ("NWL",   2,  14),  # 13 行
    "NWL18L": ("NWL",  16,  28),  # 13 行
    # ── SL 油壓頂升拖板車（Sheet: SL）
    "SL50":    ("SL",   2,  13),  # 12 行
    "SL50E":   ("SL",  15,  27),  # 13 行
    "SL50N":   ("SL",  29,  40),  # 12 行
    "SL50NSS": ("SL",  42,  53),  # 12 行
    "SL50SS":  ("SL",  55,  66),  # 12 行
    "SL50W":   ("SL",  68,  79),  # 12 行
    "SL50WE":  ("SL",  81,  92),  # 12 行
    "SL100":   ("SL",  94, 105),  # 12 行
    "SL100E":  ("SL", 107, 119),  # 13 行
    "SL100W":  ("SL", 121, 132),  # 12 行
    "SL100WE": ("SL", 134, 146),  # 13 行
    # ── ST系列 物流堆高機（Sheet: SL 系列）
    "STH50":  ("SL 系列",   2,  10),  # 9 行
    "STH65":  ("SL 系列",  12,  20),  # 9 行
    "STH80":  ("SL 系列",  22,  30),  # 9 行
    "STH100": ("SL 系列",  32,  40),  # 9 行
    "STM25":  ("SL 系列",  42,  50),  # 9 行
    "STM38":  ("SL 系列",  52,  60),  # 9 行
    "STM50":  ("SL 系列",  62,  70),  # 9 行
    "STM65":  ("SL 系列",  72,  80),  # 9 行
    "STS38":  ("SL 系列",  82,  90),  # 9 行
    # ── TT 可傾倒式盎斗台車（Sheet: TT）
    "TT15": ("TT",   2,  10),  # 9 行
    "TT25": ("TT",  12,  20),  # 9 行
    # ── TT25MH 頂高傾倒式盎斗台車（Sheet: TT 25MH）
    "TT25MH": ("TT 25MH",   2,  14),  # 13 行
    "TT25TH": ("TT 25MH",  16,  28),  # 13 行
    # ── TT25THE 頂高傾倒式電動盎斗台車（Sheet: TT 25THE）
    "AS25":     ("TT 25THE",   2,  12),  # 11 行
    "DA15":     ("TT 25THE",  14,  20),  #  7 行
    "DA25":     ("TT 25THE",  22,  28),  #  7 行
    "TT25MHE":  ("TT 25THE",  30,  43),  # 14 行
    "TT25THE":  ("TT 25THE",  45,  59),  # 15 行
    # ── T輕型 電動油壓升降平台（Sheet: T輕型）
    "2T050510B": ("T輕型",   2,  15),  # 14 行
    "T20B":      ("T輕型",  17,  30),  # 14 行
    "T30B":      ("T輕型",  32,  45),  # 14 行
    "T50B":      ("T輕型",  47,  60),  # 14 行
    "T80B":      ("T輕型",  62,  75),  # 14 行
    # ── 手推車（Sheet: 手推車）
    "L50080HT": ("手推車",   2,  10),  # 9 行
    "LS-100":   ("手推車",  12,  16),  # 5 行
    "LS-960":   ("手推車",  18,  24),  # 7 行
    "LS39059":  ("手推車",  26,  32),  # 7 行
    # ── 烏龜車（Sheet: 烏龜車）
    "50加侖桶鐵烏龜車": ("烏龜車",   2,   6),  # 5 行
    "LS-640":   ("烏龜車",   8,  14),  # 7 行
    "LS-640L":  ("烏龜車",  16,  22),  # 7 行
    "LS-670":   ("烏龜車",  24,  30),  # 7 行
    "LS-670L":  ("烏龜車",  32,  38),  # 7 行
    "LS-850":   ("烏龜車",  40,  46),  # 7 行
    "LSI-150":  ("烏龜車",  48,  52),  # 5 行
    "LSI-200":  ("烏龜車",  54,  58),  # 5 行
    "LSP-670L": ("烏龜車",  60,  65),  # 6 行
}

# 向下相容舊名稱
PRODUCT_AP_ROWS = {k: v[1:] for k, v in PRODUCT_ROWS.items()}

_PRODUCT_START_ROW = 15   # template_quote / allowance 品項區起始列
_PRODUCT_START_ROW_COMPARE = 11   # template_quote_compare 品項區起始列

# quote_type 常數
QUOTE_TYPE_REGULAR   = "regular"
QUOTE_TYPE_ALLOWANCE = "allowance"
QUOTE_TYPE_COMPARE   = "compare"


# ── 品項列 border 樣式（從 template_quote R15 讀出的規格）────────
# 僅左右垂直線；厚外框、medium 內格線
_TK = Side(style="thick")
_MD = Side(style="medium")
_NO = Side(style=None)

# col → (left_side, right_side)
_PRODUCT_ROW_BORDERS: dict[int, tuple] = {
    1:  (_TK, _NO),
    2:  (_MD, _MD),
    3:  (_NO, _NO),
    4:  (_NO, _NO),
    5:  (_NO, _MD),
    6:  (_MD, _NO),
    7:  (_MD, _MD),
    8:  (_MD, _MD),
    9:  (_MD, _NO),
    10: (_NO, _TK),
}


_NUM_FMT = '#,##0'   # 千分位數字格式


def _apply_product_row_border(ws, row: int) -> None:
    """套用 template_quote 品項列的標準 border（左右垂直線）。"""
    for col, (lft, rgt) in _PRODUCT_ROW_BORDERS.items():
        ws.cell(row=row, column=col).border = Border(left=lft, right=rgt)


def _apply_price_fmt(ws, row: int) -> None:
    """套用千分位格式到單價（H）與小計（I:J）儲存格。"""
    ws.cell(row=row, column=8).number_format = _NUM_FMT
    ws.cell(row=row, column=9).number_format = _NUM_FMT


def _apply_separator_row_border(ws, row: int) -> None:
    """空列套完整欄位垂直線，讓上下品號之間的分欄線視覺連續。"""
    _apply_product_row_border(ws, row)


# ── 品項列字型與對齊 ─────────────────────────────────────────────
_JHENG_HEI = Font(name="微軟正黑體", size=11)
# wrap_text 不設 True：讓規格文字自然向右溢出空欄，避免窄欄折行
_ALIGN_CTR  = Alignment(horizontal="center", vertical="center")
_ALIGN_LEFT = Alignment(horizontal="left",   vertical="center")


def _style_product_row(ws, row: int) -> None:
    """套用品項列的字型（微軟正黑體）與對齊（品名左齊，其他置中）。"""
    for col in range(1, 11):
        cell = ws.cell(row=row, column=col)
        cell.font      = _JHENG_HEI
        cell.alignment = _ALIGN_LEFT if col in (3, 4, 5) else _ALIGN_CTR


# ── 麥當勞流程報價單生成 ────────────────────────────────────────


def generate_quote_from_cart(
    customer: dict,
    cart_items: list[dict],
    template_path: Path,
    output_dir: Path,
    quote_no: str,
    quote_date: date,
    operator: str = "",
    shipping: dict | None = None,
    quote_type: str = QUOTE_TYPE_REGULAR,
    card_title: str = "",
) -> Path:
    """將購物車品項填入 template_quote.xlsx，生成完整報價單。

    品項資料從 template_series.xlsx 對應列複製（A-H 欄），
    按品號動態插入，品號間保留一行空列，最後更新 營業稅 / 應收總金額 公式。

    customer 欄位：company, phone, fax, contact, address, tax_id, email
    cart_items 欄位：code, name, spec, unit, qty, price
    operator  ：製表人員姓名
    """
    valid_date = quote_date + timedelta(days=15)

    # ── 依報價單類型決定起始列 ──────────────────────────────
    prod_start_row = (
        _PRODUCT_START_ROW_COMPARE
        if quote_type == QUOTE_TYPE_COMPARE
        else _PRODUCT_START_ROW
    )

    # ── 1. 開啟 template_quote ──────────────────────────────
    try:
        wb = openpyxl.load_workbook(str(template_path))
    except Exception as e:
        if "xl/drawings" in str(e).lower() or "no item named" in str(e).lower():
            wb = openpyxl.load_workbook(str(template_path), keep_links=False)
        else:
            raise
    ws = wb.active
    _template_img_count = len(ws._images)  # template 預埋圖片數（印章等）
    # 記錄每張預埋圖片的原始錨點列號，用來區分 logo（靠上）與印章（靠下）
    _template_img_orig_rows = []
    for _ti in ws._images:
        _anc = getattr(_ti, "anchor", None)
        if isinstance(_anc, str):
            _m = re.search(r"(\d+)$", _anc)
            _orig_row = int(_m.group(1)) if _m else 0
        else:
            # anchor 是 openpyxl 物件（TwoCellAnchor / OneCellAnchor）
            # 直接讀 _from.row 或 marker.row（0-indexed → +1）
            _from = getattr(_anc, "_from", None) or getattr(_anc, "marker", None)
            _orig_row = (getattr(_from, "row", 0) or 0) + 1 if _from else 0
        _template_img_orig_rows.append(_orig_row)

    # ── 2. 開啟 template_series（來源規格，多工作表）───────────────
    ap_path = template_path.parent / "template_series.xlsx"
    wb_ap = None
    if ap_path.exists():
        try:
            wb_ap = openpyxl.load_workbook(str(ap_path), keep_links=False)
        except Exception:
            wb_ap = None

    def _get_ap_sheet(sheet_name: str):
        """取得對應工作表；找不到時回退到第一個工作表。"""
        if not wb_ap:
            return None
        if sheet_name in wb_ap.sheetnames:
            return wb_ap[sheet_name]
        return wb_ap.active  # 相容舊版單一工作表

    # ── 3. 填入顧客資訊 ──────────────────────────────────────
    if quote_type == QUOTE_TYPE_COMPARE:
        # 對比報價單：直接定位填入，不掃描全表（R1-R4 為公司自身資料，不可覆寫）
        # R5：日期（標籤與值同格）
        ws.cell(5, 1).value = f"報價日期  :  {quote_date.strftime('%Y/%m/%d')}"
        # R6-R9：依標籤欄位直接定位
        _cmp_header = [
            (6, 1, customer.get("company",  "")),   # 客戶全名
            (6, 7, quote_no),                        # 報價單號
            (7, 1, customer.get("phone",    "")),    # 電話
            (7, 7, customer.get("fax",      "")),    # 傳真
            (8, 1, customer.get("contact",  "")),    # 聯絡人
            (8, 7, customer.get("tax_id",   "")),    # 統一編號
            (9, 1, customer.get("address",  "")),    # 聯絡地址
            (9, 7, customer.get("email",    "")),    # E-MAIL
        ]
        for _r, _lc, _v in _cmp_header:
            if _v:
                _vc = _find_value_col(ws, _r, _lc)
                _oc = ws.cell(_r, _vc)
                if not isinstance(_oc, _MergedCell):
                    _oc.value = _v
    else:
        header_map = {
            "客戶全名":  customer.get("company", ""),
            "電話":      customer.get("phone",   ""),
            "傳真":      customer.get("fax",     ""),
            "聯絡人":    customer.get("contact", ""),
            "聯絡地址":  customer.get("address", ""),
            "統一編號":  customer.get("tax_id",  ""),
            "EMAIL":     customer.get("email",   ""),
            "報價日期":  quote_date.strftime("%Y/%m/%d"),
            "有效日期":  valid_date.strftime("%Y/%m/%d"),
            "報價單號":  quote_no,
        }
        for label_kw, value in header_map.items():
            if value:
                _fill_field(ws, _norm(label_kw), value)

    # ── 4. 找 footer 錨點列（補助=總應付金額, 其餘=營業稅）───────────
    if quote_type == QUOTE_TYPE_ALLOWANCE:
        _anchor_kw = "總應付金額"
    else:
        _anchor_kw = "營業稅"

    tax_row = None
    for r in ws.iter_rows():
        for cell in r:
            if cell.value and _anchor_kw in str(cell.value):
                tax_row = cell.row
                break
        if tax_row:
            break
    if tax_row is None:
        tax_row = prod_start_row + 2

    # ── 5. 取出 footer 合併範圍（prod_start_row+，之後重新定位）────
    footer_merges: list[dict] = []
    stale_ranges  = [rng for rng in ws.merged_cells.ranges
                     if rng.min_row >= prod_start_row]
    for rng in stale_ranges:
        footer_merges.append({
            "row_offset":     rng.min_row - tax_row,
            "max_row_offset": rng.max_row - tax_row,
            "min_col":        rng.min_col,
            "max_col":        rng.max_col,
        })
        ws.unmerge_cells(str(rng))

    # ── 6. 刪除原有佔位品項列 ───────────────────────────────
    placeholder_count = tax_row - prod_start_row
    if placeholder_count > 0:
        ws.delete_rows(prod_start_row, placeholder_count)
        tax_row -= placeholder_count

    # ── 7. 逐品號插入資料列 ──────────────────────────────────
    insert_pos     = prod_start_row
    first_item_row = prod_start_row

    for idx, item in enumerate(cart_items):
        code  = item.get("code",  "")
        qty   = item.get("qty",   1)
        price = item.get("price", 0)

        _row_info  = PRODUCT_ROWS.get(code) if wb_ap else None
        ap_range   = (_row_info[1], _row_info[2]) if _row_info else None
        ws_ap      = _get_ap_sheet(_row_info[0]) if _row_info else None
        block_rows = (ap_range[1] - ap_range[0] + 1) if ap_range else 1

        # 插入空列（openpyxl 會自動將 tax_row 以下的列下移）
        ws.insert_rows(insert_pos, amount=block_rows)
        tax_row += block_rows

        # 設定列高
        for j in range(block_rows):
            ws.row_dimensions[insert_pos + j].height = 18.0

        if ap_range and ws_ap:
            # ── 從 template_series 複製 A-H 欄值 ─────────────
            ap_start = ap_range[0]
            for j in range(block_rows):
                dst_row = insert_pos + j
                src_row = ap_start + j
                for col in range(1, 9):   # A(1) ~ H(8)
                    src_val = ws_ap.cell(row=src_row, column=col).value
                    if src_val is not None:
                        ws.cell(row=dst_row, column=col, value=src_val)
                # 字型、對齊、border
                _style_product_row(ws, dst_row)
                _apply_product_row_border(ws, dst_row)
        else:
            # fallback：找不到 template_series 時只填基本欄位
            ws.cell(row=insert_pos, column=2, value=code)
            _style_product_row(ws, insert_pos)
            _apply_product_row_border(ws, insert_pos)

        # ── 覆寫序號、數量、單價、小計公式 ────────────────
        hrow = insert_pos                              # 品號 header 列
        ws.cell(row=hrow, column=1, value=idx + 1)    # A  序
        ws.cell(row=hrow, column=6, value=qty)        # F  數量
        ws.cell(row=hrow, column=8, value=price)      # H  單價
        ws.cell(row=hrow, column=9,                   # I  小計
                value=f"=F{hrow}*H{hrow}")
        _apply_price_fmt(ws, hrow)

        # ── 品名欄（C:E）與小計欄（I:J）合併，對齊表頭寬度 ──
        for sc, ec in ((3, 5), (9, 10)):
            try:
                ws.merge_cells(
                    start_row=hrow, end_row=hrow,
                    start_column=sc, end_column=ec,
                )
            except Exception:
                pass

        # ── 插入產品圖片（Picture/{系列}/{品號}.{ext}） ──────────
        pic_dir  = template_path.parent / "Picture"
        category = item.get("category", "")
        _pic_file = None
        for _ext in ("png", "jpg", "jpeg", "bmp"):
            # 優先：Picture/{category}/{code}.{ext}
            if category:
                _f = pic_dir / category / f"{code}.{_ext}"
                if _f.exists():
                    _pic_file = _f
                    break
            # 兜底：glob 掃全部子資料夾
            _matches = list(pic_dir.glob(f"*/{code}.{_ext}"))
            if _matches:
                _pic_file = _matches[0]
                break
        if _pic_file:
            try:
                _img = _XlImage(str(_pic_file))
                _img.width  = 205
                _img.height = 162
                _img.anchor = f"{_col_letter(7)}{hrow + 3}"
                ws.add_image(_img)
            except Exception:
                pass

        insert_pos += block_rows

        # ── 品號之間插入空列（最後一個品號後不加）────────────
        if idx < len(cart_items) - 1:
            ws.insert_rows(insert_pos, amount=1)
            ws.row_dimensions[insert_pos].height = 18.0
            _apply_separator_row_border(ws, insert_pos)   # 外框連貫
            tax_row   += 1
            insert_pos += 1

    # ── 8. 運費列（在最後品號後、稅款前）────────────────────
    if shipping and shipping.get("enabled"):
        ship_price = shipping.get("price", 1500)
        ship_promo = shipping.get("promo", False)

        # 空白分隔列（同品號間格式）
        ws.insert_rows(insert_pos, amount=1)
        ws.row_dimensions[insert_pos].height = 18.0
        _apply_separator_row_border(ws, insert_pos)
        tax_row   += 1
        insert_pos += 1

        # 運費列
        ws.insert_rows(insert_pos, amount=1)
        ws.row_dimensions[insert_pos].height = 18.0
        ship_row = insert_pos
        ws.cell(row=ship_row, column=2, value="N/A")
        ws.cell(row=ship_row, column=3, value="一次性運費")
        ws.cell(row=ship_row, column=6, value=1)
        ws.cell(row=ship_row, column=7, value="趟")
        ws.cell(row=ship_row, column=8, value=ship_price)
        ws.cell(row=ship_row, column=9, value=f"=F{ship_row}*H{ship_row}")
        _apply_price_fmt(ws, ship_row)
        _style_product_row(ws, ship_row)
        _apply_product_row_border(ws, ship_row)
        for sc, ec in ((3, 5), (9, 10)):
            try:
                ws.merge_cells(start_row=ship_row, end_row=ship_row,
                               start_column=sc, end_column=ec)
            except Exception:
                pass
        tax_row   += 1
        insert_pos += 1

        # 免運優惠文字列（紅字）
        if ship_promo:
            ws.insert_rows(insert_pos, amount=1)
            ws.row_dimensions[insert_pos].height = 18.0
            promo_row  = insert_pos
            promo_cell = ws.cell(row=promo_row, column=8,
                                 value="(配合我司車趟享免運優惠)")
            promo_cell.font      = Font(name="微軟正黑體", size=9, color="FF0000")
            promo_cell.alignment = Alignment(horizontal="center", vertical="center")
            _apply_product_row_border(ws, promo_row)
            try:
                ws.merge_cells(start_row=promo_row, end_row=promo_row,
                               start_column=8, end_column=10)
            except Exception:
                pass
            tax_row   += 1
            insert_pos += 1

    # ── 9. 更新 footer 金額公式（依 quote_type 不同） ────────
    last_item_row = insert_pos - 1

    if quote_type == QUOTE_TYPE_ALLOWANCE:
        # 補助報價單：價格已含稅，直接加總，tax_row 就是「總應付金額」列
        tot_cell = ws.cell(row=tax_row, column=8,
                           value=f"=SUM(I{first_item_row}:I{last_item_row})")
        tot_cell.number_format = _NUM_FMT

    else:
        # 一般 / 對比：tax_row = 營業稅列，另找總金額列
        tax_cell = ws.cell(row=tax_row, column=8,
                           value=f"=ROUND(SUM(I{first_item_row}:I{last_item_row})*0.05,0)")
        tax_cell.number_format = _NUM_FMT

        # 對比用「總金額」；一般用「應收總金額」
        _total_kw = "總金額" if quote_type == QUOTE_TYPE_COMPARE else "應收總金額"
        total_row = None
        for r in ws.iter_rows(min_row=tax_row):
            for cell in r:
                if cell.value and _total_kw in str(cell.value):
                    total_row = cell.row
                    break
            if total_row:
                break
        if total_row:
            tot_cell = ws.cell(row=total_row, column=8,
                               value=f"=SUM(I{first_item_row}:I{last_item_row})+H{tax_row}")
            tot_cell.number_format = _NUM_FMT

    # ── 9. 重新設定 footer 合併範圍（依最終 tax_row 位置）──────
    for fm in footer_merges:
        new_min = tax_row + fm["row_offset"]
        new_max = tax_row + fm["max_row_offset"]
        ws.merge_cells(
            start_row=new_min, end_row=new_max,
            start_column=fm["min_col"], end_column=fm["max_col"],
        )

    # ── 10. 製表人 ───────────────────────────────────────────
    if operator:
        _fill_field(ws, _norm("製表人"), operator)

    # ── 10b. 印章（template 預埋圖片）重定位到 footer ──────────
    # 一般報價單：圖片完全不動
    # 補助報價單：找「備註」    → (i,   j+4)
    # 對比報價單：找「確認章：」→ (i-3, j+2)
    if _template_img_count > 0 and quote_type != QUOTE_TYPE_REGULAR:
        _stamp_r, _stamp_c = None, None
        if quote_type == QUOTE_TYPE_COMPARE:
            for _row in ws.iter_rows():
                for _cell in _row:
                    if isinstance(_cell, _MergedCell):
                        continue
                    if _cell.value and "確認章" in re.sub(r"\s", "", str(_cell.value)):
                        if _stamp_r is None or _cell.row > _stamp_r:
                            _stamp_r, _stamp_c = _cell.row, _cell.column
            if _stamp_r:
                _stamp_anchor = f"{_col_letter(_stamp_c + 2)}{_stamp_r - 3}"
        else:
            for _row in ws.iter_rows():
                for _cell in _row:
                    if isinstance(_cell, _MergedCell):
                        continue
                    if _cell.value and re.sub(r"\s", "", str(_cell.value)).startswith("備註"):
                        if _stamp_r is None or _cell.row > _stamp_r:
                            _stamp_r, _stamp_c = _cell.row, _cell.column
            if _stamp_r:
                _stamp_anchor = f"{_col_letter(_stamp_c + 4)}{_stamp_r}"
        if _stamp_r:
            _max_orig_row = max(_template_img_orig_rows) if _template_img_orig_rows else 0
            for _i, _ti in enumerate(ws._images[:_template_img_count]):
                if _template_img_orig_rows[_i] < _max_orig_row:
                    continue  # logo，位置不動
                _ti.anchor = _stamp_anchor
                _ti.width  = 178    # 4.7 cm
                _ti.height = 177    # 4.69 cm

    # ── 10d. 補回 (含稅) 紅色字——只讓 "(含稅)" 部分為紅色 ──
    for _row in ws.iter_rows():
        for _c in _row:
            if isinstance(_c, _MergedCell):
                continue
            _val = _c.value
            if not _val or "(含稅)" not in str(_val):
                continue
            _text = str(_val)
            _idx  = _text.find("(含稅)")
            _before = _text[:_idx]
            _after  = _text[_idx + len("(含稅)"):]
            # 保留原有字型屬性，只對 (含稅) 加紅色
            _f = _c.font
            _kw: dict = {}
            if _f.name:   _kw["rFont"] = _f.name
            if _f.sz:     _kw["sz"]    = _f.sz
            if _f.bold:   _kw["b"]     = _f.bold
            if _f.italic: _kw["i"]     = _f.italic
            _base = _InlineFont(**_kw)
            _red  = _InlineFont(**{**_kw, "color": "FF0000"})
            _parts: list = []
            if _before:
                _parts.append(_TextBlock(_base, _before))
            _parts.append(_TextBlock(_red, "(含稅)"))
            if _after:
                _parts.append(_TextBlock(_base, _after))
            _c.value = _CellRichText(*_parts)

    # ── 11. 成本建議售價表（K11 起） ─────────────────────────
    _cost_path = template_path.parent / "template_cost.xlsx"
    if _cost_path.exists():
        try:
            from openpyxl.styles import PatternFill as _PFill
            _cost_wb = openpyxl.load_workbook(str(_cost_path), data_only=True)

            def _cnorm(c):
                return re.sub(r'[-\s]', '', str(c or '')).upper()

            # 建立每個工作表的標題列與品號查找表
            _sh_hdr: dict[str, list] = {}
            _sh_lkp: dict[str, dict] = {}
            for _sn in _cost_wb.sheetnames:
                _sw = _cost_wb[_sn]
                _rows = list(_sw.iter_rows(min_row=1, max_col=5, values_only=True))
                if not _rows:
                    continue
                _sh_hdr[_sn] = list(_rows[0])
                _lkp = {}
                for _r in _rows[1:]:
                    if _r[0] is not None:
                        _lkp[_cnorm(str(_r[0]))] = list(_r)
                _sh_lkp[_sn] = _lkp
            _cost_wb.close()

            # 比對每個品項
            _matched: list[tuple[str, list]] = []
            for _it in cart_items:
                _nc = _cnorm(_it.get('code', ''))
                _cat = _it.get('category', '')
                _order = ([_cat] if _cat in _sh_lkp else []) + \
                         [s for s in _sh_lkp if s != _cat]
                for _sn in _order:
                    if _nc in _sh_lkp.get(_sn, {}):
                        _matched.append((_sn, _sh_lkp[_sn][_nc]))
                        break

            if _matched:
                _CROW, _CCOL = 11, 11   # K11
                _n_rows = 1 + len(_matched)   # 標題 + 資料列數
                _n_cols = 5

                _yellow   = _PFill(fill_type='solid', fgColor='FFFF00')
                _font_hdr = Font(name='Microsoft JhengHei', bold=True)
                _font_dat = Font(name='Microsoft JhengHei')
                _thick    = Side(style='medium')
                _thin     = Side(style='thin')
                _price_fmt = '"NT$"#,##0'

                # 哪些欄（0-indexed）的值是數字（建議售價等）
                # 第0欄=品號, 第1欄=台車型號, 第2-4欄=價格
                _price_cols = {2, 3, 4}

                for _ri in range(_n_rows):
                    _row_abs = _CROW + _ri
                    _is_hdr  = (_ri == 0)
                    _vals    = _sh_hdr.get(_matched[0][0], [])[:5] if _is_hdr \
                               else _matched[_ri - 1][1][:5]

                    for _j in range(_n_cols):
                        _c = ws.cell(row=_row_abs, column=_CCOL + _j)
                        _v = _vals[_j] if _j < len(_vals) else None

                        # 值
                        if _v is not None:
                            _c.value = _v

                        # 字型
                        _c.font = _font_hdr if _is_hdr else _font_dat

                        # 標題背景
                        if _is_hdr:
                            _c.fill = _yellow

                        # 數字格式
                        if not _is_hdr and _j in _price_cols and isinstance(_v, (int, float)):
                            _c.number_format = _price_fmt

                        # 框線
                        _top    = _thick if _ri == 0            else _thin
                        _bottom = _thick if _ri == _n_rows - 1  else _thin
                        _left   = _thick if _j == 0             else _thin
                        _right  = _thick if _j == _n_cols - 1   else _thin
                        _c.border = Border(top=_top, bottom=_bottom,
                                           left=_left, right=_right)
        except Exception:
            pass

    # ── 12. 儲存 ─────────────────────────────────────────────
    output_dir.mkdir(parents=True, exist_ok=True)
    company  = customer.get("company", "客戶")
    contact  = customer.get("contact", "")
    codes    = "+".join(item.get("code", "") for item in cart_items if item.get("code"))
    roc_date = f"{quote_date.year - 1911}{quote_date.month:02d}{quote_date.day:02d}"
    # 從卡片標題取括號（含前方空白），例如 " (台中大雅)" 或 "(高雄仁武)"
    _bm      = re.search(r'\s*\([^)]+\)', card_title) if card_title else None
    _bracket = _bm.group(0) if _bm else ""
    _raw     = f"報價單-{company}{_bracket} {contact} {codes}-{roc_date}"
    safe     = re.sub(r'[\\/:*?"<>|]', "_", _raw).strip()
    out_path = output_dir / f"{safe}.xlsx"
    wb.save(str(out_path))
    wb.close()
    if wb_ap:
        wb_ap.close()
    return out_path
