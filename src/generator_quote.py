"""
generator_quote.py — 從 Trello 卡片資料填入報價單範本，生成 .xlsx
"""
import copy as _copy
import json
import re
from datetime import date, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side


# ── 描述解析 ────────────────────────────────────────────────────


def _strip_md_link(text: str) -> str:
    """去除 Trello Markdown 連結格式，只保留顯示文字。
    [text](url) → text
    """
    return re.sub(r'\[([^\]]+)\]\([^)]*\)', r'\1', text).strip()


def parse_card_desc(desc: str) -> dict:
    """解析 TR 卡片描述（「欄位：值\n」格式）→ dict。
    同時支援全形 ：與半形 : 冒號。
    Trello 的 Markdown 連結格式（如 email 超連結）也會自動還原為純文字。
    """
    result: dict[str, str] = {}
    for line in desc.splitlines():
        # 全形優先；找不到再試半形
        colon = "：" if "：" in line else (":" if ":" in line else None)
        if colon is None:
            continue
        key, _, val = line.partition(colon)
        result[key.strip()] = _strip_md_link(val.strip())
    return result


# ── 範本工具 ────────────────────────────────────────────────────


def list_templates(template_dir: Path) -> list[Path]:
    """回傳 template_dir 內所有 報價單-*.xlsx，依檔名排序。"""
    return sorted(template_dir.glob("報價單-*.xlsx"))


def _norm(text: str) -> str:
    """正規化：移除空白、全形空格、冒號、連字號，轉大寫，方便模糊比對。
    ex: "E-MAIL" → "EMAIL"、"統一編號：" → "統一編號"
    """
    return re.sub(r"[\s：:　：\-]", "", str(text)).upper()


def _label_merge_end_col(ws, row: int, label_col: int) -> int:
    """回傳 label 所在合併儲存格的最右欄（無合併則回傳 label_col 本身）。"""
    for rng in ws.merged_cells.ranges:
        if (rng.min_row <= row <= rng.max_row
                and rng.min_col <= label_col <= rng.max_col):
            return rng.max_col
    return label_col


def _find_value_col(ws, row: int, label_col: int) -> int:
    """跳過 label 的合併範圍，找同列第一個可填入的值格。"""
    # 先確定 label merge 的右邊界，從其右一格開始找
    start = _label_merge_end_col(ws, row, label_col) + 1
    for c in range(start, start + 8):
        cell = ws.cell(row=row, column=c)
        val  = cell.value
        if val is None:
            return c
        # 只含冒號或空白也視為空
        if isinstance(val, str) and _norm(val) == "":
            return c
    return start


# label 關鍵字 → desc 欄位名（None 表示由呼叫端直接給值）
_LABEL_DESC_MAP: list[tuple[str, str | None]] = [
    ("客戶全名",  "公司名"),
    ("CUSTOMERNAME", "公司名"),   # 英文 label fallback
    ("電話",      None),          # 手機 or 電話，由呼叫端決定
    ("傳真",      "傳真"),
    ("聯絡人",    "聯絡人"),
    ("聯絡地址",  "地址"),
    ("統一編號",  "統一編號"),
    ("報價日期",  None),          # auto
    ("有效日期",  None),          # auto
    ("報價單號",  None),          # auto
]


def _fill_field(ws, label_kw_norm: str, value) -> bool:
    """掃描整張工作表，找到含 label_kw 的格後填入相鄰格。"""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            if label_kw_norm in _norm(str(cell.value)):
                target_col = _find_value_col(ws, cell.row, cell.column)
                out_cell = ws.cell(row=cell.row, column=target_col, value=value)
                # 強制純文字，防止 Excel 自動將 email / URL 轉成超連結
                if isinstance(value, str) and ("@" in value or value.startswith("http")):
                    out_cell.number_format = "@"
                return True
    return False


# ── 需求解析 ────────────────────────────────────────────────────


def _parse_demand_items(demand_str: str) -> list[tuple[str, int]]:
    """從需求字串中找出所有「品號*數量」格式，回傳 [(品號, 數量), ...]。

    支援多行或空白/逗號分隔：
        "AP-50B*1"          → [("AP-50B", 1)]
        "AP-50B*1 AP-30B*2" → [("AP-50B", 1), ("AP-30B", 2)]
    """
    items: list[tuple[str, int]] = []
    for token in re.split(r"[\s,，]+", demand_str):
        m = re.match(r"([A-Za-z0-9][A-Za-z0-9\-]*)\*(\d+)$", token.strip())
        if m:
            items.append((m.group(1), int(m.group(2))))
    return items


def _extract_demand_from_desc(raw_desc: str) -> str:
    """從原始描述字串中取出「需求:」或「需求：」之後的所有內容。

    卡片描述常見格式：
        需求:
        AP-50B*2
        AP-30B*1

    parse_card_desc 只抓同行 val，會漏掉換行後的品項，
    此函式直接從原始文字擷取 需求 標題之後的多行內容。
    """
    for keyword in ("需求：", "需求:"):
        idx = raw_desc.find(keyword)
        if idx == -1:
            continue
        after = raw_desc[idx + len(keyword):]
        # 同行有值就包含進來，後續行繼續收集直到遇到下一個「key:」行為止
        lines = []
        for line in after.splitlines():
            stripped = line.strip()
            # 遇到下一個欄位（含冒號的行）就停止
            if stripped and ("：" in stripped or ":" in stripped) and not re.match(
                r"[A-Za-z0-9][A-Za-z0-9\-]*\*\d+", stripped
            ):
                break
            lines.append(stripped)
        return " ".join(l for l in lines if l)
    return ""


# ── 報價單號 ────────────────────────────────────────────────────


def next_quote_no(output_dir: Path, operator_code: str, quote_date: date) -> str:
    """依輸出資料夾內當日已有的報價單數，產生下一組報價單號。

    格式：{代號}{民國年}{月:02d}{日:02d}{序號:03d}
    Ex: K114052100 1 → 「K1140521001」
    """
    roc_year = quote_date.year - 1911
    prefix   = f"{operator_code}{roc_year}{quote_date.month:02d}{quote_date.day:02d}"
    count    = 0
    if output_dir.exists():
        for f in output_dir.iterdir():
            if f.stem.startswith(f"報價單_") and prefix in f.stem:
                count += 1
    return f"{prefix}{count + 1:03d}"


# ── 主要生成函式 ────────────────────────────────────────────────


def generate_quote(
    card: dict,
    template_path: Path,
    output_dir: Path,
    quote_no: str,
    quote_date: date,
) -> Path:
    """將 Trello 卡片資料填入報價單範本並儲存。

    回傳輸出 .xlsx 的路徑。
    """
    desc_fields = parse_card_desc(card.get("desc", ""))
    valid_date  = quote_date + timedelta(days=15)

    def _get(*keys: str) -> str:
        """依序嘗試多個 key，回傳第一個非空值，否則回傳空字串。"""
        for k in keys:
            v = desc_fields.get(k)
            if v:
                return v
        return ""

    # 客戶全名：可能叫「公司名」或「公司名稱」
    company = _get("公司名", "公司名稱").strip()

    # 電話：優先手機，fallback 電話
    phone = _get("手機", "電話")

    fill_map: dict[str, object] = {
        "客戶全名":  company,
        "電話":      phone,
        "傳真":      _get("傳真"),
        "聯絡人":    _get("聯絡人"),
        "聯絡地址":  _get("地址", "聯絡地址"),
        # 統一編號：繁體/簡體「統」，或縮寫「统編」
        "統一編號":  _get("統一編號", "统一編號", "统編", "統編"),
        # EMAIL：可能叫「電子信箱」或直接寫「E-MAIL」
        "EMAIL":     _get("電子信箱", "E-MAIL", "EMAIL", "E-Mail"),
        "報價日期":  quote_date.strftime("%Y/%m/%d"),
        "有效日期":  valid_date.strftime("%Y/%m/%d"),
        "報價單號":  quote_no,
    }

    try:
        wb = openpyxl.load_workbook(str(template_path))
    except Exception as e:
        # 部分範本含有損壞的圖片/圖表引用（xl/drawings/NULL），
        # 改用 keep_links=False 再試一次
        if "xl/drawings" in str(e).lower() or "no item named" in str(e).lower():
            wb = openpyxl.load_workbook(str(template_path), keep_links=False)
        else:
            raise
    ws = wb.active

    for label_kw, value in fill_map.items():
        if value:
            _fill_field(ws, _norm(label_kw), value)

    # ── 需求明細：「品號*數量」→ 從第 15 列起填入 ──────────────
    # 需求項目可能換行寫在「需求:」下面，parse_card_desc 會漏掉，
    # 改用 _extract_demand_from_desc 直接從原始描述抓。
    demand_raw = _extract_demand_from_desc(card.get("desc", ""))
    if demand_raw:
        items = _parse_demand_items(demand_raw)
        for i, (_part_no, qty) in enumerate(items):
            row = 15 + i
            ws.cell(row=row, column=6, value=qty)                   # F = 數量
            ws.cell(row=row, column=9, value=f"=F{row}*H{row}")    # I = 小計

    output_dir.mkdir(parents=True, exist_ok=True)
    safe_company = re.sub(r'[\\/:*?"<>|]', "_", company).strip() or "客戶"
    out_path     = output_dir / f"報價單_{safe_company}.xlsx"
    wb.save(str(out_path))
    wb.close()
    return out_path


# ── 產品目錄 ────────────────────────────────────────────────────


def load_product_catalog(template_dir: Path) -> list[dict]:
    """從 template_dir/products.json 載入產品目錄，找不到回傳空清單。"""
    path = template_dir / "products.json"
    if not path.exists():
        return []
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return []


# ── template_AP 品號列範圍對照表 ────────────────────────────────
# 每個品號在 template_AP.xlsx 中對應的 (起始列, 結束列)

PRODUCT_AP_ROWS: dict[str, tuple[int, int]] = {
    "AP-15WB(經銷)": (2,  7),   # 簡化版，6 行
    "AP-15WB":       (9,  22),  # 完整版，14 行
    "AP-30B":        (24, 37),  # 14 行
    "AP-30SB":       (39, 51),  # 13 行
    "AP-50B":        (53, 65),  # 13 行
    "AP-50WB":       (67, 79),  # 13 行
    "AP-80B":        (80, 92),  # 13 行
}

_PRODUCT_START_ROW = 15   # template_quote 品項區起始列


# ── 品項列 border 樣式（從 template_quote R15 讀出的規格）────────
# 僅左右垂直線；厚外框、medium 內格線
_TK = Side(style="thick")
_MD = Side(style="medium")
_NO = Side(style=None)

# col → (left_side, right_side)
_PRODUCT_ROW_BORDERS: dict[int, tuple] = {
    1:  (_TK, _NO),
    2:  (_MD, _MD),
    3:  (_NO, _NO),
    4:  (_NO, _NO),
    5:  (_NO, _MD),
    6:  (_MD, _NO),
    7:  (_MD, _MD),
    8:  (_MD, _MD),
    9:  (_MD, _NO),
    10: (_NO, _TK),
}


def _apply_product_row_border(ws, row: int) -> None:
    """套用 template_quote 品項列的標準 border（左右垂直線）。"""
    for col, (lft, rgt) in _PRODUCT_ROW_BORDERS.items():
        ws.cell(row=row, column=col).border = Border(left=lft, right=rgt)


def _apply_separator_row_border(ws, row: int) -> None:
    """空列套完整欄位垂直線，讓上下品號之間的分欄線視覺連續。"""
    _apply_product_row_border(ws, row)


# ── 品項列字型與對齊 ─────────────────────────────────────────────
_JHENG_HEI = Font(name="微軟正黑體", size=11)
# wrap_text 不設 True：讓規格文字自然向右溢出空欄，避免窄欄折行
_ALIGN_CTR  = Alignment(horizontal="center", vertical="center")
_ALIGN_LEFT = Alignment(horizontal="left",   vertical="center")


def _style_product_row(ws, row: int) -> None:
    """套用品項列的字型（微軟正黑體）與對齊（品名左齊，其他置中）。"""
    for col in range(1, 11):
        cell = ws.cell(row=row, column=col)
        cell.font      = _JHENG_HEI
        cell.alignment = _ALIGN_LEFT if col in (3, 4, 5) else _ALIGN_CTR


# ── 麥當勞流程報價單生成 ────────────────────────────────────────


def generate_quote_from_cart(
    customer: dict,
    cart_items: list[dict],
    template_path: Path,
    output_dir: Path,
    quote_no: str,
    quote_date: date,
    operator: str = "",
) -> Path:
    """將購物車品項填入 template_quote.xlsx，生成完整報價單。

    品項資料從 template_AP.xlsx 對應列複製（A-H 欄），
    按品號動態插入，品號間保留一行空列，最後更新 營業稅 / 應收總金額 公式。

    customer 欄位：company, phone, fax, contact, address, tax_id, email
    cart_items 欄位：code, name, spec, unit, qty, price
    operator  ：製表人員姓名
    """
    valid_date = quote_date + timedelta(days=15)

    # ── 1. 開啟 template_quote ──────────────────────────────
    try:
        wb = openpyxl.load_workbook(str(template_path))
    except Exception as e:
        if "xl/drawings" in str(e).lower() or "no item named" in str(e).lower():
            wb = openpyxl.load_workbook(str(template_path), keep_links=False)
        else:
            raise
    ws = wb.active

    # ── 2. 開啟 template_AP（來源規格）──────────────────────
    ap_path = template_path.parent / "template_AP.xlsx"
    wb_ap, ws_ap = None, None
    if ap_path.exists():
        try:
            wb_ap = openpyxl.load_workbook(str(ap_path), keep_links=False)
            ws_ap = wb_ap.active
        except Exception:
            wb_ap, ws_ap = None, None

    # ── 3. 填入顧客資訊（欄位掃描填入）─────────────────────
    header_map = {
        "客戶全名":  customer.get("company", ""),
        "電話":      customer.get("phone",   ""),
        "傳真":      customer.get("fax",     ""),
        "聯絡人":    customer.get("contact", ""),
        "聯絡地址":  customer.get("address", ""),
        "統一編號":  customer.get("tax_id",  ""),
        "EMAIL":     customer.get("email",   ""),
        "報價日期":  quote_date.strftime("%Y/%m/%d"),
        "有效日期":  valid_date.strftime("%Y/%m/%d"),
        "報價單號":  quote_no,
    }
    for label_kw, value in header_map.items():
        if value:
            _fill_field(ws, _norm(label_kw), value)

    # ── 4. 找 營業稅 列（刪除 / 插入前先記下位置）───────────
    tax_row = None
    for r in ws.iter_rows():
        for cell in r:
            if cell.value and "營業稅" in str(cell.value):
                tax_row = cell.row
                break
        if tax_row:
            break
    if tax_row is None:
        tax_row = 17

    # ── 5. 取出 footer 合併範圍（R15+，之後重新定位）───────────
    #    openpyxl 的 delete_rows / insert_rows 不會自動更新 merged_cells，
    #    需手動移除再重新加入。
    footer_merges: list[dict] = []
    stale_ranges  = [rng for rng in ws.merged_cells.ranges
                     if rng.min_row >= _PRODUCT_START_ROW]
    for rng in stale_ranges:
        footer_merges.append({
            "row_offset":     rng.min_row - tax_row,
            "max_row_offset": rng.max_row - tax_row,
            "min_col":        rng.min_col,
            "max_col":        rng.max_col,
        })
        ws.unmerge_cells(str(rng))

    # ── 6. 刪除原有佔位品項列（R15 到 tax_row-1）────────────
    placeholder_count = tax_row - _PRODUCT_START_ROW
    if placeholder_count > 0:
        ws.delete_rows(_PRODUCT_START_ROW, placeholder_count)
        tax_row -= placeholder_count   # 營業稅列上移

    # ── 7. 逐品號插入資料列 ──────────────────────────────────
    insert_pos     = _PRODUCT_START_ROW
    first_item_row = _PRODUCT_START_ROW

    for idx, item in enumerate(cart_items):
        code  = item.get("code",  "")
        qty   = item.get("qty",   1)
        price = item.get("price", 0)

        ap_range  = PRODUCT_AP_ROWS.get(code) if ws_ap else None
        block_rows = (ap_range[1] - ap_range[0] + 1) if ap_range else 1

        # 插入空列（openpyxl 會自動將 tax_row 以下的列下移）
        ws.insert_rows(insert_pos, amount=block_rows)
        tax_row += block_rows

        # 設定列高
        for j in range(block_rows):
            ws.row_dimensions[insert_pos + j].height = 18.0

        if ap_range and ws_ap:
            # ── 從 template_AP 複製 A-H 欄值 ─────────────
            ap_start = ap_range[0]
            for j in range(block_rows):
                dst_row = insert_pos + j
                src_row = ap_start + j
                for col in range(1, 9):   # A(1) ~ H(8)
                    src_val = ws_ap.cell(row=src_row, column=col).value
                    if src_val is not None:
                        ws.cell(row=dst_row, column=col, value=src_val)
                # 字型、對齊、border
                _style_product_row(ws, dst_row)
                _apply_product_row_border(ws, dst_row)
        else:
            # fallback：找不到 template_AP 時只填基本欄位
            ws.cell(row=insert_pos, column=2, value=code)
            _style_product_row(ws, insert_pos)
            _apply_product_row_border(ws, insert_pos)

        # ── 覆寫序號、數量、單價、小計公式 ────────────────
        hrow = insert_pos                              # 品號 header 列
        ws.cell(row=hrow, column=1, value=idx + 1)    # A  序
        ws.cell(row=hrow, column=6, value=qty)        # F  數量
        ws.cell(row=hrow, column=8, value=price)      # H  單價
        ws.cell(row=hrow, column=9,                   # I  小計
                value=f"=F{hrow}*H{hrow}")

        # ── 品名欄（C:E）與小計欄（I:J）合併，對齊表頭寬度 ──
        for sc, ec in ((3, 5), (9, 10)):
            try:
                ws.merge_cells(
                    start_row=hrow, end_row=hrow,
                    start_column=sc, end_column=ec,
                )
            except Exception:
                pass

        insert_pos += block_rows

        # ── 品號之間插入空列（最後一個品號後不加）────────────
        if idx < len(cart_items) - 1:
            ws.insert_rows(insert_pos, amount=1)
            ws.row_dimensions[insert_pos].height = 18.0
            _apply_separator_row_border(ws, insert_pos)   # 外框連貫
            tax_row   += 1
            insert_pos += 1

    # ── 8. 更新 營業稅 / 應收總金額 公式 ────────────────────
    last_item_row = insert_pos - 1

    ws.cell(row=tax_row, column=8,
            value=f"=ROUND(SUM(I{first_item_row}:I{last_item_row})*0.05,0)")

    # 找 應收總金額 列
    total_row = None
    for r in ws.iter_rows(min_row=tax_row):
        for cell in r:
            if cell.value and "應收總金額" in str(cell.value):
                total_row = cell.row
                break
        if total_row:
            break
    if total_row:
        ws.cell(row=total_row, column=8,
                value=f"=SUM(I{first_item_row}:I{last_item_row})+H{tax_row}")

    # ── 9. 重新設定 footer 合併範圍（依最終 tax_row 位置）──────
    for fm in footer_merges:
        new_min = tax_row + fm["row_offset"]
        new_max = tax_row + fm["max_row_offset"]
        ws.merge_cells(
            start_row=new_min, end_row=new_max,
            start_column=fm["min_col"], end_column=fm["max_col"],
        )

    # ── 10. 製表人 ───────────────────────────────────────────
    if operator:
        _fill_field(ws, _norm("製表人"), operator)

    # ── 11. 儲存 ─────────────────────────────────────────────
    output_dir.mkdir(parents=True, exist_ok=True)
    company  = customer.get("company", "客戶")
    safe     = re.sub(r'[\\/:*?"<>|]', "_", company).strip() or "客戶"
    out_path = output_dir / f"報價單_{safe}.xlsx"
    wb.save(str(out_path))
    wb.close()
    if wb_ap:
        wb_ap.close()
    return out_path
