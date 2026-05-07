"""
generator_inspection.py — 報價單 → 驗機單 轉換 v5

策略（大幅簡化）：
  1. 直接用 openpyxl 讀寫，不動 XML / zipfile
  2. 清除 J 欄（col>=10）所有儲存格
  3. 標題「報價單」→「驗機單」
  4. 清除指定欄位值（電話、傳真…）
  5. 品項單價/小計、營業稅、應收總金額歸零
  6. A1 插入公司 Logo 圖片
  7. 找「訂購請簽章回傳」，將其下方 4 格改成粗下框線
"""

import re, shutil
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage

OUTPUT_DIR   = Path(__file__).parent.parent / "output"
LOGO_PATH    = Path(__file__).parent.parent / "template" / "logo.png"
TITLE_RE     = re.compile(r'報[\s\u3000]*價[\s\u3000]*單')
CLEAR_KW     = {"電話", "傳真", "聯絡人", "統一編號", "聯絡地址", "EMAIL"}
ZERO_KW      = {"營業稅", "應收總金額", "合計"}


def _norm(s) -> str:
    return re.sub(r'[\s\u3000\-：:]+', '', str(s or '')).upper()


def _strip_ws(s) -> str:
    return re.sub(r'[\s\u3000]+', '', str(s or ''))


def _zero(v):
    """數字直接歸零；NT$x,xxx 字串改為 NT$0；其他原樣"""
    if isinstance(v, (int, float)):
        return 0
    if isinstance(v, str):
        if re.sub(r'[NT$,\s]', '', v).replace('.', '', 1).isdigit():
            return "NT$0"
    return v


def generate_inspection(src_path: str, data: dict) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    customer = data["header"].get("customer", "客戶")
    out_path = OUTPUT_DIR / f"驗機單-{customer}.xlsx"
    shutil.copy(src_path, out_path)

    wb = load_workbook(out_path)
    ws = wb.active

    # ── ① 清除 K 欄（col >= 11）所有儲存格值與格式 ──────────────
    from openpyxl.styles import PatternFill, Border, Font, Alignment
    _no_fill   = PatternFill(fill_type=None)
    _no_border = Border()
    _def_font  = Font()
    _def_align = Alignment()

    for row in ws.iter_rows():
        for cell in row:
            if cell.column >= 11 and not isinstance(cell, MergedCell):
                cell.value      = None
                cell.fill       = _no_fill
                cell.border     = _no_border
                cell.font       = _def_font
                cell.alignment  = _def_align

    # 移除 K 欄以後的圖片
    def _img_col(img) -> int:
        a = img.anchor
        if isinstance(a, str):
            m = re.match(r'([A-Za-z]+)', a)
            if m:
                col = 0
                for ch in m.group(1).upper():
                    col = col * 26 + (ord(ch) - 64)
                return col
        else:
            try:
                return img.anchor._from.col + 1   # 0-based → 1-based
            except Exception:
                pass
        return 0

    ws._images = [img for img in ws._images if _img_col(img) < 11]

    # ── ② 找標題列並替換 ────────────────────────────────────────
    title_row = -1
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell) or not cell.value:
                continue
            if _strip_ws(str(cell.value)) == "報價單":
                cell.value = TITLE_RE.sub('驗 機 單', str(cell.value))
                title_row = cell.row
            elif "報價單號" in str(cell.value):
                cell.value = str(cell.value).replace("報價單號", "驗機單號")

    # ── ③ 清除指定欄位的值（標籤保留）──────────────────────────
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell) or not cell.value:
                continue
            if any(kw in _norm(str(cell.value)) for kw in CLEAR_KW):
                for c in range(cell.column + 1, cell.column + 8):
                    t = ws.cell(row=cell.row, column=c)
                    if not isinstance(t, MergedCell) and t.value:
                        t.value = None
                        break
    
    # ── ④ 動態找「單價」/「小計」欄，全欄歸零；合計類關鍵字歸零 ──
    def _is_numeric(v):
        if isinstance(v, (int, float)):
            return True
        if isinstance(v, str):
            stripped = re.sub(r'[NT$,\s]', '', v).replace('.', '', 1).lstrip('-')
            return bool(stripped) and stripped.isdigit()
        return False

    def _should_clear(v):
        """數值或公式都要清除"""
        if v is None:
            return False
        if isinstance(v, (int, float)):
            return True
        if isinstance(v, str):
            if v.startswith('='):
                return True
            return _is_numeric(v)
        return False

    # 掃描表頭，動態定位「單價」「小計」欄號
    price_col    = 8   # 預設 fallback
    subtotal_col = 9
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell) or not cell.value:
                continue
            text = _norm(str(cell.value))
            if text == "單價" and price_col == 8:
                price_col = cell.column
            elif text in ("小計", "金額") and subtotal_col == 9:
                subtotal_col = cell.column
        if price_col != 8 or subtotal_col != 9:
            break   # 找到後就停

    for row in ws.iter_rows():
        rn = row[0].row

        # 單價欄 → 清空
        cp = ws.cell(row=rn, column=price_col)
        if not isinstance(cp, MergedCell) and _should_clear(cp.value):
            cp.value = None

        # 小計欄 → 歸 0
        cs = ws.cell(row=rn, column=subtotal_col)
        if not isinstance(cs, MergedCell) and _should_clear(cs.value):
            cs.value = 0

        # 掃描此列：找合計類關鍵字，將同列所有數值/公式歸零
        for cell in row:
            if isinstance(cell, MergedCell) or not cell.value:
                continue
            if any(kw in _norm(str(cell.value)) for kw in ZERO_KW):
                for c in range(2, 11):
                    t = ws.cell(row=rn, column=c)
                    if isinstance(t, MergedCell) or t.value is None:
                        continue
                    new_v = _zero(t.value)
                    if new_v != t.value:
                        t.value = new_v

    # ── ⑤ 插入 Logo 至 A1 ────────────────────────────────────
    if LOGO_PATH.exists():
        img = XLImage(str(LOGO_PATH))
        img.anchor = "A1"
        ws.add_image(img)

    # ── ⑥ 找「訂購請簽章回傳」，下方 4 格改成粗下框線 ────────────────
    from openpyxl.styles import Side
    sig_row = -1
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell) or not cell.value:
                continue
            if "訂購" in str(cell.value) and "簽章" in str(cell.value):
                sig_row = cell.row
                sig_col = cell.column
                break
        if sig_row > 0:
            break

    if sig_row > 0:
        # 在下方 4 格（row + 4）設定細下框線
        sig_cell = ws.cell(row=sig_row + 4, column=sig_col)
        if not isinstance(sig_cell, MergedCell):
            sig_cell.border = Border(bottom=Side(style='thin'))

    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    import sys
    sys.path.insert(0, str(Path(__file__).parent))
    from parser import parse

    if len(sys.argv) < 2:
        print("用法：python generator_inspection.py 報價單.xlsx")
        sys.exit(1)

    src  = sys.argv[1]
    if not Path(src).exists():
        print(f"找不到檔案：{src}")
        sys.exit(1)

    data = parse(src)
    out  = generate_inspection(src, data)
    print(f"驗機單已輸出：{out}")
