"""
generator_schedule.py — 從 Timetree 行事曆抓取出貨排程，生成 Excel 工作表
"""
import re
from datetime import datetime, timezone, timedelta, date as date_type
from pathlib import Path

import requests
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

_SCHEDULE_FILE = Path(r"Z:\會計\5.出貨相關\出貨行程表.xlsx")

_RED    = "FFFF0000"
_THIN   = Side(style="thin")
_THICK  = Side(style="thick")
_ALL_BORDER   = Border(left=_THIN,  right=_THIN,  top=_THIN,  bottom=_THIN)
_THICK_BORDER = Border(left=_THICK, right=_THICK, top=_THICK, bottom=_THICK)
_CENTER = Alignment(horizontal="center", vertical="center")

CALENDAR_ID = 25024642
_API_URL    = f"https://timetreeapp.com/api/v1/calendar/{CALENDAR_ID}/events"
_TW_TZ      = timezone(timedelta(hours=8))


def _extract_seq(title: str) -> float:
    """Return leading (N) sequence number, or inf if absent."""
    t = re.sub(r'【[^】]*】', '', title).strip()
    m = re.match(r'^\((\d+)\)', t)
    return int(m.group(1)) if m else float('inf')


def _split_prefixes(title: str) -> tuple[str, str]:
    """Split off leading parenthesised groups from title (after stripping 【】).

    Numeric groups like (1)(2) are skipped silently (used only for ordering).
    Text groups like (拉回)(拉回+送回) are collected and returned as the first element.
    Returns (text_prefix, remaining_title).
    """
    t = re.sub(r'【[^】]*】', '', title).strip()
    text_parts: list[str] = []
    while True:
        m = re.match(r'^\(\d+\)\s*', t)
        if m:
            t = t[m.end():]
            continue
        m = re.match(r'^(\([^)]+\))\s*', t)
        if m:
            text_parts.append(m.group(1))
            t = t[m.end():]
        else:
            break
    return "".join(text_parts), t


def _extract_company(title: str) -> str:
    """Return company name, stripping all leading (N)/(text) prefixes."""
    _, rest = _split_prefixes(title)
    m = re.match(r'^(.+?)(?=[\(（])', rest)  # match both half-width ( and full-width （
    if m:
        return m.group(1).strip(' -')
    return rest.split()[0] if rest else title


def _location_cell(ev: dict) -> str:
    company = _extract_company(ev.get("title", ""))
    loc = (ev.get("location") or "").strip()
    if loc:
        return f"{company}({loc})"
    _, rest = _split_prefixes(ev.get("title", ""))
    m = re.search(r'[\(（]([^)）]+)[)）]', rest)  # half-width or full-width brackets
    area = m.group(1) if m else ""
    return f"{company}({area})" if area else company


def _note_suffix(ev: dict) -> str:
    """Return note suffix: text-prefix (e.g. (拉回)) prepended, then person/product info."""
    company = _extract_company(ev.get("title", ""))
    text_prefix, rest = _split_prefixes(ev.get("title", ""))
    rest = re.sub(r'^' + re.escape(company) + r'\s*(?:[\(（][^)）]*[)）])?\s*', '', rest).strip()
    rest = re.sub(r'^[-\s]+', '', rest).strip()
    return f"{text_prefix} {rest}".strip() if text_prefix else rest


def _note_cell(seq: int, ev: dict) -> str:
    company = _extract_company(ev.get("title", ""))
    suffix = _note_suffix(ev)
    return f"{seq}.{company}-{suffix}" if suffix else f"{seq}.{company}"


def fetch_events(target: date_type, session_id: str, csrf_token: str) -> list:
    """Return active Timetree events on target date (Taiwan time)."""
    start_tw = datetime(target.year, target.month, target.day, tzinfo=_TW_TZ)
    since_ms = int((start_tw - timedelta(days=60)).timestamp() * 1000)

    headers = {
        "accept": "application/json, text/plain, */*",
        "content-type": "application/json",
        "referer": "https://timetreeapp.com/",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "x-csrf-token": csrf_token,
        "x-timetreea": "web/2.1.0/zh-TW",
        "cookie": f"_session_id={session_id}",
    }

    resp = requests.get(_API_URL, params={"since": since_ms}, headers=headers, timeout=15)
    resp.raise_for_status()

    result = []
    for ev in resp.json().get("events", []):
        if ev.get("deactivated_at"):
            continue
        dt = datetime.fromtimestamp(ev.get("start_at", 0) / 1000, tz=_TW_TZ)
        if dt.date() == target:
            result.append({**ev, "_dt": dt})

    # numbered (N) events first in numeric order, then all-day, then by start time
    result.sort(key=lambda e: (_extract_seq(e.get("title", "")), not e.get("all_day"), e["_dt"]))
    return result


def events_to_rows(events: list) -> list:
    """Convert event list to row dicts for GUI preview/editing.

    Each row: {"ev": dict, "location": str, "note_suffix": str, "travel_time": str}
    """
    return [{"ev": ev, "location": _location_cell(ev),
             "note_suffix": _note_suffix(ev), "travel_time": ""}
            for ev in events]


def _maps_query(row: dict) -> str:
    """Extract best address string for Google Maps from a row."""
    loc = (row["ev"].get("location") or "").strip()
    if loc:
        return loc
    m = re.search(r'\(([^)]+)\)', row["location"])
    return m.group(1) if m else row["location"]


def sort_rows_by_location(rows: list, api_key: str, south_to_north: bool = True) -> tuple[list, list]:
    """Sort rows by latitude via Google Maps Geocoding API.

    Returns (sorted_rows, failed_list). Failed rows are appended at the end unchanged.
    """
    _GEO_URL = "https://maps.googleapis.com/maps/api/geocode/json"
    geocoded: list[tuple[float, dict]] = []
    failed: list[dict] = []

    for row in rows:
        query = _maps_query(row)
        resp = requests.get(_GEO_URL, params={"address": query, "key": api_key}, timeout=15)
        resp.raise_for_status()
        data = resp.json()
        if data.get("status") == "OK" and data.get("results"):
            lat = data["results"][0]["geometry"]["location"]["lat"]
            geocoded.append((lat, row))
        else:
            failed.append(row)

    geocoded.sort(key=lambda x: x[0], reverse=not south_to_north)
    return [r for _, r in geocoded] + failed, failed


def _format_duration(seconds: int) -> str:
    minutes = round(seconds / 60)
    if minutes < 60:
        return f"{minutes}分"
    h, m = divmod(minutes, 60)
    return f"{h}時{m}分" if m else f"{h}時"


def calculate_travel_times(rows: list, api_key: str, origin: str) -> tuple:
    """Call Google Maps Distance Matrix API for each stop and update travel_time in-place.

    Returns (rows, failed_list) where failed_list contains (index, location, api_status).
    """
    _DM_URL = "https://maps.googleapis.com/maps/api/distancematrix/json"
    failed = []
    prev = origin
    for i, row in enumerate(rows):
        dest = _maps_query(row)
        resp = requests.get(_DM_URL, params={
            "origins":      prev,
            "destinations": dest,
            "mode":         "driving",
            "language":     "zh-TW",
            "key":          api_key,
        }, timeout=15)
        resp.raise_for_status()
        data = resp.json()
        top = data.get("status", "UNKNOWN")
        if top != "OK":
            row["travel_time"] = ""
            failed.append((i + 1, dest, f"API錯誤:{top}"))
            prev = dest
            continue
        try:
            el = data["rows"][0]["elements"][0]
            if el["status"] == "OK":
                row["travel_time"] = _format_duration(el["duration"]["value"])
            else:
                row["travel_time"] = ""
                failed.append((i + 1, dest, el["status"]))
        except (IndexError, KeyError):
            row["travel_time"] = ""
            failed.append((i + 1, dest, "解析失敗"))
        prev = dest
    return rows, failed


def generate_schedule(target: date_type, session_id: str, csrf_token: str,
                      rows: list | None = None) -> Path:
    """Write a new sheet into _SCHEDULE_FILE for target date.

    rows: optional pre-processed list from events_to_rows (may be edited by user).
          If None, fetch from Timetree and derive automatically.
    """
    if rows is None:
        rows = events_to_rows(fetch_events(target, session_id, csrf_token))

    wb = openpyxl.load_workbook(str(_SCHEDULE_FILE))

    roc_year = target.year - 1911
    sheet_name = f"{roc_year}年{target.month}月{target.day}日"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    def _merge_ab_cf(r, center_c=False):
        ws.merge_cells(f"A{r}:B{r}")
        ws.cell(row=r, column=1).alignment = _CENTER
        ws.merge_cells(f"C{r}:F{r}")
        if center_c:
            ws.cell(row=r, column=3).alignment = _CENTER

    def _merge_gh_border(r):
        ws.merge_cells(f"G{r}:H{r}")
        for col in (7, 8):
            ws.cell(row=r, column=col).border = _THICK_BORDER

    def _table_border(start_row, end_row):
        for r in range(start_row, end_row + 1):
            for col in range(1, 9):
                ws.cell(row=r, column=col).border = _ALL_BORDER

    # Column widths
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 11.5
    ws.column_dimensions["H"].width = 40

    # ── 上方事件表格 ─────────────────────────────────────────
    table_end = 1 + len(rows)

    ws["A1"] = f"{target.month:02d}/{target.day:02d}行程順序"
    ws["C1"] = "地點"
    ws.cell(row=1, column=7, value="Google時間").alignment = _CENTER
    ws.cell(row=1, column=8, value="備註").alignment = _CENTER
    _merge_ab_cf(1, center_c=True)

    for i, row in enumerate(rows, 1):
        r = i + 1
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=3, value=row["location"])
        if row.get("travel_time"):
            ws.cell(row=r, column=7, value=row["travel_time"]).alignment = _CENTER
        _merge_ab_cf(r)

    _table_border(1, table_end)

    # ── 下方備註列（每筆之間空一列）────────────────────────────
    note_start = table_end + 2
    for i, row in enumerate(rows, 1):
        r = note_start + (i - 1) * 2
        company = row["location"].split("(")[0].strip()
        suffix = row["note_suffix"]
        note_text = f"{i}.{company}-{suffix}" if suffix else f"{i}.{company}"
        ws.cell(row=r, column=1, value=note_text)
        if "拉回" in row["ev"].get("title", ""):
            continue
        ws.merge_cells(f"A{r}:F{r}")
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = _THICK_BORDER
        ws.cell(row=r, column=7, value="物流士確認          □")
        _merge_gh_border(r)

    # ── 頁尾 ─────────────────────────────────────────────────
    footer_r = note_start + len(rows) * 2

    ws.cell(row=footer_r, column=7, value="要給司機零用金+手機").font = Font(color=_RED)

    ws.cell(row=footer_r + 1, column=7, value="物流士確認")
    _merge_gh_border(footer_r + 1)

    ws.cell(row=footer_r + 2, column=7).value = CellRichText(
        TextBlock(InlineFont(color=_RED), "物流士確認          主管覆核(蓋章) "),
        TextBlock(InlineFont(color="FF000000"), "□"),
    )
    _merge_gh_border(footer_r + 2)

    wb.save(str(_SCHEDULE_FILE))
    wb.close()
    return _SCHEDULE_FILE
